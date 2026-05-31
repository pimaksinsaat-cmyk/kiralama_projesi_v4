"""
Read-only Firma.bakiye / defter bakiye uyusmazlik raporu.

Bu script yalnizca raporlama yapar; veritabaninda hicbir kaydi guncellemez.

Kullanim:
  docker exec kiralama_web python scripts/audit_firma_bakiye_defter_mismatch.py
  docker exec kiralama_web python scripts/audit_firma_bakiye_defter_mismatch.py --xlsx /tmp/firma_bakiye_raporu.xlsx
  docker exec kiralama_web python scripts/audit_firma_bakiye_defter_mismatch.py --csv /tmp/firma_bakiye_raporu.csv
  docker exec kiralama_web python scripts/audit_firma_bakiye_defter_mismatch.py --tolerance 0.01 --today 2026-05-26

Not:
  Proje Docker Compose icinde calisirken DATABASE_URL host'u "db" olur.
  Windows terminalinden dogrudan "python scripts/..." calistirmak bu host'u
  cozemeyebilir; bu durumda yukaridaki docker exec komutunu kullanin.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Iterable, Optional

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from app import create_app
from app.cari.models import HizmetKaydi, Odeme
from app.extensions import db
from app.firmalar.models import Firma
from app.kiralama.models import KiralamaKalemi
from app.services.kiralama_services import KiralamaService, to_decimal
from sqlalchemy.exc import OperationalError

DAHILI_FIRMA_ADLARI = ("DAHILI ISLEMLER", "DAHİLİ İŞLEMLER", "Dahili Kasa İşlemleri")


@dataclass
class FirmaBakiyeMismatch:
    firma_id: int
    firma_adi: str
    firma_bakiye_cache: Decimal
    defter_net_bakiye: Decimal
    fark: Decimal
    defter_net_bakiye_kdvli: Decimal
    cari_bakiye_kdvli_cache: Decimal
    cari_cache_fark: Decimal
    hizmet_kaydi_sayisi: int
    odeme_sayisi: int
    sebep_tahmini: str

    def to_row(self) -> dict[str, str]:
        return {
            "firma_id": str(self.firma_id),
            "firma_adi": self.firma_adi,
            "firma_bakiye_cache": str(self.firma_bakiye_cache),
            "defter_net_bakiye": str(self.defter_net_bakiye),
            "fark": str(self.fark),
            "defter_net_bakiye_kdvli": str(self.defter_net_bakiye_kdvli),
            "cari_bakiye_kdvli_cache": str(self.cari_bakiye_kdvli_cache),
            "cari_cache_fark": str(self.cari_cache_fark),
            "hizmet_kaydi_sayisi": str(self.hizmet_kaydi_sayisi),
            "odeme_sayisi": str(self.odeme_sayisi),
            "sebep_tahmini": self.sebep_tahmini,
        }


@dataclass
class DisKiralamaMismatch:
    kalem_id: int
    firma_id: int
    form_no: str
    beklenen_tutar: Decimal
    mevcut_hizmet_tutari: Decimal
    hizmet_id: int
    fark: Decimal


@dataclass
class AuditReport:
    firma_mismatches: list[FirmaBakiyeMismatch]
    dis_kiralama_mismatches: list[DisKiralamaMismatch]
    checked_firma_count: int
    checked_dis_kiralama_count: int


def _abs_diff(left: Decimal, right: Decimal) -> Decimal:
    return abs(to_decimal(left) - to_decimal(right))


def _exceeds_tolerance(left: Decimal, right: Decimal, tolerance: Decimal) -> bool:
    return _abs_diff(left, right) > tolerance


def _classify_mismatch(
    firma_bakiye_cache: Decimal,
    defter_net_bakiye: Decimal,
    cari_bakiye_kdvli_cache: Decimal,
    defter_net_bakiye_kdvli: Decimal,
    tolerance: Decimal,
) -> str:
    reasons: list[str] = []

    if defter_net_bakiye == 0 and firma_bakiye_cache != 0:
        reasons.append("DEFTER_ZERO_CACHE_STALE")
    elif firma_bakiye_cache == 0 and defter_net_bakiye != 0:
        reasons.append("FIRMA_BAKIYE_ZERO_CACHE")
    else:
        reasons.append("FIRMA_BAKIYE_PARTIAL_STALE")

    if _abs_diff(cari_bakiye_kdvli_cache, defter_net_bakiye_kdvli) <= tolerance:
        reasons.append("CARI_CACHE_OK")
    elif cari_bakiye_kdvli_cache == 0 and defter_net_bakiye_kdvli != 0:
        reasons.append("CARI_CACHE_ZERO")
    else:
        reasons.append("CARI_CACHE_STALE")

    return ";".join(reasons)


def audit_firma_bakiye_mismatches(tolerance: Decimal | float | str = Decimal("0.01")) -> tuple[list[FirmaBakiyeMismatch], int]:
    tolerance = to_decimal(tolerance)
    firmalar = (
        Firma.query.filter(
            Firma.is_deleted == False,
            Firma.firma_adi.notin_(DAHILI_FIRMA_ADLARI),
        )
        .order_by(Firma.id.asc())
        .all()
    )

    rows: list[FirmaBakiyeMismatch] = []
    for firma in firmalar:
        ozet = firma.bakiye_ozeti
        defter_net = to_decimal(ozet.get("net_bakiye"))
        defter_net_kdvli = to_decimal(ozet.get("net_bakiye_kdvli"))
        firma_cache = to_decimal(firma.bakiye)
        cari_cache = to_decimal(firma.cari_bakiye_kdvli)

        if not _exceeds_tolerance(firma_cache, defter_net, tolerance):
            continue

        hizmet_sayisi = HizmetKaydi.query.filter(
            HizmetKaydi.firma_id == firma.id,
            HizmetKaydi.is_deleted == False,
        ).count()
        odeme_sayisi = Odeme.query.filter(
            Odeme.firma_musteri_id == firma.id,
            Odeme.is_deleted == False,
        ).count()

        rows.append(
            FirmaBakiyeMismatch(
                firma_id=firma.id,
                firma_adi=firma.firma_adi,
                firma_bakiye_cache=firma_cache,
                defter_net_bakiye=defter_net,
                fark=defter_net - firma_cache,
                defter_net_bakiye_kdvli=defter_net_kdvli,
                cari_bakiye_kdvli_cache=cari_cache,
                cari_cache_fark=defter_net_kdvli - cari_cache,
                hizmet_kaydi_sayisi=hizmet_sayisi,
                odeme_sayisi=odeme_sayisi,
                sebep_tahmini=_classify_mismatch(
                    firma_cache,
                    defter_net,
                    cari_cache,
                    defter_net_kdvli,
                    tolerance,
                ),
            )
        )

    return rows, len(firmalar)


def _dis_kiralama_hizmetleri(kalem: KiralamaKalemi) -> list[HizmetKaydi]:
    return (
        HizmetKaydi.query.filter(
            HizmetKaydi.ozel_id == kalem.id,
            HizmetKaydi.firma_id == kalem.harici_ekipman_tedarikci_id,
            HizmetKaydi.yon == "gelen",
            HizmetKaydi.is_deleted == False,
            db.or_(
                HizmetKaydi.aciklama.like("Dış Kiralama%"),
                HizmetKaydi.aciklama.like("Dis Kiralama%"),
            ),
        )
        .order_by(HizmetKaydi.id.asc())
        .all()
    )


def audit_dis_kiralama_amount_mismatches(
    today: Optional[date] = None,
    tolerance: Decimal | float | str = Decimal("0.01"),
) -> tuple[list[DisKiralamaMismatch], int]:
    if today is None:
        today = date.today()
    tolerance = to_decimal(tolerance)

    kalemler = (
        KiralamaKalemi.query.filter(
            KiralamaKalemi.is_dis_tedarik_ekipman == True,
            KiralamaKalemi.harici_ekipman_tedarikci_id.isnot(None),
            KiralamaKalemi.harici_ekipman_tedarikci_id > 0,
            KiralamaKalemi.is_deleted == False,
        )
        .order_by(KiralamaKalemi.id.asc())
        .all()
    )

    rows: list[DisKiralamaMismatch] = []
    for kalem in kalemler:
        kiralama = kalem.kiralama
        if not kiralama or kiralama.is_deleted:
            continue

        hizmetler = _dis_kiralama_hizmetleri(kalem)
        if not hizmetler:
            continue

        hizmet = hizmetler[0]
        beklenen = to_decimal(KiralamaService._hesapla_bekleyen_alis_kalem_tutari(kalem, referans_tarih=today))
        mevcut = to_decimal(hizmet.tutar)
        if not _exceeds_tolerance(mevcut, beklenen, tolerance):
            continue

        rows.append(
            DisKiralamaMismatch(
                kalem_id=kalem.id,
                firma_id=kalem.harici_ekipman_tedarikci_id,
                form_no=kiralama.kiralama_form_no or "-",
                beklenen_tutar=beklenen,
                mevcut_hizmet_tutari=mevcut,
                hizmet_id=hizmet.id,
                fark=beklenen - mevcut,
            )
        )

    return rows, len(kalemler)


def run_audit(
    today: Optional[date] = None,
    tolerance: Decimal | float | str = Decimal("0.01"),
) -> AuditReport:
    tolerance = to_decimal(tolerance)
    firma_rows, firma_count = audit_firma_bakiye_mismatches(tolerance=tolerance)
    dis_rows, dis_count = audit_dis_kiralama_amount_mismatches(today=today, tolerance=tolerance)
    return AuditReport(
        firma_mismatches=firma_rows,
        dis_kiralama_mismatches=dis_rows,
        checked_firma_count=firma_count,
        checked_dis_kiralama_count=dis_count,
    )


def write_csv(rows: Iterable[FirmaBakiyeMismatch], path: str) -> None:
    fieldnames = [
        "firma_id",
        "firma_adi",
        "firma_bakiye_cache",
        "defter_net_bakiye",
        "fark",
        "defter_net_bakiye_kdvli",
        "cari_bakiye_kdvli_cache",
        "cari_cache_fark",
        "hizmet_kaydi_sayisi",
        "odeme_sayisi",
        "sebep_tahmini",
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_row())


def _autosize_columns(ws, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _style_header(row) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    side = Side(style="thin", color="D9E2F3")
    border = Border(bottom=side)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _format_money_columns(ws, columns: Iterable[str]) -> None:
    for col in columns:
        for cell in ws[col][1:]:
            cell.number_format = '#,##0.00'


def write_xlsx(report: AuditReport, path: str) -> None:
    """Biçimli Excel raporu üretir. Veritabanına yazmaz."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Ozet"
    ws_firma = wb.create_sheet("Firma Uyusmazliklari")
    ws_dis = wb.create_sheet("Dis Kiralama")

    ws_summary["A1"] = "Firma Bakiye / Defter Uyusmazlik Raporu"
    ws_summary["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws_summary["A2"] = "READ-ONLY - veritabaninda hicbir kayit guncellenmez."
    ws_summary["A2"].font = Font(italic=True, color="666666")
    summary_rows = [
        ("Kontrol edilen firma", report.checked_firma_count),
        ("Firma.bakiye drift", len(report.firma_mismatches)),
        ("Kontrol edilen dis kiralama kalemi", report.checked_dis_kiralama_count),
        ("Dis kiralama tutar uyusmazligi", len(report.dis_kiralama_mismatches)),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=4):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)
    for row_idx in range(4, 8):
        ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor="EAF2F8")
    _autosize_columns(ws_summary)

    firma_headers = [
        "firma_id",
        "firma_adi",
        "firma_bakiye_cache",
        "defter_net_bakiye",
        "fark",
        "defter_net_bakiye_kdvli",
        "cari_bakiye_kdvli_cache",
        "cari_cache_fark",
        "hizmet_kaydi_sayisi",
        "odeme_sayisi",
        "sebep_tahmini",
    ]
    ws_firma.append(firma_headers)
    for row in report.firma_mismatches:
        ws_firma.append([
            row.firma_id,
            row.firma_adi,
            float(row.firma_bakiye_cache),
            float(row.defter_net_bakiye),
            float(row.fark),
            float(row.defter_net_bakiye_kdvli),
            float(row.cari_bakiye_kdvli_cache),
            float(row.cari_cache_fark),
            row.hizmet_kaydi_sayisi,
            row.odeme_sayisi,
            row.sebep_tahmini,
        ])
    _style_header(ws_firma[1])
    ws_firma.freeze_panes = "A2"
    ws_firma.auto_filter.ref = ws_firma.dimensions
    _format_money_columns(ws_firma, ["C", "D", "E", "F", "G", "H"])
    for cell in ws_firma["K"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    if ws_firma.max_row > 1:
        table = Table(displayName="FirmaUyusmazliklari", ref=f"A1:K{ws_firma.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws_firma.add_table(table)
    _autosize_columns(ws_firma, max_width=55)

    dis_headers = ["kalem_id", "firma_id", "form_no", "beklenen_tutar", "mevcut_hizmet_tutari", "hizmet_id", "fark"]
    ws_dis.append(dis_headers)
    for row in report.dis_kiralama_mismatches:
        ws_dis.append([
            row.kalem_id,
            row.firma_id,
            row.form_no,
            float(row.beklenen_tutar),
            float(row.mevcut_hizmet_tutari),
            row.hizmet_id,
            float(row.fark),
        ])
    _style_header(ws_dis[1])
    ws_dis.freeze_panes = "A2"
    ws_dis.auto_filter.ref = ws_dis.dimensions
    _format_money_columns(ws_dis, ["D", "E", "G"])
    if ws_dis.max_row > 1:
        table = Table(displayName="DisKiralamaUyusmazliklari", ref=f"A1:G{ws_dis.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws_dis.add_table(table)
    _autosize_columns(ws_dis, max_width=45)

    for ws in (ws_summary, ws_firma, ws_dis):
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="top",
                    wrap_text=cell.alignment.wrap_text,
                )

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


def print_report(report: AuditReport, detailed: bool = True) -> None:
    print("=== Firma bakiye / defter uyusmazlik raporu (READ-ONLY) ===")
    print("No database changes are written by this script.")
    print()
    print("=== Ozet ===")
    print(f"  Kontrol edilen firma: {report.checked_firma_count}")
    print(f"  Firma.bakiye drift: {len(report.firma_mismatches)}")
    print(f"  Kontrol edilen dis kiralama kalemi: {report.checked_dis_kiralama_count}")
    print(f"  Dis kiralama tutar uyusmazligi: {len(report.dis_kiralama_mismatches)}")

    print()
    print("=== Firma.bakiye cache uyusmazliklari ===")
    if not detailed:
        print(f"Detaylar Excel/CSV raporuna yazildi. Toplam bulgu: {len(report.firma_mismatches)}")
        print()
        print("=== Dis kiralama tutar uyusmazliklari ===")
        print(f"Detaylar Excel raporuna yazildi. Toplam bulgu: {len(report.dis_kiralama_mismatches)}")
        return

    if not report.firma_mismatches:
        print("Bulgu yok.")
    else:
        headers = [
            "firma_id",
            "firma_adi",
            "firma_bakiye_cache",
            "defter_net_bakiye",
            "fark",
            "defter_net_bakiye_kdvli",
            "cari_bakiye_kdvli_cache",
            "cari_cache_fark",
            "hizmet_kaydi_sayisi",
            "odeme_sayisi",
            "sebep_tahmini",
        ]
        print("\t".join(headers))
        for row in report.firma_mismatches:
            data = row.to_row()
            print("\t".join(data[h] for h in headers))

    print()
    print("=== Dis kiralama tutar uyusmazliklari ===")
    if not report.dis_kiralama_mismatches:
        print("Bulgu yok.")
    else:
        print("kalem_id\tfirma_id\tform_no\tbeklenen_tutar\tmevcut_hizmet_tutari\thizmet_id\tfark")
        for row in report.dis_kiralama_mismatches:
            print(
                "\t".join(
                    [
                        str(row.kalem_id),
                        str(row.firma_id),
                        row.form_no,
                        str(row.beklenen_tutar),
                        str(row.mevcut_hizmet_tutari),
                        str(row.hizmet_id),
                        str(row.fark),
                    ]
                )
            )


def _parse_today(value: Optional[str]) -> date:
    if not value:
        return date.today()
    return datetime.strptime(value, "%Y-%m-%d").date()


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Read-only Firma.bakiye / defter bakiye uyusmazlik raporu. No database changes."
    )
    parser.add_argument("--tolerance", type=str, default="0.01", help="Tutar toleransi (TL)")
    parser.add_argument("--today", type=str, default=None, help="Dis kiralama canli tahakkuk referans tarihi YYYY-MM-DD")
    parser.add_argument("--csv", type=str, default=None, help="Firma uyusmazliklarini CSV dosyasina yaz")
    parser.add_argument(
        "--xlsx",
        nargs="?",
        const="firma_bakiye_raporu.xlsx",
        default=None,
        help="Bicimli Excel raporu yaz. Path verilmezse firma_bakiye_raporu.xlsx kullanilir.",
    )
    parser.add_argument("--details", action="store_true", help="Excel/CSV yazarken terminalde tum detay satirlarini da bas")
    parser.add_argument(
        "--fail-on-findings",
        action="store_true",
        help="Bulgu varsa cikis kodunu 1 yap. Varsayilan raporlama komutlari basarili cikis kodu 0 doner.",
    )
    args = parser.parse_args(argv)

    app = create_app()
    with app.app_context():
        try:
            report = run_audit(today=_parse_today(args.today), tolerance=Decimal(args.tolerance))
        except OperationalError as exc:
            message = str(exc)
            if 'could not translate host name "db"' in message or "Name or service not known" in message:
                print("DB baglanti hatasi: Windows terminalindeki Python, Docker icindeki 'db' host adini cozemiyor.")
                print()
                print("Bu raporu Docker web konteynerinin icinden calistirin:")
                print("  docker exec kiralama_web python scripts/audit_firma_bakiye_defter_mismatch.py")
                print()
                print("CSV icin:")
                print("  docker exec kiralama_web python scripts/audit_firma_bakiye_defter_mismatch.py --csv /tmp/firma_bakiye_raporu.csv")
                print()
                print("No database changes were written.")
                return 2
            raise
        compact_terminal = bool(args.xlsx or args.csv) and not args.details
        print_report(report, detailed=not compact_terminal)
        if args.csv:
            write_csv(report.firma_mismatches, args.csv)
            print(f"\nCSV yazildi: {args.csv}")
        if args.xlsx:
            write_xlsx(report, args.xlsx)
            print(f"\nExcel yazildi: {args.xlsx}")
        print("\nNo database changes were written.")
        has_findings = bool(report.firma_mismatches or report.dis_kiralama_mismatches)
        return 1 if args.fail_on_findings and has_findings else 0


if __name__ == "__main__":
    raise SystemExit(main())
