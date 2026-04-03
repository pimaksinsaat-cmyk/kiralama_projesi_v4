from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


base = Path(__file__).resolve().parents[1]
out = base / "app" / "static" / "Teslim_Tutanagi_Alan_Eslestirme.xlsx"

rows = [
    ["Sablon Alani", "Tip", "Kaynak", "Kaynak Alan", "Doldurulan Deger", "Not"],
    ["form_no", "Ust seviye", "app/dokumanlar/engine_teslim_tutanagi.py", "kiralama.kiralama_form_no", "Kiralama form no", "Dolu"],
    ["gunun_tarihi", "Ust seviye", "app/dokumanlar/engine_teslim_tutanagi.py", "kiralama.kiralama_olusturma_tarihi", "dd.mm.yyyy", "Ilk yazdirmada set edilir"],
    ["musteri_unvan", "Ust seviye", "app/dokumanlar/engine_teslim_tutanagi.py", "musteri.firma_adi", "Buyuk harf unvan", "Dolu"],
    ["musteri_vergi", "Ust seviye", "app/dokumanlar/engine_teslim_tutanagi.py", "musteri.vergi_dairesi + musteri.vergi_no", "Vergi dairesi / vergi no", "Dolu"],
    ["musteri_adres", "Ust seviye", "app/dokumanlar/engine_teslim_tutanagi.py", "musteri.iletisim_bilgileri", "Musteri adres/iletisim", "Dolu"],
    ["makine_kullanim_yeri", "Ust seviye", "app/dokumanlar/engine_teslim_tutanagi.py", "kiralama.makine_calisma_adresi", "Kullanim yeri", "Yeni eklendi"],
    ["kalemler", "Dongu", "app/dokumanlar/teslim_tutanagi_hazirla.py", "for kalem in kiralama.kalemler", "Satir listesi", "Word tablosu"],
    ["k.ekipman", "Kalem alani", "app/dokumanlar/teslim_tutanagi_hazirla.py", "marka + model (fallback kod/tip)", "Ekipman adi", "Guncellendi"],
    ["k.ekipman_marka", "Kalem alani", "app/dokumanlar/teslim_tutanagi_hazirla.py", "kalem.ekipman.marka veya harici_ekipman_marka", "Marka", "Yeni eklendi"],
    ["k.ekipman_model", "Kalem alani", "app/dokumanlar/teslim_tutanagi_hazirla.py", "kalem.ekipman.model veya harici_ekipman_model", "Model", "Yeni eklendi"],
    ["k.seri_no", "Kalem alani", "app/dokumanlar/teslim_tutanagi_hazirla.py", "kalem.ekipman.seri_no veya harici_ekipman_seri_no", "Seri numarasi", "Dolu"],
    ["k.teslim_tarihi", "Kalem alani", "app/dokumanlar/teslim_tutanagi_hazirla.py", "kalem.kiralama_baslangici", "dd.mm.yyyy", "Dolu"],
    ["k.makine_kullanim_yeri", "Kalem alani", "app/dokumanlar/teslim_tutanagi_hazirla.py", "kiralama.makine_calisma_adresi", "Kullanim yeri", "Yeni eklendi"],
]

wb = Workbook()
ws = wb.active
ws.title = "Alan Eslestirme"

for row in rows:
    ws.append(row)

header_fill = PatternFill("solid", fgColor="1F4E78")
for cell in ws[1]:
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

for col, width in {"A": 24, "B": 14, "C": 48, "D": 44, "E": 26, "F": 22}.items():
    ws.column_dimensions[col].width = width

meta = wb.create_sheet("Meta")
meta["A1"] = "Olusturma Tarihi"
meta["B1"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
meta["A2"] = "Sablon"
meta["B2"] = "app/static/templates/Teslim_Tutanagi_TASLAK.docx"
meta["A3"] = "Not"
meta["B3"] = "Alanlar teslim_tutanagi_hazirla ve engine_teslim_tutanagi kaynaklarina gore eslestirilmistir."
meta.column_dimensions["A"].width = 24
meta.column_dimensions["B"].width = 120

out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print(out)
