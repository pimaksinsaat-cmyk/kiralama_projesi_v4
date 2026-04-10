from app.filo import filo_bp
from flask import render_template, redirect, url_for, flash, request, current_app, jsonify, send_file
from sqlalchemy.orm import joinedload, subqueryload
from sqlalchemy import or_, and_
from datetime import datetime, date, timedelta, timezone
from flask_login import current_user
from io import BytesIO
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook, load_workbook

# Servisler ve Doğrulama
from app.services.filo_services import EkipmanService, BakimService
from app.services.ekipman_rapor_services import EkipmanRaporuService
from app.services.base import ValidationError

# Modeller
from app.filo.models import Ekipman, BakimKaydi
from app.kiralama.models import Kiralama, KiralamaKalemi
from app.subeler.models import Sube
from app.araclar.models import Arac as NakliyeAraci
from app.firmalar.models import Firma
from app.utils import ensure_active_sube_exists, normalize_turkish_upper, tr_ilike
from app.extensions import db
from app.services.operation_log_service import OperationLogService

# Formlar
from app.filo.forms import EkipmanForm 
import locale

class ListPagination:
    """Bellek içi liste için sayfalama yardımcısı."""
    def __init__(self, items, page, per_page):
        self.total = len(items)
        self.per_page = per_page
        self.page = page
        self.pages = max(1, (self.total + per_page - 1) // per_page)
        self.page = min(max(page, 1), self.pages)
        start = (self.page - 1) * per_page
        self.items = items[start:start + per_page]
        self.has_prev = self.page > 1
        self.has_next = self.page < self.pages
        self.prev_num = self.page - 1
        self.next_num = self.page + 1

# Türkçe yerel ayarlarını dene
try:
    locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')
except:
    pass

# --- GÜVENLİK YARDIMCISI ---
def get_actor_id():
    """Kullanıcı giriş sistemi aktifse işlemi yapanın ID'sini alır."""
    if hasattr(current_app, 'login_manager'):
        try:
            if current_user.is_authenticated:
                return current_user.id
        except Exception:
            pass
    return None


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {'1', 'true', 'evet', 'yes', 'x'}:
        return True
    if text in {'0', 'false', 'hayir', 'hayır', 'no', ''}:
        return False
    return default


def _to_int(value, default=0):
    if value is None or str(value).strip() == '':
        return default
    try:
        return int(float(str(value).replace(',', '.')))
    except (ValueError, TypeError):
        return default


def _to_float(value):
    if value is None or str(value).strip() == '':
        return None
    try:
        return float(str(value).replace(',', '.'))
    except (ValueError, TypeError):
        return None


def _to_decimal(value, default='0'):
    if value is None or str(value).strip() == '':
        return Decimal(default)
    try:
        cleaned = str(value).replace(' ', '').replace(',', '.')
        return Decimal(cleaned)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


# -------------------------------------------------------------------------
# 1. Makine Parkı Listeleme (Sadece Aktifler)
# -------------------------------------------------------------------------
@filo_bp.route('/')
@filo_bp.route('/index')
def index():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        if per_page not in {10, 25, 50, 100}:
            per_page = 25
        q = request.args.get('q', '', type=str)

        # Detayli filtre parametreleri (filo/index modal filtresi)
        sube_id = request.args.get('sube_id', type=int)
        tip = request.args.get('tip', type=str)
        marka = request.args.get('marka', type=str)
        enerji = request.args.get('enerji', type=str)
        ortam = request.args.get('ortam', type=str)
        y_min = request.args.get('y_min', type=float)
        k_min = request.args.get('k_min', type=float)
        agirlik_max = request.args.get('agirlik_max', type=float)
        genislik_max = request.args.get('genislik_max', type=float)
        uzunluk_max = request.args.get('uzunluk_max', type=float)
        ky_max = request.args.get('ky_max', type=float)
        
        # SADECE AKTİF VE BİZİM OLANLAR
        base_query = Ekipman.query.filter(
            and_(
                Ekipman.firma_tedarikci_id.is_(None),
                Ekipman.is_active == True 
            )
        ).options(
            joinedload(Ekipman.sube),
            subqueryload(Ekipman.kiralama_kalemleri).options(
                joinedload(KiralamaKalemi.kiralama).joinedload(Kiralama.firma_musteri)
            )
        )
        
        if q:
            base_query = base_query.filter(
                or_(
                    tr_ilike(Ekipman.kod, f'%{q}%'),
                    tr_ilike(Ekipman.tipi, f'%{q}%'),
                    Ekipman.seri_no.ilike(f'%{q}%'),
                    tr_ilike(Ekipman.marka, f'%{q}%'),
                )
            )

        # Detayli filtreleri uygula
        if sube_id:
            base_query = base_query.filter(Ekipman.sube_id == sube_id)
        if tip:
            base_query = base_query.filter(Ekipman.tipi == tip)
        if marka:
            base_query = base_query.filter(Ekipman.marka == marka)
        if enerji:
            base_query = base_query.filter(Ekipman.yakit == enerji)
        if ortam == 'ic':
            base_query = base_query.filter(Ekipman.ic_mekan_uygun == True)
        elif ortam == 'dis':
            base_query = base_query.filter(Ekipman.arazi_tipi_uygun == True)

        if y_min is not None:
            base_query = base_query.filter(Ekipman.calisma_yuksekligi >= y_min)
        if k_min is not None:
            base_query = base_query.filter(Ekipman.kaldirma_kapasitesi >= k_min)
        if agirlik_max is not None:
            base_query = base_query.filter(Ekipman.agirlik <= agirlik_max)
        if genislik_max is not None:
            base_query = base_query.filter(Ekipman.genislik <= genislik_max)
        if uzunluk_max is not None:
            base_query = base_query.filter(Ekipman.uzunluk <= uzunluk_max)
        if ky_max is not None:
            base_query = base_query.filter(Ekipman.kapali_yukseklik <= ky_max)
        
        pagination = base_query.order_by(Ekipman.kod).paginate(
            page=page, per_page=per_page, error_out=False
        )
        ekipmanlar = pagination.items
        recent_threshold = datetime.now(timezone.utc) - timedelta(days=1)
        
        for ekipman in ekipmanlar:
            ekipman.aktif_kiralama_bilgisi = None
            ekipman.iade_iptal_kalem_id = None
            ekipman.son_musteri_adi = None
            if ekipman.calisma_durumu == 'kirada':
                aktif_kalemler = [k for k in ekipman.kiralama_kalemleri if not k.sonlandirildi]
                if aktif_kalemler:
                    ekipman.aktif_kiralama_bilgisi = max(aktif_kalemler, key=lambda k: k.id)
            else:
                sonlandirilmis_kalemler = [
                    k for k in ekipman.kiralama_kalemleri
                    if k.sonlandirildi and k.is_active and k.updated_at
                ]
                if sonlandirilmis_kalemler:
                    en_guncel_kalem = max(
                        sonlandirilmis_kalemler,
                        key=lambda k: k.updated_at or datetime.min.replace(tzinfo=timezone.utc)
                    )
                    kalem_updated_at = en_guncel_kalem.updated_at if en_guncel_kalem.updated_at.tzinfo else en_guncel_kalem.updated_at.replace(tzinfo=timezone.utc)
                    if kalem_updated_at >= recent_threshold:
                        ekipman.iade_iptal_kalem_id = en_guncel_kalem.id
                    
                    # Son müşteri adını ata (boşta durumda)
                    if en_guncel_kalem.kiralama and en_guncel_kalem.kiralama.firma_musteri:
                        ekipman.son_musteri_adi = en_guncel_kalem.kiralama.firma_musteri.firma_adi
        
        subeler = Sube.query.all()
        tipler = [
            t[0] for t in Ekipman.query.with_entities(Ekipman.tipi)
            .filter(Ekipman.tipi.isnot(None), Ekipman.tipi != '')
            .distinct()
            .order_by(Ekipman.tipi)
            .all()
        ]
        markalar = [
            m[0] for m in Ekipman.query.with_entities(Ekipman.marka)
            .filter(Ekipman.marka.isnot(None), Ekipman.marka != '')
            .distinct()
            .order_by(Ekipman.marka)
            .all()
        ]
        enerji_kaynaklari = [
            y[0] for y in Ekipman.query.with_entities(Ekipman.yakit)
            .filter(Ekipman.yakit.isnot(None), Ekipman.yakit != '')
            .distinct()
            .order_by(Ekipman.yakit)
            .all()
        ]
        nakliye_araclari = NakliyeAraci.aktif_nakliye_query().order_by(NakliyeAraci.plaka).all()
        nakliye_tedarikci_listesi = Firma.query.filter_by(is_tedarikci=True).order_by(Firma.firma_adi).all()
    
    except Exception as e:
        flash(f"Hata: {str(e)}", "danger")
        ekipmanlar = []
        pagination = None
        per_page = 25
        subeler = []
        tipler = []
        markalar = []
        enerji_kaynaklari = []
        nakliye_araclari = []
        nakliye_tedarikci_listesi = []

    return render_template(
        'filo/index.html',
        ekipmanlar=ekipmanlar,
        pagination=pagination,
        per_page=per_page,
        q=q,
        subeler=subeler,
        tipler=tipler,
        markalar=markalar,
        enerji_kaynaklari=enerji_kaynaklari,
        nakliye_araclari=nakliye_araclari,
        nakliye_tedarikci_listesi=nakliye_tedarikci_listesi,
    )


@filo_bp.route('/excel-disari-aktar', methods=['GET'])
def excel_disari_aktar():
    ekipmanlar = Ekipman.query.filter(
        Ekipman.firma_tedarikci_id.is_(None),
        Ekipman.is_active == True,
    ).options(joinedload(Ekipman.sube)).order_by(Ekipman.kod).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'MakineParki'

    headers = [
        'Kod', 'Seri No', 'Marka', 'Model', 'Tipi', 'Yakit',
        'Calisma Yuksekligi', 'Kaldirma Kapasitesi', 'Uretim Yili',
        'Giris Maliyeti', 'Para Birimi', 'Sube', 'Calisma Durumu',
        'Agirlik', 'Genislik', 'Uzunluk', 'Kapali Yukseklik',
        'Ic Mekan Uygun', 'Arazi Tipi Uygun', 'Aktif Mi'
    ]
    ws.append(headers)

    for e in ekipmanlar:
        ws.append([
            e.kod,
            e.seri_no,
            e.marka,
            e.model,
            e.tipi,
            e.yakit,
            e.calisma_yuksekligi,
            e.kaldirma_kapasitesi,
            e.uretim_yili,
            float(e.giris_maliyeti or 0),
            e.para_birimi,
            e.sube.isim if e.sube else '',
            e.calisma_durumu,
            e.agirlik,
            e.genislik,
            e.uzunluk,
            e.kapali_yukseklik,
            'Evet' if e.ic_mekan_uygun else 'Hayir',
            'Evet' if e.arazi_tipi_uygun else 'Hayir',
            'Evet' if e.is_active else 'Hayir',
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"makine_parki_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@filo_bp.route('/excel-ice-yukle', methods=['POST'])
def excel_ice_yukle():
    confirm_password = (
        request.form.get('confirm_password')
        or request.form.get('password')
        or request.form.get('sifre')
        or ''
    ).strip()
    if not (
        getattr(current_user, 'is_authenticated', False)
        and hasattr(current_user, 'check_password')
        and current_user.check_password(confirm_password)
    ):
        flash('Excel içe aktarım için kullanıcı şifrenizi doğru girmeniz gerekiyor.', 'danger')
        return redirect(url_for('filo.index'))

    file = request.files.get('excel_file')
    if not file or not file.filename:
        flash('Lütfen bir Excel dosyası seçiniz.', 'warning')
        return redirect(url_for('filo.index'))

    if not file.filename.lower().endswith('.xlsx'):
        flash('Sadece .xlsx uzantılı dosyalar destekleniyor.', 'danger')
        return redirect(url_for('filo.index'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as exc:
        flash(f'Excel dosyası okunamadı: {exc}', 'danger')
        return redirect(url_for('filo.index'))

    sube_map = {s.isim.strip().lower(): s.id for s in Sube.query.all() if s.isim}

    created = 0
    updated = 0
    skipped = 0
    errors = []

    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_idx, row in enumerate(rows, start=2):
        try:
            kod = (row[0] or '').strip() if row[0] else ''
            seri_no = (row[1] or '').strip() if row[1] else ''

            if not kod:
                skipped += 1
                continue

            marka = (row[2] or '').strip() if row[2] else ''
            model = (row[3] or '').strip() if row[3] else ''
            tipi = (row[4] or '').strip() if row[4] else ''
            yakit = (row[5] or '').strip() if row[5] else ''

            calisma_yuksekligi = _to_int(row[6], default=0)
            kaldirma_kapasitesi = _to_int(row[7], default=0)
            uretim_yili = _to_int(row[8], default=date.today().year)
            giris_maliyeti = _to_decimal(row[9], default='0')
            para_birimi = normalize_turkish_upper((row[10] or 'TRY').strip() if row[10] else 'TRY')[:3]

            sube_adi = (row[11] or '').strip() if row[11] else ''
            calisma_durumu = (row[12] or 'bosta').strip() if row[12] else 'bosta'

            agirlik = _to_float(row[13])
            genislik = _to_float(row[14])
            uzunluk = _to_float(row[15])
            kapali_yukseklik = _to_float(row[16])

            ic_mekan_uygun = _to_bool(row[17], default=False)
            arazi_tipi_uygun = _to_bool(row[18], default=False)
            is_active = _to_bool(row[19], default=True)

            sube_id = sube_map.get(sube_adi.lower()) if sube_adi else None

            mevcut = Ekipman.query.filter_by(kod=kod).first()

            if mevcut and mevcut.firma_tedarikci_id is not None:
                skipped += 1
                continue

            if not seri_no:
                seri_no = kod

            if not mevcut:
                seri_cakisiyor = Ekipman.query.filter_by(seri_no=seri_no, firma_tedarikci_id=None).first()
                if seri_cakisiyor:
                    errors.append(f'Satır {row_idx}: Seri no çakıştı ({seri_no}).')
                    continue

                yeni = Ekipman(
                    kod=kod,
                    seri_no=seri_no,
                    marka=marka or 'Bilinmiyor',
                    model=model,
                    tipi=tipi or '',
                    yakit=yakit or '',
                    calisma_yuksekligi=calisma_yuksekligi,
                    kaldirma_kapasitesi=kaldirma_kapasitesi,
                    uretim_yili=uretim_yili,
                    giris_maliyeti=giris_maliyeti,
                    para_birimi=para_birimi,
                    sube_id=sube_id,
                    calisma_durumu=calisma_durumu,
                    agirlik=agirlik,
                    genislik=genislik,
                    uzunluk=uzunluk,
                    kapali_yukseklik=kapali_yukseklik,
                    ic_mekan_uygun=ic_mekan_uygun,
                    arazi_tipi_uygun=arazi_tipi_uygun,
                    is_active=is_active,
                    firma_tedarikci_id=None,
                )
                db.session.add(yeni)
                created += 1
            else:
                if seri_no and seri_no != mevcut.seri_no:
                    seri_cakisiyor = Ekipman.query.filter(
                        Ekipman.id != mevcut.id,
                        Ekipman.seri_no == seri_no,
                        Ekipman.firma_tedarikci_id.is_(None),
                    ).first()
                    if seri_cakisiyor:
                        errors.append(f'Satır {row_idx}: Seri no çakıştı ({seri_no}).')
                        continue

                mevcut.seri_no = seri_no
                mevcut.marka = marka or mevcut.marka
                mevcut.model = model
                mevcut.tipi = tipi
                mevcut.yakit = yakit

                mevcut.calisma_yuksekligi = calisma_yuksekligi
                mevcut.kaldirma_kapasitesi = kaldirma_kapasitesi
                mevcut.uretim_yili = uretim_yili
                mevcut.giris_maliyeti = giris_maliyeti
                mevcut.para_birimi = para_birimi
                mevcut.sube_id = sube_id
                mevcut.calisma_durumu = calisma_durumu
                mevcut.agirlik = agirlik
                mevcut.genislik = genislik
                mevcut.uzunluk = uzunluk
                mevcut.kapali_yukseklik = kapali_yukseklik
                mevcut.ic_mekan_uygun = ic_mekan_uygun
                mevcut.arazi_tipi_uygun = arazi_tipi_uygun
                mevcut.is_active = is_active
                updated += 1

        except Exception as exc:
            errors.append(f'Satır {row_idx}: {exc}')

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f'Excel içe aktarma başarısız: {exc}', 'danger')
        return redirect(url_for('filo.index'))

    flash(
        f'Excel içe aktarım tamamlandı. Yeni: {created}, Güncellenen: {updated}, Atlanan: {skipped}.',
        'success',
    )
    if errors:
        flash('Bazı satırlar işlenemedi: ' + ' | '.join(errors[:5]), 'warning')

    return redirect(url_for('filo.index'))

# -------------------------------------------------------------------------
# 2. Yeni Makine Ekleme
# -------------------------------------------------------------------------
@filo_bp.route('/ekle', methods=['GET', 'POST'])
def ekle():
    guard_response = ensure_active_sube_exists(
        warning_message="Makine eklemeden önce en az bir aktif şube / depo tanımlamalısınız."
    )
    if guard_response:
        return guard_response

    form = EkipmanForm()
    form.sube_id.choices = [(s.id, s.isim) for s in Sube.query.all()]
    
    try:
        son_ekipman = Ekipman.query.filter(Ekipman.firma_tedarikci_id.is_(None)).order_by(Ekipman.kod.desc()).first()
        son_kod = son_ekipman.kod if son_ekipman else 'Henüz kayıt yok'
    except:
        son_kod = '...'

    if form.validate_on_submit():
        try:
            # Kullanıcı dostu özel arşiv uyarılarını koruyoruz (Servisten önce formda yakalıyoruz)
            mevcut_makine = Ekipman.query.filter_by(kod=form.kod.data).first()
            if mevcut_makine:
                durum = "zaten listenizde" if mevcut_makine.is_active else "ARŞİVDE (Pasif Durumda)"
                flash(f"HATA: '{form.kod.data}' kodlu makine {durum} mevcut.", "warning" if not mevcut_makine.is_active else "danger")
                return render_template('filo/ekipman_form.html', form=form, son_kod=son_kod, is_edit=False)
            
            mevcut_seri = Ekipman.query.filter_by(seri_no=form.seri_no.data, firma_tedarikci_id=None).first()
            if mevcut_seri:
                durum = f"zaten mevcut (Kod: {mevcut_seri.kod})" if mevcut_seri.is_active else "ARŞİVDE mevcut"
                flash(f"HATA: '{form.seri_no.data}' seri numaralı bir makine {durum}!", "warning" if not mevcut_seri.is_active else "danger")
                return render_template('filo/ekipman_form.html', form=form, son_kod=son_kod, is_edit=False)

            # BaseForm sayesinde temiz veri aktarımı (MoneyField, giris_maliyeti'ni otomatik Decimal yapar)
            data = {k: v for k, v in form.data.items() if k in EkipmanService.updatable_fields}
            data['firma_tedarikci_id'] = None
            data['calisma_durumu'] = 'bosta'
            data['is_active'] = True
            
            yeni_ekipman = Ekipman(**data)
            EkipmanService.save(yeni_ekipman, actor_id=get_actor_id())
            OperationLogService.log(
                module='filo', action='create',
                user_id=get_actor_id(),
                username=getattr(current_user, 'username', None),
                entity_type='Ekipman', entity_id=yeni_ekipman.id,
                description=f"{yeni_ekipman.kod} makine eklendi.",
                success=True
            )
            flash('Yeni makine başarıyla eklendi!', 'success')
            return redirect(url_for('filo.index'))
            
        except ValidationError as e:
            OperationLogService.log(
                module='filo', action='create',
                user_id=get_actor_id(),
                username=getattr(current_user, 'username', None),
                entity_type='Ekipman',
                description=f"Makine ekleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "danger")
        except Exception as e:
            OperationLogService.log(
                module='filo', action='create',
                user_id=get_actor_id(),
                username=getattr(current_user, 'username', None),
                entity_type='Ekipman',
                description=f"Makine ekleme hatası: {str(e)}",
                success=False
            )
            flash(f"Kayıt hatası: {str(e)}", "danger")
    
    return render_template('filo/ekipman_form.html', form=form, son_kod=son_kod,is_edit=False)
    
# -------------------------------------------------------------------------
# 3. Makine Bilgilerini Düzelt
# -------------------------------------------------------------------------
@filo_bp.route('/duzelt/<int:id>', methods=['GET', 'POST'])
def duzelt(id):
    guard_response = ensure_active_sube_exists(
        warning_message="Makine düzenlemeden önce en az bir aktif şube / depo tanımlamalısınız."
    )
    if guard_response:
        return guard_response

    ekipman = EkipmanService.get_by_id(id)
    if not ekipman or not ekipman.is_active or ekipman.firma_tedarikci_id is not None:
        flash("Geçerli bir makine bulunamadı.", "danger")
        return redirect(url_for('filo.index'))
        
    # obj=ekipman dendiğinde MoneyField sayesinde giris_maliyeti otomatik formatlanır!
    form = EkipmanForm(obj=ekipman)
    form.sube_id.choices = [(s.id, s.isim) for s in Sube.query.all()]

    # Bazi kayitlarda form baglama/format farkindan dolayi giris_maliyeti bos gelebiliyor.
    # Duzenleme ekraninda mevcut degeri her zaman gostermek icin GET'te zorunlu doldur.
    if request.method == 'GET':
        form.giris_maliyeti.data = ekipman.giris_maliyeti if ekipman.giris_maliyeti is not None else None

    if form.validate_on_submit():
        try:
            data = {k: v for k, v in form.data.items() if k in EkipmanService.updatable_fields}
            
            # Şube ve Durum Güncelleme Koruması
            if ekipman.calisma_durumu == 'kirada':
                # Kiradaki makinenin şubesi ve durumu formdan değiştirilemez
                data.pop('sube_id', None)
                data.pop('calisma_durumu', None)
            else:
                data['calisma_durumu'] = 'bosta'

            EkipmanService.update(id, data, actor_id=get_actor_id())
            OperationLogService.log(
                module='filo', action='update',
                user_id=get_actor_id(),
                username=getattr(current_user, 'username', None),
                entity_type='Ekipman', entity_id=id,
                description=f"{ekipman.kod} makine bilgileri güncellendi.",
                success=True
            )
            flash('Makine bilgileri güncellendi!', 'success')
            return redirect(url_for('filo.index', page=request.args.get('page', 1, type=int), q=request.args.get('q', '')))
            
        except ValidationError as e:
            OperationLogService.log(
                module='filo', action='update',
                user_id=get_actor_id(),
                username=getattr(current_user, 'username', None),
                entity_type='Ekipman', entity_id=id,
                description=f"Makine güncelleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "danger")
        except Exception as e:
            OperationLogService.log(
                module='filo', action='update',
                user_id=get_actor_id(),
                username=getattr(current_user, 'username', None),
                entity_type='Ekipman', entity_id=id,
                description=f"Makine güncelleme hatası: {str(e)}",
                success=False
            )
            flash(f"Hata: {str(e)}", "danger")

    return render_template('filo/ekipman_form.html', form=form, ekipman=ekipman, subeler=Sube.query.all(), is_edit=True)

# -------------------------------------------------------------------------
# 4. Şube Nakil Merkezi (Modal Hızlı Transfer)
# -------------------------------------------------------------------------
@filo_bp.route('/sube_degistir/<int:id>', methods=['POST'])
def sube_degistir(id):
    try:
        yeni_sube_id = request.form.get('yeni_sube_id', type=int)
        if not yeni_sube_id:
            flash("Uyarı: Lütfen geçerli bir şube seçin.", "warning")
            return redirect(request.referrer or url_for('filo.index'))
            
        EkipmanService.sube_transfer(id, yeni_sube_id, actor_id=get_actor_id())
        OperationLogService.log(
            module='filo', action='sube_transfer',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=id,
            description=f"Makine #{id} şube transferi yapıldı (şube: {yeni_sube_id}).",
            success=True
        )
        flash(f"Nakil Başarılı! Makine yeni deposuna aktarıldı.", "success")
    except ValidationError as e:
        OperationLogService.log(
            module='filo', action='sube_transfer',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=id,
            description=f"Şube transfer hatası: {str(e)}",
            success=False
        )
        flash(f"Hata: {str(e)}", "danger")
    except Exception as e:
        OperationLogService.log(
            module='filo', action='sube_transfer',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=id,
            description=f"Şube transfer hatası: {str(e)}",
            success=False
        )
        flash(f"Nakil hatası: {str(e)}", "danger")
        
    return redirect(request.referrer or url_for('filo.index'))

# -------------------------------------------------------------------------
# 5. Makine Bilgi ve Geçmişi
# -------------------------------------------------------------------------
@filo_bp.route('/bilgi/<int:id>', methods=['GET'])
def bilgi(id):
    ekipman = Ekipman.query.filter(
        Ekipman.id == id,
        Ekipman.firma_tedarikci_id.is_(None) 
    ).options(
        subqueryload(Ekipman.kiralama_kalemleri).options(
            joinedload(KiralamaKalemi.kiralama).joinedload(Kiralama.firma_musteri)
        )
    ).first_or_404()
    
    kalemler = sorted(ekipman.kiralama_kalemleri, key=lambda k: k.id, reverse=True)
    referrer = request.args.get('back') or request.referrer or url_for('filo.index')
    return render_template('filo/bilgi.html', ekipman=ekipman, kalemler=kalemler, referrer=referrer)

# -------------------------------------------------------------------------
# 6. Kiralama Sonlandırma (İnce Tarih Kontrolü ve Şube Ataması)
# -------------------------------------------------------------------------
@filo_bp.route('/sonlandir', methods=['POST'])
def sonlandir():
    try:
        ekipman_id = request.form.get('ekipman_id', type=int)
        bitis_tarihi_str = request.form.get('bitis_tarihi') 
        donus_sube_id = request.form.get('donus_sube_id', type=int)
        
        if not (ekipman_id and bitis_tarihi_str and donus_sube_id):
            flash('Eksik bilgi! Lütfen tarih ve dönüş şubesini seçiniz.', 'danger')
            return redirect(url_for('filo.index'))

        EkipmanService.kiralama_sonlandir(ekipman_id, bitis_tarihi_str, donus_sube_id, actor_id=get_actor_id())
        OperationLogService.log(
            module='filo', action='kiralama_sonlandir',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=ekipman_id,
            description=f"Makine #{ekipman_id} kiralama sonlandırıldı, depoya iade alındı.",
            success=True
        )
        flash(f"Kiralama sonlandırıldı. Makine depoya iade alındı.", 'success')

    except ValidationError as e:
        OperationLogService.log(
            module='filo', action='kiralama_sonlandir',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=ekipman_id,
            description=f"Kiralama sonlandırma hatası: {str(e)}",
            success=False
        )
        flash(f"Hata: {str(e)}", 'danger')
    except Exception as e:
        OperationLogService.log(
            module='filo', action='kiralama_sonlandir',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=ekipman_id,
            description=f"Kiralama sonlandırma hatası: {str(e)}",
            success=False
        )
        flash(f"Sistem Hatası: {str(e)}", 'danger')
        
    return redirect(url_for('filo.index', page=request.args.get('page', 1, type=int), q=request.args.get('q', '')))

# -------------------------------------------------------------------------
# 7. Bakım ve Servis İşlemleri
# -------------------------------------------------------------------------
@filo_bp.route('/bakima_al', methods=['POST'])
def bakima_al():
    """Filo listesinden makineyi bakıma al ve servis kaydı oluştur"""
    try:
        ekipman_id = request.form.get('ekipman_id', type=int)
        tarih = request.form.get('tarih')
        calisma_saati = request.form.get('calisma_saati', type=int)

        if ekipman_id is None or not tarih:
            flash('Eksik bilgi! Lütfen bakım başlangıç tarihini giriniz.', 'danger')
            return redirect(url_for('filo.index'))

        bakim_verileri = {
            'tarih': tarih,
            'bakim_tipi': 'ariza',
            'servis_tipi': 'ic_servis',
            'servis_veren_firma_id': None,
            'servis_veren_kisi': None,
            'aciklama': 'Filo listesinden bakıma alındı. Detaylar bakımdaki makine kartından yönetilecek.',
            'calisma_saati': calisma_saati,
            'sonraki_bakim_tarihi': None,
            'toplam_iscilik_maliyeti': Decimal('0'),
            'durum': 'acik',
        }

        BakimService.bakim_kaydet(ekipman_id, bakim_verileri, actor_id=get_actor_id())
        OperationLogService.log(
            module='filo', action='bakima_al',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=ekipman_id,
            description=f"Makine #{ekipman_id} bakıma alındı.",
            success=True
        )
        flash("Makine başarıyla bakıma alındı ve servis kaydı oluşturuldu.", 'success')
    except ValidationError as e:
        OperationLogService.log(
            module='filo', action='bakima_al',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=ekipman_id,
            description=f"Bakıma alma hatası: {str(e)}",
            success=False
        )
        flash(f"Uyarı: {str(e)}", 'warning')
    except Exception as e:
        OperationLogService.log(
            module='filo', action='bakima_al',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=ekipman_id,
            description=f"Bakıma alma hatası: {str(e)}",
            success=False
        )
        flash(f"Hata: {str(e)}", 'danger')

    return redirect(url_for('filo.index'))

# -------------------------------------------------------------------------
# 8. Arşiv, Silme ve Harici Makineler
# -------------------------------------------------------------------------
@filo_bp.route('/sil/<int:id>', methods=['POST'])
def sil(id):
    try:
        ekipman = EkipmanService.get_by_id(id)
        EkipmanService.delete(id, actor_id=get_actor_id())
        OperationLogService.log(
            module='filo', action='delete',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=id,
            description=f"{getattr(ekipman, 'kod', id)} arşive kaldırıldı.",
            success=True
        )
        flash('Makine arşive kaldırıldı.', 'success')
    except ValidationError as e:
        OperationLogService.log(
            module='filo', action='delete',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=id,
            description=f"Arşivleme hatası: {str(e)}",
            success=False
        )
        flash(str(e), 'danger')
    return redirect(url_for('filo.index'))

@filo_bp.route('/arsiv')
def arsiv():
    ekipmanlar = EkipmanService.find_by(is_active=False, firma_tedarikci_id=None)
    return render_template('filo/arsiv.html', ekipmanlar=ekipmanlar)

@filo_bp.route('/geri_yukle/<int:id>', methods=['POST'])
def geri_yukle(id):
    try:
        ekipman = EkipmanService.get_by_id(id)
        EkipmanService.update(id, {'is_active': True}, actor_id=get_actor_id())
        OperationLogService.log(
            module='filo', action='restore',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=id,
            description=f"{getattr(ekipman, 'kod', id)} arşivden geri yüklendi.",
            success=True
        )
        flash("Makine geri yüklendi.", "success")
    except ValidationError as e:
        flash(str(e), "danger")
    return redirect(url_for('filo.arsiv'))

@filo_bp.route('/harici')
def harici():
    ekipmanlar = Ekipman.query.filter(Ekipman.firma_tedarikci_id.isnot(None), Ekipman.is_active == True).all()
    return render_template('filo/harici.html', ekipmanlar=ekipmanlar)


# -------------------------------------------------------------------------
# 9. Finansal Raporlama - Makine ROI ve Amorti Analizi
# -------------------------------------------------------------------------
@filo_bp.route('/finansal_rapor/<int:ekipman_id>', methods=['GET', 'POST'])
def finansal_rapor(ekipman_id):
    """
    Makinenin finansal analizini gösterir:
    - Başlangıç maliyeti
    - Kiralama gelirleri (dönem bazında)
    - Servis masrafları
    - ROI ve amorti durumu
    """
    ekipman = Ekipman.query.get_or_404(ekipman_id)
    
    # Form'dan tarih aralığı al veya varsayılan değerler kullan
    start_date = None
    end_date = date.today()
    
    if request.method == 'POST':
        start_str = request.form.get('start_date')
        end_str = request.form.get('end_date')
        
        if start_str:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        if end_str:
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    else:
        # İlk ziyarette: Yılbaşından bugüne (YTD)
        start_date = date(end_date.year, 1, 1)
    
    # Finansal analiz hesapla
    ozet = EkipmanRaporuService.get_finansal_ozet(ekipman_id, start_date, end_date)
    
    # Kiralama detayları (tablo için)
    kiralama_detaylari = EkipmanRaporuService.get_kiralama_detaylari(ekipman_id, start_date, end_date)
    
    # Durum etiketi (Türkçe)
    durum_etiketleri = {
        'amorti_olmadi_zarar': 'Henüz Amorti Olmadı (Zarar)',
        'amorti_surecinde': 'Amorti Süreci İçinde',
        'amorti_oldu': 'Kendini Amorti Etti',
        'kar_asamasi': 'Kâr Aşamasında'
    }
    
    ozet['durum_etiket'] = durum_etiketleri.get(ozet['durum'], ozet['durum'])
    
    return render_template(
        'filo/finansal_rapor.html',
        ekipman=ekipman,
        ozet=ozet,
        kiralama_detaylari=kiralama_detaylari,
        start_date=start_date,
        end_date=end_date
    )


@filo_bp.route('/finansal_rapor_api/<int:ekipman_id>')
def finansal_rapor_api(ekipman_id):
    """
    Finansal rapor verilerini JSON formatında döner (grafik oluşturma için)
    """
    ekipman = Ekipman.query.get(ekipman_id)
    if not ekipman:
        return jsonify({'error': 'Makine bulunamadı'}), 404
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    ozet = EkipmanRaporuService.get_finansal_ozet(ekipman_id, start_date, end_date)
    
    return jsonify(ozet)


# -------------------------------------------------------------------------
# 10. Kiralama Geçmişi (Detaylı Tablo)
# -------------------------------------------------------------------------
@filo_bp.route('/kiralama_gecmisi/<int:ekipman_id>', methods=['GET'])
def kiralama_gecmisi(ekipman_id):
    """
    Makinenin tüm kiralama geçmişini gösterir:
    - Her kiralama kalemi için TRY, USD, EUR cinsinden gelir
    - Döviz kuru bilgileri
    - Toplam satırlar
    """
    ekipman = Ekipman.query.get_or_404(ekipman_id)
    
    # Tarih aralığı URL parametrelerinden al veya varsayılanları kullan
    end_date = date.today()
    start_str = request.args.get('start_date')
    end_str = request.args.get('end_date')
    start_date = datetime.strptime(start_str, '%Y-%m-%d').date() if start_str else date(end_date.year, 1, 1)
    if end_str:
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    
    # Kiralama detayları al
    kiralama_detaylari_tumu = EkipmanRaporuService.get_kiralama_detaylari(ekipman_id, start_date, end_date)
    
    # Totalleri tam liste üzerinden hesapla
    total_gun = sum(d['gun_sayisi'] for d in kiralama_detaylari_tumu)
    total_try = sum(d['gelir_try'] for d in kiralama_detaylari_tumu)
    total_usd = sum(d['gelir_usd'] for d in kiralama_detaylari_tumu)
    total_eur = sum(d['gelir_eur'] for d in kiralama_detaylari_tumu)
    
    # Sayfalama
    _ALLOWED_PER_PAGE = {10, 20, 25, 50, 100}
    per_page = request.args.get('per_page', 25, type=int)
    if per_page not in _ALLOWED_PER_PAGE:
        per_page = 25
    page = request.args.get('page', 1, type=int)
    pagination = ListPagination(kiralama_detaylari_tumu, page, per_page)
    
    # Kısır döngüyü kırmak için asıl geri adresini koru
    return_to_param = request.args.get('return_to')
    referrer_param = request.args.get('referrer')
    header_referrer = request.referrer

    referrer = return_to_param or referrer_param
    if not referrer:
        if header_referrer and ('/filo/bilgi/' not in header_referrer) and ('/filo/kiralama_gecmisi/' not in header_referrer):
            referrer = header_referrer
        else:
            referrer = url_for('filo.index')

    # Bilgi sayfasından geri gelindiğinde yine bu geçmiş sayfasına dönsün
    bilgi_back_url = url_for('filo.kiralama_gecmisi', ekipman_id=ekipman_id, return_to=referrer)
    
    return render_template(
        'filo/kiralama_gecmisi.html',
        ekipman=ekipman,
        kiralama_detaylari=pagination.items,
        pagination=pagination,
        per_page=per_page,
        total_gun=total_gun,
        total_try=total_try,
        total_usd=total_usd,
        total_eur=total_eur,
        start_date=start_date,
        end_date=end_date,
        referrer=referrer,
        bilgi_back_url=bilgi_back_url
    )
