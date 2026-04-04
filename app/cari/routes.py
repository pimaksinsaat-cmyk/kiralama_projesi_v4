from io import BytesIO

from flask import render_template, redirect, url_for, flash, request, send_file
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.cari import cari_bp
# DÜZELTME: KasaHizliIslemForm ve KasaTransferForm import listesine eklendi
from app.cari.forms import OdemeForm, HizmetKaydiForm, KasaForm, KasaTransferForm, KasaHizliIslemForm
from app.cari.models import Kasa, Odeme, HizmetKaydi
from app.firmalar.models import Firma

# --- YENİ SERVİS MİMARİSİ İÇE AKTARIMLARI ---
from app.services.cari_services import (
    KasaService, OdemeService, HizmetKaydiService, 
    CariRaporService, get_dahili_islem_firmasi
)
from app.services.base import ValidationError
from app.services.operation_log_service import OperationLogService

from decimal import Decimal
from datetime import date, datetime
import traceback
import logging

# -------------------------------------------------------------------------
# 🛠️ YARDIMCI FONKSİYONLAR
# -------------------------------------------------------------------------
class ListPagination:
    def __init__(self, total, page, per_page):
        self.total = max(0, int(total or 0))
        self.page = max(1, int(page or 1))
        self.per_page = max(1, int(per_page or 1))

        if self.total == 0:
            self.pages = 0
        else:
            self.pages = (self.total + self.per_page - 1) // self.per_page

        if self.pages and self.page > self.pages:
            self.page = self.pages

    @property
    def has_prev(self):
        return self.page > 1

    @property
    def has_next(self):
        return self.page < self.pages

    @property
    def prev_num(self):
        return self.page - 1

    @property
    def next_num(self):
        return self.page + 1

    def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
        last = 0
        for num in range(1, self.pages + 1):
            if (
                num <= left_edge
                or (self.page - left_current - 1 < num < self.page + right_current)
                or num > self.pages - right_edge
            ):
                if last + 1 != num:
                    yield None
                yield num
                last = num


def get_actor():
    """Audit Log için işlemi yapan kullanıcının ID'sini döndürür."""
    return current_user.id if current_user.is_authenticated else None

# -------------------------------------------------------------------------
# 1. ÖDEME VE TAHSİLAT
# -------------------------------------------------------------------------
@cari_bp.route('/odeme/ekle', methods=['GET', 'POST'])
def odeme_ekle():
    firma_id = request.args.get('firma_id', type=int)
    yon_param = request.args.get('yon', 'tahsilat')
    form = OdemeForm()
    
    # Seçenekleri dinamik yükle (is_deleted kontrolü servis üzerinden değil modelden geçici olarak yapılıyor)
    form.firma_musteri_id.choices = [(f.id, f.firma_adi) for f in Firma.query.filter_by(is_active=True, is_deleted=False).all()]
    form.kasa_id.choices = [(k.id, f"{k.kasa_adi} ({k.bakiye} {k.para_birimi})") 
                            for k in Kasa.query.filter_by(is_active=True, is_deleted=False).all()]
    
    if request.method == 'GET':
        if firma_id: form.firma_musteri_id.data = firma_id
        form.tarih.data = date.today()
        form.yon.data = yon_param

    if form.validate_on_submit():
        try:
            yeni_odeme = Odeme(
                firma_musteri_id=form.firma_musteri_id.data,
                kasa_id=form.kasa_id.data,
                tarih=form.tarih.data,
                tutar=form.tutar.data, # MoneyField sayesinde otomatik Decimal
                yon=form.yon.data,
                aciklama=form.aciklama.data,
                fatura_no=form.fatura_no.data,
                vade_tarihi=form.vade_tarihi.data
            )
            
            OdemeService.save(yeni_odeme, is_new=True, actor_id=get_actor())
            OperationLogService.log(
                module='cari', action='odeme_ekle',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Odeme', entity_id=yeni_odeme.id,
                description=f"{yeni_odeme.yon.upper()} {yeni_odeme.tutar} - {getattr(yeni_odeme, 'aciklama', '')}",
                success=True
            )
            flash('İşlem başarıyla kaydedildi.', 'success')
            return redirect(url_for('firmalar.bilgi', id=form.firma_musteri_id.data))
            
        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='odeme_ekle',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Odeme',
                description=f"Ödeme ekleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")
        except Exception as e:
            OperationLogService.log(
                module='cari', action='odeme_ekle',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Odeme',
                description=f"Ödeme ekleme hatası: {str(e)}",
                success=False
            )
            flash(f"Hata: {str(e)}", "danger")
            
    return render_template('cari/odeme_ekle.html', form=form)

@cari_bp.route('/odeme/duzelt/<int:id>', methods=['GET', 'POST'])
def odeme_duzelt(id):
    odeme = OdemeService.get_by_id(id)
    if not odeme or odeme.is_deleted:
        flash("Kayıt bulunamadı", "danger")
        return redirect(request.referrer or url_for('cari.finans_menu'))

    form = OdemeForm(obj=odeme)
    form.firma_musteri_id.choices = [(f.id, f.firma_adi) for f in Firma.query.filter_by(is_active=True, is_deleted=False).all()]
    form.kasa_id.choices = [(k.id, f"{k.kasa_adi} ({k.bakiye} {k.para_birimi})") 
                            for k in Kasa.query.filter_by(is_active=True, is_deleted=False).all()]

    # Yön bilgisine göre başlık ve buton metni belirle
    if odeme.yon == 'odeme':
        page_title = "Ödeme Düzenle"
        submit_text = "Ödemeyi Kaydet"
    else:
        page_title = "Tahsilat Düzenle"
        submit_text = "Tahsilatı Kaydet"

    if form.validate_on_submit():
        try:
            odeme.firma_musteri_id = form.firma_musteri_id.data
            odeme.kasa_id = form.kasa_id.data
            odeme.tarih = form.tarih.data
            odeme.tutar = form.tutar.data
            odeme.yon = form.yon.data
            odeme.aciklama = form.aciklama.data
            odeme.fatura_no = form.fatura_no.data
            odeme.vade_tarihi = form.vade_tarihi.data

            OdemeService.save(odeme, is_new=False, actor_id=get_actor())
            OperationLogService.log(
                module='cari', action='odeme_duzelt',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Odeme', entity_id=id,
                description=f"Ödeme #{id} güncellendi.",
                success=True
            )
            flash('İşlem güncellendi.', 'success')
            return redirect(url_for('firmalar.bilgi', id=odeme.firma_musteri_id))
        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='odeme_duzelt',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Odeme', entity_id=id,
                description=f"Ödeme güncelleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")

    return render_template('cari/odeme_ekle.html', form=form, title=page_title, submit_text=submit_text)

@cari_bp.route('/odeme/sil/<int:id>', methods=['POST'])
def odeme_sil(id):
    odeme = OdemeService.get_by_id(id)
    if not odeme:
        flash('Ödeme kaydı bulunamadı.', 'danger')
        return redirect(request.referrer or url_for('cari.kasa_listesi'))
        
    f_id = odeme.firma_musteri_id
    try:
        OdemeService.delete(id, actor_id=get_actor())
        OperationLogService.log(
            module='cari', action='odeme_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='Odeme', entity_id=id,
            description=f"Ödeme #{id} silindi.",
            success=True
        )
        flash('Ödeme/Tahsilat kaydı silindi ve kasa bakiyesi düzeltildi.', 'success')
    except ValidationError as e:
        OperationLogService.log(
            module='cari', action='odeme_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='Odeme', entity_id=id,
            description=f"Ödeme silme hatası: {str(e)}",
            success=False
        )
        flash(str(e), "warning")
    except Exception as e:
        OperationLogService.log(
            module='cari', action='odeme_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='Odeme', entity_id=id,
            description=f"Ödeme silme hatası: {str(e)}",
            success=False
        )
        flash(f'Hata: {str(e)}', 'danger')
        
    return redirect(url_for('firmalar.bilgi', id=f_id))

# -------------------------------------------------------------------------
# 2. HİZMET / FATURA
# -------------------------------------------------------------------------
@cari_bp.route('/hizmet/ekle', methods=['GET', 'POST'])
def hizmet_ekle():
    firma_id = request.args.get('firma_id', type=int)
    form = HizmetKaydiForm()
    form.firma_id.choices = [(f.id, f.firma_adi) for f in Firma.query.filter_by(is_active=True, is_deleted=False).all()]
    
    if request.method == 'GET':
        if firma_id:
            form.firma_id.data = firma_id
        form.tarih.data = date.today()

    firma = None
    if form.firma_id.data:
        firma = Firma.query.get(form.firma_id.data)

    if form.validate_on_submit():
        try:
            logging.warning(f"[HIZMET_EKLE] Form verileri: firma_id={form.firma_id.data}, tarih={form.tarih.data}, tutar={form.tutar.data}, yon={form.yon.data}, aciklama={form.aciklama.data}, fatura_no={form.fatura_no.data}, kdv_orani={form.kdv_orani.data}")
            yeni_hizmet = HizmetKaydi(
                firma_id=form.firma_id.data,
                tarih=form.tarih.data,
                tutar=form.tutar.data,
                yon=form.yon.data,
                aciklama=form.aciklama.data,
                fatura_no=form.fatura_no.data,
                kdv_orani=int(form.kdv_orani.data) if form.kdv_orani.data is not None else None  # Decimal yerine int
            )
            HizmetKaydiService.save(yeni_hizmet, actor_id=get_actor())
            OperationLogService.log(
                module='cari', action='hizmet_ekle',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='HizmetKaydi', entity_id=yeni_hizmet.id,
                description=f"Hizmet/Fatura kaydı eklendi: {getattr(yeni_hizmet, 'aciklama', '')}",
                success=True
            )
            flash('Fatura kaydedildi.', 'success')
            return redirect(url_for('firmalar.bilgi', id=form.firma_id.data))
        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='hizmet_ekle',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='HizmetKaydi',
                description=f"Hizmet ekleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")

    return render_template('cari/hizmet_ekle.html', form=form, firma=firma)

@cari_bp.route('/hizmet/duzelt/<int:id>', methods=['GET', 'POST'])
def hizmet_duzelt(id):
    hizmet = HizmetKaydiService.get_by_id(id)
    if not hizmet or hizmet.is_deleted:
        flash('Kayıt bulunamadı', 'danger')
        return redirect(request.referrer)
    
    form = HizmetKaydiForm(obj=hizmet)
    form.firma_id.choices = [(f.id, f.firma_adi) for f in Firma.query.filter_by(is_active=True, is_deleted=False).all()]

    if form.validate_on_submit():
        try:
            hizmet.firma_id = form.firma_id.data
            hizmet.tarih = form.tarih.data
            hizmet.tutar = form.tutar.data
            hizmet.kdv_orani = int(form.kdv_orani.data) if form.kdv_orani.data is not None else None
            hizmet.yon = form.yon.data
            hizmet.aciklama = form.aciklama.data
            hizmet.fatura_no = form.fatura_no.data
            
            HizmetKaydiService.save(hizmet, is_new=False, actor_id=get_actor())
            OperationLogService.log(
                module='cari', action='hizmet_duzelt',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='HizmetKaydi', entity_id=id,
                description=f"Hizmet/Fatura #{id} güncellendi.",
                success=True
            )
            flash('Fatura güncellendi.', 'success')
            return redirect(url_for('firmalar.bilgi', id=hizmet.firma_id))
        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='hizmet_duzelt',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='HizmetKaydi', entity_id=id,
                description=f"Hizmet güncelleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")
            
    return render_template('cari/hizmet_duzelt.html', form=form, hizmet=hizmet, title="Fatura Düzenle")

@cari_bp.route('/hizmet/sil/<int:id>', methods=['POST'])
def hizmet_sil(id):
    hizmet = HizmetKaydiService.get_by_id(id)
    if not hizmet: return redirect(request.referrer)
        
    f_id = hizmet.firma_id
    try:
        HizmetKaydiService.delete(id, actor_id=get_actor())
        OperationLogService.log(
            module='cari', action='hizmet_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='HizmetKaydi', entity_id=id,
            description=f"Hizmet/Fatura #{id} silindi.",
            success=True
        )
        flash('Fatura kaydı silindi.', 'success')
    except ValidationError as e:
        OperationLogService.log(
            module='cari', action='hizmet_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='HizmetKaydi', entity_id=id,
            description=f"Hizmet silme hatası: {str(e)}",
            success=False
        )
        flash(str(e), 'warning')
        
    return redirect(url_for('firmalar.bilgi', id=f_id))

@cari_bp.route('/hizmet/detay/<int:id>', methods=['GET'])
def hizmet_detay(id):
    hizmet = HizmetKaydiService.get_by_id(id)
    if not hizmet or hizmet.is_deleted:
        flash('Kayıt bulunamadı', 'danger')
        return redirect(request.referrer or url_for('main.index'))
    return render_template('cari/hizmet_detay.html', hizmet=hizmet)

# -------------------------------------------------------------------------
# 3. KASA YÖNETİMİ VE TRANSFER
# -------------------------------------------------------------------------
@cari_bp.route('/kasa/listesi')
def kasa_listesi():
    kasalar = KasaService.find_by(is_active=True, is_deleted=False)
    
    # Modal transfer formu
    transfer_form = KasaTransferForm()
    kasa_choices = [(k.id, f"{k.kasa_adi} ({k.para_birimi})") for k in kasalar]
    transfer_form.kaynak_kasa_id.choices = kasa_choices
    transfer_form.hedef_kasa_id.choices = kasa_choices
    
    # Modal hızlı işlem formu
    hizli_form = KasaHizliIslemForm()
    hizli_form.kasa_id.choices = kasa_choices

    kasalar_json = [
        {
            'id': str(k.id),
            'adi': k.kasa_adi or '',
            'birim': k.para_birimi or '',
            'bakiye': float(k.bakiye or 0),
        }
        for k in kasalar
    ]

    return render_template('cari/kasa_listesi.html', 
                           kasalar=kasalar, 
                           form=transfer_form, 
                           hizli_form=hizli_form,
                           kasalar_json=kasalar_json)

@cari_bp.route('/kasa/transfer', methods=['POST'])
def kasa_transfer():
    kasalar = KasaService.find_by(is_active=True, is_deleted=False)
    form = KasaTransferForm()
    kasa_choices = [(k.id, k.kasa_adi) for k in kasalar]
    form.kaynak_kasa_id.choices = kasa_choices
    form.hedef_kasa_id.choices = kasa_choices

    if form.validate_on_submit():
        try:
            KasaService.transfer_yap(
                form.kaynak_kasa_id.data, 
                form.hedef_kasa_id.data, 
                form.tutar.data, 
                actor_id=get_actor()
            )
            OperationLogService.log(
                module='cari', action='kasa_transfer',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa',
                description=f"Kasa transfer: {form.tutar.data} (kasa {form.kaynak_kasa_id.data} → {form.hedef_kasa_id.data}).",
                success=True
            )
            flash('Para transferi başarıyla tamamlandı.', 'success')
        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='kasa_transfer',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa',
                description=f"Kasa transfer hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{getattr(form, field).label.text}: {error}", "danger")
                
    return redirect(url_for('cari.kasa_listesi'))

@cari_bp.route('/kasa/hizli_islem', methods=['POST'])
def kasa_hizli_islem():
    """
    KasaHizliIslemForm kullanarak manuel parse işlemlerinden kurtulduk.
    """
    form = KasaHizliIslemForm()
    kasalar = KasaService.find_by(is_active=True, is_deleted=False)
    form.kasa_id.choices = [(k.id, k.kasa_adi) for k in kasalar]

    if form.validate_on_submit():
        try:
            dahili = get_dahili_islem_firmasi()
            yon = 'tahsilat' if form.islem_yonu.data == 'giris' else 'odeme'
            
            odeme = Odeme(
                firma_musteri_id=dahili.id,
                kasa_id=form.kasa_id.data,
                tarih=date.today(),
                tutar=form.tutar.data, # MoneyField sayesinde temiz Decimal gelir
                yon=yon,
                aciklama=f"Hızlı Kasa İşlemi: {form.aciklama.data}"
            )

            OdemeService.save(odeme, actor_id=get_actor())
            OperationLogService.log(
                module='cari', action='kasa_hizli_islem',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa', entity_id=form.kasa_id.data,
                description=f"Hızlı kasa işlemi: {form.islem_yonu.data} {form.tutar.data} - {form.aciklama.data}",
                success=True
            )
            flash("Kasa işlemi başarıyla kaydedildi", "success")

        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='kasa_hizli_islem',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa',
                description=f"Hızlı kasa işlem hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")
        except Exception as e:
            OperationLogService.log(
                module='cari', action='kasa_hizli_islem',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa',
                description=f"Hızlı kasa işlem hatası: {str(e)}",
                success=False
            )
            flash(f"Hızlı kasa işlem hatası: {str(e)}", "danger")
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{getattr(form, field).label.text}: {error}", "danger")

    return redirect(url_for('cari.kasa_listesi'))

@cari_bp.route('/kasa/ekle', methods=['GET', 'POST'])
def kasa_ekle():
    form = KasaForm()

    if form.validate_on_submit():
        try:
            yeni_kasa = Kasa(
                kasa_adi=form.kasa_adi.data,
                tipi=form.tipi.data,
                para_birimi=form.para_birimi.data,
                banka_sube_adi=(form.banka_sube_adi.data or '').strip() or None,
                bakiye=form.bakiye.data or 0
            )
            KasaService.save(yeni_kasa, actor_id=get_actor())
            OperationLogService.log(
                module='cari', action='kasa_ekle',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa', entity_id=yeni_kasa.id,
                description=f"{yeni_kasa.kasa_adi} kasası oluşturuldu.",
                success=True
            )
            flash(f'{yeni_kasa.kasa_adi} başarıyla oluşturuldu.', 'success')
            return redirect(url_for('cari.kasa_listesi'))
        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='kasa_ekle',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa',
                description=f"Kasa ekleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")
            
    return render_template('cari/kasa_ekle.html', form=form, title="Yeni Kasa Ekle")

@cari_bp.route('/kasa/duzelt/<int:id>', methods=['GET', 'POST'])
def kasa_duzelt(id):
    kasa = KasaService.get_by_id(id)
    if not kasa or kasa.is_deleted:
        flash("Kasa bulunamadı", "danger")
        return redirect(url_for('cari.kasa_listesi'))
        
    form = KasaForm(obj=kasa)

    if form.validate_on_submit():
        try:
            kasa.kasa_adi = form.kasa_adi.data
            kasa.tipi = form.tipi.data
            kasa.para_birimi = form.para_birimi.data
            kasa.banka_sube_adi = (form.banka_sube_adi.data or '').strip() or None
            KasaService.save(kasa, is_new=False, actor_id=get_actor())
            OperationLogService.log(
                module='cari', action='kasa_duzelt',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa', entity_id=id,
                description=f"Kasa #{id} ({kasa.kasa_adi}) güncellendi.",
                success=True
            )
            flash('Kasa bilgileri güncellendi.', 'success')
            return redirect(url_for('cari.kasa_listesi'))
        except ValidationError as e:
            OperationLogService.log(
                module='cari', action='kasa_duzelt',
                user_id=get_actor(), username=getattr(current_user, 'username', None),
                entity_type='Kasa', entity_id=id,
                description=f"Kasa güncelleme hatası: {str(e)}",
                success=False
            )
            flash(str(e), "warning")
            
    return render_template('cari/kasa_ekle.html', form=form, title="Kasa Düzenle")

@cari_bp.route('/kasa/sil/<int:id>', methods=['POST'])
def kasa_sil(id):
    hedef_kasa_id = request.form.get('hedef_kasa_id', type=int)
    try:
        KasaService.kasa_kapat_ve_devret(id, hedef_kasa_id, actor_id=get_actor())
        OperationLogService.log(
            module='cari', action='kasa_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='Kasa', entity_id=id,
            description=f"Kasa #{id} kapatıldı ve bakiye devredildi.",
            success=True
        )
        flash('Kasa hesabı başarıyla kapatıldı.', 'success')
    except ValidationError as e:
        OperationLogService.log(
            module='cari', action='kasa_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='Kasa', entity_id=id,
            description=f"Kasa kapatma hatası: {str(e)}",
            success=False
        )
        flash(str(e), 'warning')
    except Exception as e:
        OperationLogService.log(
            module='cari', action='kasa_sil',
            user_id=get_actor(), username=getattr(current_user, 'username', None),
            entity_type='Kasa', entity_id=id,
            description=f"Kasa kapatma hatası: {str(e)}",
            success=False
        )
        flash(f"Hata: {str(e)}", 'danger')
        
    return redirect(url_for('cari.kasa_listesi'))

@cari_bp.route('/kasa/hareketleri/<int:id>')
def kasa_hareketleri(id):
    kasa = KasaService.get_by_id(id)
    if not kasa or kasa.is_deleted:
        flash("Kasa bulunamadı", "danger")
        return redirect(url_for('cari.kasa_listesi'))

    hareketler = Odeme.query.filter_by(kasa_id=kasa.id, is_deleted=False)\
                      .order_by(Odeme.tarih.desc(), Odeme.id.desc()).all()
    return render_template('cari/kasa_hareketleri.html', kasa=kasa, hareketler=hareketler)

# -------------------------------------------------------------------------
# 4. RAPORLAR VE MENÜ
# -------------------------------------------------------------------------
@cari_bp.route('/finans-menu')
def finans_menu():
    return render_template('cari/finans_menu.html')


def _get_cari_durum_raporu_verisi(sort_by, sort_dir, q):
    if sort_by not in {'firma_adi', 'bakiye', 'bakiye_kdvli', 'durum'}:
        sort_by = 'firma_adi'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'asc'

    rapor, genel_toplam = CariRaporService.get_durum_raporu()

    if q:
        q_lower = q.casefold()
        rapor = [s for s in rapor if q_lower in (s.get('firma_adi') or '').casefold()]

    if sort_by == 'firma_adi':
        sort_key = lambda s: (s.get('firma_adi') or '').casefold()
    elif sort_by in ('bakiye', 'bakiye_kdvli'):
        key_name = 'bakiye_kdvli' if sort_by == 'bakiye_kdvli' else 'bakiye'
        sort_key = lambda s: float(s.get(key_name) or 0)
    else:
        def durum_key(s):
            bakiye = float(s.get('bakiye') or 0)
            return 'borclu' if bakiye > 0 else ('alacakli' if bakiye < 0 else 'kapali')
        sort_key = durum_key

    rapor = sorted(rapor, key=sort_key, reverse=(sort_dir == 'desc'))
    return rapor, genel_toplam, sort_by, sort_dir

@cari_bp.route('/cari-durum-raporu')
def cari_durum_raporu():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        sort_by = request.args.get('sort_by', 'firma_adi', type=str)
        sort_dir = request.args.get('sort_dir', 'asc', type=str)
        q = (request.args.get('q', '', type=str) or '').strip()

        allowed_per_page = {10, 25, 50, 100}
        allowed_sort_by = {'firma_adi', 'bakiye', 'durum'}
        allowed_sort_dir = {'asc', 'desc'}

        if per_page not in allowed_per_page:
            per_page = 25
        if sort_by not in allowed_sort_by:
            sort_by = 'firma_adi'
        if sort_dir not in allowed_sort_dir:
            sort_dir = 'asc'

        rapor, genel_toplam, sort_by, sort_dir = _get_cari_durum_raporu_verisi(sort_by, sort_dir, q)

        toplam_kayit = len(rapor)
        pagination = ListPagination(total=toplam_kayit, page=page, per_page=per_page)
        baslangic = (pagination.page - 1) * pagination.per_page
        bitis = baslangic + pagination.per_page
        rapor_sayfa = rapor[baslangic:bitis]

        return render_template(
            'cari/cari_durum_raporu.html',
            rapor=rapor_sayfa,
            genel_toplam=genel_toplam,
            pagination=pagination,
            per_page=per_page,
            sort_by=sort_by,
            sort_dir=sort_dir,
            q=q,
        )
    except ValidationError as e:
        flash(str(e), "danger")
        return redirect(url_for('cari.finans_menu'))

@cari_bp.route('/cari-durum-raporu/yazdir')
def cari_durum_raporu_yazdir():
    try:
        sort_by = request.args.get('sort_by', 'firma_adi', type=str)
        sort_dir = request.args.get('sort_dir', 'asc', type=str)
        q = (request.args.get('q', '', type=str) or '').strip()
        rapor, genel_toplam, sort_by, sort_dir = _get_cari_durum_raporu_verisi(sort_by, sort_dir, q)

        return render_template(
            'cari/cari_durum_raporu_yazdir.html',
            rapor=rapor,
            genel_toplam=genel_toplam,
            q=q,
            rapor_tarihi=date.today().strftime('%d.%m.%Y'),
        )
    except ValidationError as e:
        flash(str(e), "danger")
        return redirect(url_for('cari.cari_durum_raporu'))


@cari_bp.route('/cari-durum-raporu/excel')
def cari_durum_raporu_excel():
    try:
        sort_by = request.args.get('sort_by', 'firma_adi', type=str)
        sort_dir = request.args.get('sort_dir', 'asc', type=str)
        q = (request.args.get('q', '', type=str) or '').strip()
        rapor, genel_toplam, _, _ = _get_cari_durum_raporu_verisi(sort_by, sort_dir, q)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Cari Durum Raporu'
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        title_font = Font(name='Calibri', size=14, bold=True, color='1F1F1F')
        meta_font = Font(name='Calibri', size=10, color='44546A')
        header_font = Font(name='Calibri', size=10, bold=True, color='18324A')
        body_font = Font(name='Calibri', size=10, color='1F1F1F')
        total_font = Font(name='Calibri', size=10, bold=True, color='1F1F1F')

        header_fill = PatternFill(fill_type='solid', fgColor='DBE7F3')
        total_fill = PatternFill(fill_type='solid', fgColor='F0F0F0')

        thin_blue = Side(style='thin', color='D9E2F0')
        dark_blue = Side(style='medium', color='18324A')

        header_border = Border(top=dark_blue, bottom=dark_blue)
        cell_border = Border(bottom=thin_blue)
        total_border = Border(top=dark_blue)

        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        def set_cell(row, column, value, font=None, fill=None, border=None, alignment=None, number_format=None):
            cell = sheet.cell(row=row, column=column, value=value)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
            if number_format:
                cell.number_format = number_format
            return cell

        sheet.merge_cells('A1:G1')
        set_cell(1, 1, 'Firma Cari Durum Raporu', font=title_font, alignment=left_alignment)
        sheet.merge_cells('A2:D2')
        alt_baslik = 'Tum finansal hareketlerin ozetidir'
        if q:
            alt_baslik = f'{alt_baslik} - Filtre: "{q}"'
        set_cell(2, 1, alt_baslik, font=meta_font, alignment=left_alignment)
        sheet.merge_cells('E2:G2')
        set_cell(2, 5, f'Rapor Tarihi: {date.today().strftime("%d.%m.%Y")}', font=meta_font, alignment=left_alignment)
        sheet.merge_cells('A3:B3')
        set_cell(3, 1, f'Toplam Firma: {len(rapor)}', font=meta_font, alignment=left_alignment)
        sheet.merge_cells('C3:D3')
        set_cell(3, 3, f"Genel Toplam Borç: {float(genel_toplam.get('borc_kdvli') or 0):,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.'), font=meta_font, alignment=left_alignment)
        sheet.merge_cells('E3:F3')
        set_cell(3, 5, f"Genel Toplam Alacak: {float(genel_toplam.get('alacak_kdvli') or 0):,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.'), font=meta_font, alignment=left_alignment)
        set_cell(3, 7, f"Bakiye: {float(genel_toplam.get('bakiye_kdvli') or 0):,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.'), font=meta_font, alignment=right_alignment)

        headers = ['#', 'Firma Adı', 'Tipi', 'Toplam Borç (KDV Dahil)', 'Toplam Alacak (KDV Dahil)', 'Güncel Bakiye (KDV Dahil)', 'Durum']
        header_row = 5
        for col_idx, header in enumerate(headers, start=1):
            align = center_alignment if col_idx in {1, 3, 7} else right_alignment if col_idx in {4, 5, 6} else left_alignment
            set_cell(header_row, col_idx, header, font=header_font, fill=header_fill, border=header_border, alignment=align)

        current_row = 6
        for index, satir in enumerate(rapor, start=1):
            bakiye = float(satir.get('bakiye') or 0)
            if bakiye > 0:
                durum = 'Borçlu'
            elif bakiye < 0:
                durum = 'Alacaklı'
            else:
                durum = 'Kapalı'

            row_data = [
                index,
                satir.get('firma_adi') or '-',
                satir.get('tipi') or 'Firma',
                float(satir.get('borc_kdvli') or 0),
                float(satir.get('alacak_kdvli') or 0),
                float(satir.get('bakiye_kdvli') or 0),
                durum,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                if col_idx in {4, 5, 6}:
                    set_cell(current_row, col_idx, value, font=body_font, border=cell_border, alignment=right_alignment, number_format='#,##0.00')
                else:
                    align = center_alignment if col_idx in {1, 3, 7} else left_alignment
                    set_cell(current_row, col_idx, value, font=body_font, border=cell_border, alignment=align)
            current_row += 1

        if rapor:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            set_cell(current_row, 1, 'GENEL TOPLAM:', font=total_font, fill=total_fill, border=total_border, alignment=right_alignment)
            set_cell(current_row, 4, float(genel_toplam.get('borc_kdvli') or 0), font=total_font, fill=total_fill, border=total_border, alignment=right_alignment, number_format='#,##0.00')
            set_cell(current_row, 5, float(genel_toplam.get('alacak_kdvli') or 0), font=total_font, fill=total_fill, border=total_border, alignment=right_alignment, number_format='#,##0.00')
            set_cell(current_row, 6, float(genel_toplam.get('bakiye_kdvli') or 0), font=total_font, fill=total_fill, border=total_border, alignment=right_alignment, number_format='#,##0.00')
            set_cell(current_row, 7, '', font=total_font, fill=total_fill, border=total_border, alignment=center_alignment)

        column_widths = {
            'A': 6,
            'B': 34,
            'C': 14,
            'D': 18,
            'E': 18,
            'F': 18,
            'G': 12,
        }
        for column_letter, width in column_widths.items():
            sheet.column_dimensions[column_letter].width = width

        for row_idx in range(6, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = 28

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        file_name = f'cari_durum_raporu_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except ValidationError as e:
        flash(str(e), 'danger')
        return redirect(url_for('cari.cari_durum_raporu'))
