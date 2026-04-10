from io import BytesIO

from flask import render_template, redirect, url_for, flash, request, send_file
from flask_login import current_user
from app.extensions import db
from app.nakliyeler import nakliye_bp
from app.nakliyeler.models import Nakliye
from app.nakliyeler.forms import NakliyeForm
from app.services.nakliye_services import CariServis
from app.services.operation_log_service import OperationLogService
from app.firmalar.models import Firma
from app.araclar.models import Arac
from decimal import Decimal, InvalidOperation
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _get_nakliye_araclari():
    return Arac.aktif_nakliye_query().order_by(Arac.plaka).all()

def _actor():
    return current_user.id if current_user.is_authenticated else None

def _uname():
    return getattr(current_user, 'username', None)

# -------------------------------------------------------------------------
# YARDIMCI FONKSİYON: Decimal Hata Çözücü
# -------------------------------------------------------------------------
def to_decimal(value):
    """
    Formlardan veya veritabanından gelen karmaşık sayı formatlarını 
    güvenli bir şekilde Decimal nesnesine çevirir.
    """
    if value is None or value == '':
        return Decimal('0.00')
    if isinstance(value, Decimal):
        return value
    try:
        # Virgüllü formatı (1.250,50) Python'un anlayacağı (1250.50) formatına çevir
        clean_val = str(value).replace('.', '').replace(',', '.')
        return Decimal(clean_val)
    except (InvalidOperation, ValueError):
        return Decimal('0.00')


def _nakliye_filtered_query(baslangic, bitis, secili_plaka, secili_taseron_id, secili_firma_id):
    query = Nakliye.query.filter(Nakliye.tutar > 0)

    if baslangic:
        try:
            query = query.filter(Nakliye.tarih >= datetime.strptime(baslangic, '%Y-%m-%d').date())
        except ValueError:
            pass
    if bitis:
        try:
            query = query.filter(Nakliye.tarih <= datetime.strptime(bitis, '%Y-%m-%d').date())
        except ValueError:
            pass
    if secili_plaka:
        query = query.filter(Nakliye.plaka == secili_plaka)
    if secili_taseron_id and secili_taseron_id.isdigit():
        query = query.filter(Nakliye.taseron_firma_id == int(secili_taseron_id))
    if secili_firma_id and secili_firma_id.isdigit():
        query = query.filter(Nakliye.firma_id == int(secili_firma_id))

    return query


def _nakliye_kdv_orani(nakliye):
    if nakliye.kiralama_id and nakliye.kiralama and nakliye.kiralama.kalemler:
        return nakliye.kiralama.kalemler[0].nakliye_satis_kdv or 0
    return nakliye.kdv_orani or 0


def _turkce_tarih(iso_date):
    if iso_date and len(iso_date) == 10 and '-' in iso_date:
        return f"{iso_date[8:10]}.{iso_date[5:7]}.{iso_date[0:4]}"
    return iso_date

# ---------------------------------------------------
# 1. NAKLİYE SEFER LİSTESİ (Filtreleme)
# ---------------------------------------------------
@nakliye_bp.route('/')
def index():
    # Filtreleme parametrelerini yakala
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    if per_page not in {10, 25, 50, 100}:
        per_page = 25

    bugun = date.today()

    baslangic_explicit = request.args.get('baslangic')
    bitis_explicit = request.args.get('bitis')
    secili_plaka = request.args.get('plaka')
    secili_taseron_id = request.args.get('taseron_id')
    secili_firma_id = request.args.get('firma_id')

    # Firma/taşeron/plaka filtresi aktifken varsayılan tarih aralığı uygulanmaz;
    # böylece eski tarihli kiralama bağlantılı nakliyeler de görünür.
    spesifik_filtre_var = bool(secili_firma_id or secili_taseron_id or secili_plaka)
    if spesifik_filtre_var:
        baslangic = baslangic_explicit  # None ise tarih kısıtı yok
        bitis = bitis_explicit
    else:
        baslangic = baslangic_explicit or (bugun - timedelta(days=15)).isoformat()
        bitis = bitis_explicit or bugun.isoformat()

    try:
        query = Nakliye.query.filter(Nakliye.tutar > 0)

        # Filtreleri uygula
        if baslangic:
            try:
                baslangic_date = datetime.strptime(baslangic, '%Y-%m-%d').date()
                query = query.filter(Nakliye.tarih >= baslangic_date)
            except ValueError: pass
        if bitis:
            try:
                bitis_date = datetime.strptime(bitis, '%Y-%m-%d').date()
                query = query.filter(Nakliye.tarih <= bitis_date)
            except ValueError: pass
        if secili_plaka:
            query = query.filter(Nakliye.plaka == secili_plaka)
        if secili_taseron_id and secili_taseron_id.isdigit():
            query = query.filter(Nakliye.taseron_firma_id == int(secili_taseron_id))
        if secili_firma_id and secili_firma_id.isdigit():
            query = query.filter(Nakliye.firma_id == int(secili_firma_id))

        ordered_query = query.order_by(Nakliye.tarih.desc())
        pagination = ordered_query.paginate(page=page, per_page=per_page, error_out=False)
        nakliyeler = pagination.items
        filtered_all = ordered_query.all()

        # Dropdown listelerini hazırla
        plakalar = db.session.query(Nakliye.plaka).filter(Nakliye.plaka.isnot(None)).distinct().all()
        plaka_listesi = [p[0] for p in plakalar if p[0] and p[0].strip() != ""]
        taseron_listesi = Firma.query.filter_by(is_tedarikci=True, is_active=True).order_by(Firma.firma_adi).all()
        firma_listesi = [{'id': f.id, 'firma_adi': f.firma_adi}
                         for f in Firma.query.filter_by(is_active=True).order_by(Firma.firma_adi).all()]

        # İstatistikler
        stats = {
            'sefer_sayisi': len(filtered_all),
            'ciro': sum(n.toplam_tutar or 0 for n in filtered_all),
            'maliyet': sum(n.taseron_maliyet or 0 for n in filtered_all),
            'kar': sum(n.tahmini_kar or 0 for n in filtered_all)
        }

        return render_template('nakliyeler/index.html',
                               nakliyeler=nakliyeler,
                               pagination=pagination,
                               per_page=per_page,
                               stats=stats,
                               baslangic=baslangic,
                               bitis=bitis,
                               plaka_listesi=plaka_listesi,
                               taseron_listesi=taseron_listesi,
                               firma_listesi=firma_listesi,
                               secili_taseron_id=secili_taseron_id,
                               secili_plaka=secili_plaka,
                               secili_firma_id=secili_firma_id)
    except Exception as e:
        db.session.rollback()
        from flask import current_app
        current_app.logger.error(f"Nakliye Liste Yükleme Hatası: {str(e)}")
        flash("Liste yüklenirken bir hata oluştu.", "danger")
        return render_template('nakliyeler/index.html',
                               nakliyeler=[], pagination=None, per_page=per_page,
                               stats={'sefer_sayisi': 0, 'ciro': 0, 'maliyet': 0, 'kar': 0},
                               baslangic=baslangic, bitis=bitis,
                               plaka_listesi=[], taseron_listesi=[], firma_listesi=[],
                               secili_taseron_id=secili_taseron_id,
                               secili_plaka=secili_plaka,
                               secili_firma_id=secili_firma_id)


# ---------------------------------------------------
# YAZDIR: Filtrelenmiş TÜM Nakliye Kayıtları (Sayfalama Yok)
# ---------------------------------------------------
@nakliye_bp.route('/yazdir')
def yazdir():
    bugun = date.today()

    baslangic = request.args.get('baslangic')
    bitis = request.args.get('bitis')
    secili_plaka = request.args.get('plaka')
    secili_taseron_id = request.args.get('taseron_id')
    secili_firma_id = request.args.get('firma_id')

    if not (secili_firma_id or secili_taseron_id or secili_plaka):
        baslangic = baslangic or (bugun - timedelta(days=15)).isoformat()
        bitis = bitis or bugun.isoformat()

    query = _nakliye_filtered_query(baslangic, bitis, secili_plaka, secili_taseron_id, secili_firma_id)
    nakliyeler = query.order_by(Nakliye.tarih.desc()).all()

    stats = {
        'sefer_sayisi': len(nakliyeler),
        'ciro': sum(n.toplam_tutar or 0 for n in nakliyeler),
        'maliyet': sum(n.taseron_maliyet or 0 for n in nakliyeler),
        'kar': sum(n.tahmini_kar or 0 for n in nakliyeler)
    }

    return render_template('nakliyeler/yazdir.html',
                           nakliyeler=nakliyeler,
                           stats=stats,
                           baslangic=baslangic,
                           bitis=bitis,
                           secili_plaka=secili_plaka,
                           secili_taseron_id=secili_taseron_id,
                           secili_firma_id=secili_firma_id,
                           rapor_tarihi=date.today().strftime('%d.%m.%Y'))


@nakliye_bp.route('/excel')
def excel_aktar():
    bugun = date.today()

    baslangic = request.args.get('baslangic')
    bitis = request.args.get('bitis')
    secili_plaka = request.args.get('plaka')
    secili_taseron_id = request.args.get('taseron_id')
    secili_firma_id = request.args.get('firma_id')

    if not (secili_firma_id or secili_taseron_id or secili_plaka):
        baslangic = baslangic or (bugun - timedelta(days=15)).isoformat()
        bitis = bitis or bugun.isoformat()

    nakliyeler = _nakliye_filtered_query(
        baslangic,
        bitis,
        secili_plaka,
        secili_taseron_id,
        secili_firma_id,
    ).order_by(Nakliye.tarih.desc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Nakliye'
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    title_font = Font(name='Calibri', size=14, bold=True, color='1F1F1F')
    meta_font = Font(name='Calibri', size=10, color='44546A')
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    body_font = Font(name='Calibri', size=10, color='1F1F1F')
    total_font = Font(name='Calibri', size=10, bold=True, color='1F1F1F')

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    total_fill = PatternFill(fill_type='solid', fgColor='EDF3F8')

    thin_blue = Side(style='thin', color='D9E2F0')
    thin_header = Side(style='thin', color='1F4E78')
    medium_total = Side(style='medium', color='9FBAD0')

    header_border = Border(top=thin_header, bottom=thin_header, left=thin_header, right=thin_header)
    cell_border = Border(bottom=thin_blue)
    total_border = Border(top=medium_total)

    left_alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
    center_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    right_alignment = Alignment(vertical='center', horizontal='right')

    sheet.merge_cells('A1:J1')
    sheet['A1'] = 'Nakliye Sefer Listesi'
    sheet['A1'].font = title_font
    sheet['A1'].alignment = left_alignment

    sheet.merge_cells('A2:F2')
    sheet['A2'] = f"Tarih Araligi: {_turkce_tarih(baslangic)} - {_turkce_tarih(bitis)}"
    sheet['A2'].font = meta_font
    sheet['A2'].alignment = left_alignment

    report_text = f"Rapor Tarihi: {date.today().strftime('%d.%m.%Y')}"
    if secili_plaka:
        sheet.merge_cells('G2:H2')
        sheet['G2'] = report_text
        sheet['G2'].font = meta_font
        sheet['G2'].alignment = left_alignment
        sheet.merge_cells('I2:J2')
        sheet['I2'] = f"Plaka: {secili_plaka}"
        sheet['I2'].font = meta_font
        sheet['I2'].alignment = left_alignment
    else:
        sheet.merge_cells('G2:J2')
        sheet['G2'] = report_text
        sheet['G2'].font = meta_font
        sheet['G2'].alignment = left_alignment

    headers = ['#', 'Tarih', 'Müşteri Firma', 'Güzergah / Açıklama', 'Plaka / Tedarikçi', 'Matrah', 'KDV %', 'KDV', 'Tutar', 'Durum']
    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_alignment if col_idx in {1, 2, 7, 10} else left_alignment

    current_row = 5
    toplam_matrah = 0.0
    toplam_kdv = 0.0
    toplam_tutar = 0.0

    for index, nakliye in enumerate(nakliyeler, start=1):
        kdv_orani = float(_nakliye_kdv_orani(nakliye) or 0)
        matrah = float(nakliye.tutar or 0)
        kdv = matrah * (kdv_orani / 100.0)
        tutar = matrah + kdv

        toplam_matrah += matrah
        toplam_kdv += kdv
        toplam_tutar += tutar

        guzergah = nakliye.guzergah or ''
        if nakliye.aciklama:
            guzergah = f"{guzergah}\n{nakliye.aciklama}" if guzergah else nakliye.aciklama

        if nakliye.nakliye_tipi == 'taseron' and nakliye.taseron_firma:
            plaka_tedarikci = nakliye.taseron_firma.firma_adi or ''
            if nakliye.plaka:
                plaka_tedarikci = f"{plaka_tedarikci}\n{nakliye.plaka}"
        else:
            plaka_tedarikci = nakliye.plaka or '-'

        row_data = [
            index,
            nakliye.tarih.strftime('%d.%m.%Y') if nakliye.tarih else '',
            nakliye.firma.firma_adi if nakliye.firma else '-',
            guzergah,
            plaka_tedarikci,
            matrah,
            kdv_orani,
            kdv,
            tutar,
            'Islendi' if nakliye.cari_islendi_mi else 'Bekliyor',
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=current_row, column=col_idx, value=value)
            cell.font = body_font
            cell.border = cell_border

            if col_idx in {6, 7, 8, 9}:
                cell.alignment = right_alignment
                if col_idx == 7:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '#,##0.00'
            elif col_idx in {1, 2, 10}:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

        current_row += 1

    if nakliyeler:
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        total_label = sheet.cell(row=current_row, column=1, value='TOPLAM:')
        total_label.font = total_font
        total_label.fill = total_fill
        total_label.border = total_border
        total_label.alignment = right_alignment

        for col_idx in range(2, 6):
            merged_cell = sheet.cell(row=current_row, column=col_idx)
            merged_cell.fill = total_fill
            merged_cell.border = total_border

        total_values = {
            6: toplam_matrah,
            8: toplam_kdv,
            9: toplam_tutar,
        }
        for col_idx in range(6, 11):
            cell = sheet.cell(row=current_row, column=col_idx)
            cell.fill = total_fill
            cell.border = total_border
            cell.font = total_font
            cell.alignment = right_alignment if col_idx in {6, 8, 9} else center_alignment
            if col_idx in total_values:
                cell.value = total_values[col_idx]
                cell.number_format = '#,##0.00'

    column_widths = {
        'A': 6,
        'B': 14,
        'C': 28,
        'D': 34,
        'E': 22,
        'F': 14,
        'G': 10,
        'H': 14,
        'I': 14,
        'J': 12,
    }
    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    for row_idx in range(5, current_row + 1):
        sheet.row_dimensions[row_idx].height = 34

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    file_name = f"nakliye_listesi_{bugun.strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ---------------------------------------------------
# 2. YENİ NAKLİYE KAYDI (Otomasyonlu)
# ---------------------------------------------------
@nakliye_bp.route('/ekle', methods=['GET', 'POST'])
def ekle():
    form = NakliyeForm()

    # Form seçeneklerini dinamik doldur
    try:
        firmalar = Firma.query.filter_by(is_active=True).order_by(Firma.firma_adi).all()
        form.firma_id.choices = [(0, '-- Seçiniz --')] + [(f.id, f.firma_adi) for f in firmalar]
        form.taseron_firma_id.choices = [(0, '--- Taşeron Seçiniz ---')] + [(f.id, f.firma_adi) for f in firmalar if f.is_tedarikci]

        araclar = _get_nakliye_araclari()
        form.arac_id.choices = [(0, '--- Dış Nakliye / Belirtilmemiş ---')] + [(a.id, a.plaka) for a in araclar]
    except Exception as e:
        db.session.rollback()
        from flask import current_app
        current_app.logger.error(f"Nakliye Ekle seçenek yükleme hatası: {str(e)}")
        flash("Form seçenekleri yüklenemedi. Lütfen tekrar deneyin.", "danger")
        return redirect(url_for('nakliyeler.index'))

    kiralama_id = request.args.get('kiralama_id', type=int)
    kdv_orani = request.args.get('kdv_orani', type=int)

    # KDV oranını form varsayılanına ayarla
    if kdv_orani:
        form.kdv_orani.data = kdv_orani

    if form.validate_on_submit():
        try:
            nakliye = Nakliye()
            form.populate_obj(nakliye)
            nakliye.tevkifat_orani = form.tevkifat_orani.data or None

            if not nakliye.arac_id or nakliye.arac_id <= 0:
                nakliye.arac_id = None
            if not nakliye.taseron_firma_id or nakliye.taseron_firma_id <= 0:
                nakliye.taseron_firma_id = None
            
            # Öz mal araç seçildiyse plaka senkronize et
            if nakliye.nakliye_tipi == 'oz_mal':
                nakliye.taseron_firma_id = None
                if nakliye.arac_id:
                    secili_arac = Arac.aktif_nakliye_query().filter(Arac.id == nakliye.arac_id).first()
                    if secili_arac:
                        nakliye.plaka = secili_arac.plaka
                    else:
                        nakliye.arac_id = None
                        nakliye.plaka = None
                else:
                    nakliye.plaka = None
            else:
                nakliye.arac_id = None

            # Decimal dönüşümleri
            nakliye.tutar = to_decimal(form.tutar.data)
            nakliye.taseron_maliyet = to_decimal(form.taseron_maliyet.data) if nakliye.nakliye_tipi == 'taseron' else Decimal('0.00')
            if kiralama_id: nakliye.kiralama_id = kiralama_id
            
            nakliye.hesapla_ve_guncelle()

            # Veritabanına ekle
            db.session.add(nakliye)
            db.session.flush() 
            
            # --- 🚀 CARİ SERVİS OTOMASYONU ---
            CariServis.musteri_nakliye_senkronize_et(nakliye)
            CariServis.taseron_maliyet_senkronize_et(nakliye)
            
            db.session.commit()
            OperationLogService.log(
                module='nakliyeler', action='create',
                user_id=_actor(), username=_uname(),
                entity_type='Nakliye', entity_id=nakliye.id,
                description=f"Nakliye seferi eklendi (#{nakliye.id}).",
                success=True
            )
            flash('Nakliye seferi ve bağlı cari kayıtlar başarıyla oluşturuldu.', 'success')
            return redirect(url_for('nakliyeler.index'))
            
        except Exception as e:
            db.session.rollback()
            OperationLogService.log(
                module='nakliyeler', action='create',
                user_id=_actor(), username=_uname(),
                entity_type='Nakliye',
                description=f"Nakliye ekleme hatası: {str(e)}",
                success=False
            )
            flash(f'Kayıt hatası: {str(e)}', 'danger')

    return render_template('nakliyeler/ekle.html', form=form)

# ---------------------------------------------------
# 3. ARAÇ YÖNETİMİ (Lojistik Parkı)
# ---------------------------------------------------
@nakliye_bp.route('/arac/liste')
def arac_liste():
    araclar = Arac.query.all()
    return render_template('nakliyeler/arac_liste.html', araclar=araclar)

@nakliye_bp.route('/arac/ekle', methods=['GET', 'POST'])
def arac_ekle():
    from .forms import AracForm # Döngüsel importu önlemek için
    from app.utils import normalize_turkish_upper
    form = AracForm()
    if form.validate_on_submit():
        try:
            plaka_norm = normalize_turkish_upper(form.plaka.data)
            mevcut = Arac.query.filter_by(plaka=plaka_norm).first()
            if mevcut:
                flash(f"{form.plaka.data} zaten kayıtlı!", "danger")
            else:
                yeni_arac = Arac(
                    plaka=plaka_norm,
                    arac_tipi=form.arac_tipi.data,
                    marka_model=form.marka_model.data
                )
                db.session.add(yeni_arac)
                db.session.commit()
                flash('Araç eklendi.', 'success')
                return redirect(url_for('nakliyeler.arac_liste'))
        except Exception as e:
            db.session.rollback()
            flash(f'Hata: {str(e)}', 'danger')
    return render_template('nakliyeler/arac_ekle.html', form=form)

# ---------------------------------------------------
# 4. DÜZENLEME (Otomasyonlu & Kilitli)
# ---------------------------------------------------
@nakliye_bp.route('/duzenle/<int:id>', methods=['GET', 'POST'])
def duzenle(id):
    nakliye = Nakliye.query.get_or_404(id)

    if nakliye.kiralama_id:
        flash('Bu kayıt kiralama modülüne bağlıdır.', 'warning')
        return redirect(url_for('nakliyeler.index'))

    form = NakliyeForm(obj=nakliye)

    # Seçenekleri doldur
    try:
        firmalar = Firma.query.filter_by(is_active=True).all()
        form.firma_id.choices = [(0, '-- Seçiniz --')] + [(f.id, f.firma_adi) for f in firmalar]
        form.taseron_firma_id.choices = [(0, '--- Taşeron Seçiniz ---')] + [(f.id, f.firma_adi) for f in firmalar if f.is_tedarikci]

        araclar = _get_nakliye_araclari()
        form.arac_id.choices = [(0, '--- Dış Nakliye / Belirtilmemiş ---')] + [(a.id, a.plaka) for a in araclar]
    except Exception as e:
        db.session.rollback()
        from flask import current_app
        current_app.logger.error(f"Nakliye Düzenle seçenek yükleme hatası: {str(e)}")
        flash("Form seçenekleri yüklenemedi. Lütfen tekrar deneyin.", "danger")
        return redirect(url_for('nakliyeler.index'))

    if form.validate_on_submit():
        try:
            form.populate_obj(nakliye)
            nakliye.tevkifat_orani = form.tevkifat_orani.data or None
            if not nakliye.arac_id or nakliye.arac_id <= 0:
                nakliye.arac_id = None
            if not nakliye.taseron_firma_id or nakliye.taseron_firma_id <= 0:
                nakliye.taseron_firma_id = None

            if nakliye.nakliye_tipi == 'oz_mal':
                nakliye.taseron_firma_id = None
                if nakliye.arac_id:
                    secili_arac = Arac.aktif_nakliye_query().filter(Arac.id == nakliye.arac_id).first()
                    if secili_arac:
                        nakliye.plaka = secili_arac.plaka
                    else:
                        nakliye.arac_id = None
                        nakliye.plaka = None
                else:
                    nakliye.plaka = None
            else:
                nakliye.arac_id = None

            nakliye.tutar = to_decimal(form.tutar.data)
            nakliye.taseron_maliyet = to_decimal(form.taseron_maliyet.data) if nakliye.nakliye_tipi == 'taseron' else Decimal('0.00')

            nakliye.hesapla_ve_guncelle()

            CariServis.musteri_nakliye_senkronize_et(nakliye)
            CariServis.taseron_maliyet_senkronize_et(nakliye)

            db.session.commit()
            OperationLogService.log(
                module='nakliyeler', action='update',
                user_id=_actor(), username=_uname(),
                entity_type='Nakliye', entity_id=nakliye.id,
                description=f"Nakliye #{nakliye.id} güncellendi.",
                success=True
            )
            flash('Kayıt güncellendi.', 'success')
            return redirect(url_for('nakliyeler.index'))

        except Exception as e:
            db.session.rollback()
            OperationLogService.log(
                module='nakliyeler', action='update',
                user_id=_actor(), username=_uname(),
                entity_type='Nakliye', entity_id=id,
                description=f"Nakliye güncelleme hatası: {str(e)}",
                success=False
            )
            flash(f'Hata: {str(e)}', 'danger')
            # Rollback sonrası nakliye objesi expire oldu, yeniden yükle
            nakliye = db.session.get(Nakliye, id)

    return render_template('nakliyeler/duzenle.html', form=form, nakliye=nakliye)

# ---------------------------------------------------
# 5. SİLME (Tam Temizlik)
# ---------------------------------------------------
@nakliye_bp.route('/sil/<int:id>', methods=['POST'])
def sil(id):
    nakliye = Nakliye.query.get_or_404(id)
    
    if nakliye.kiralama_id:
        flash('Kiralama bağlantılı kayıtlar silinemez.', 'danger')
        return redirect(url_for('nakliyeler.index'))
    
    try:
        CariServis.nakliye_cari_temizle(nakliye.id)
        db.session.delete(nakliye)
        db.session.commit()
        OperationLogService.log(
            module='nakliyeler', action='delete',
            user_id=_actor(), username=_uname(),
            entity_type='Nakliye', entity_id=id,
            description=f"Nakliye #{id} silindi.",
            success=True
        )
        flash('Kayıt silindi.', 'success')
    except Exception as e:
        db.session.rollback()
        OperationLogService.log(
            module='nakliyeler', action='delete',
            user_id=_actor(), username=_uname(),
            entity_type='Nakliye', entity_id=id,
            description=f"Nakliye silme hatası: {str(e)}",
            success=False
        )
        flash(f'Hata: {str(e)}', 'danger')
        
    return redirect(url_for('nakliyeler.index'))

@nakliye_bp.route('/detay/<int:id>')
def detay(id):
    nakliye = Nakliye.query.get_or_404(id)
    return render_template('nakliyeler/detay.html', nakliye=nakliye)