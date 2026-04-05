from io import BytesIO

from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from sqlalchemy import or_
from sqlalchemy.orm import joinedload, subqueryload
from decimal import Decimal
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.extensions import db
from app.servis import servis_bp
from app.filo.models import BakimKaydi, KullanilanParca, Ekipman, StokKarti, StokHareket
from app.firmalar.models import Firma
from app.personel.models import Personel
from app.services.base import ValidationError
from app.services.filo_services import BakimService, EkipmanService
from app.services.operation_log_service import OperationLogService


def get_actor_id():
    return current_user.id if current_user.is_authenticated else None


BAKIM_TIPI_LABELS = {
    'ariza': 'Ariza',
    'periyodik': 'Periyodik Bakim',
    'genel_kontrol': 'Genel Kontrol',
}

SERVIS_TIPI_LABELS = {
    'ic_servis': 'Ic Servis',
    'dis_servis': 'Dis Servis',
    'yerinde_servis': 'Yerinde Servis',
}

DURUM_LABELS = {
    'acik': 'Acik',
    'parca_bekliyor': 'Parca Bekliyor',
    'tamamlandi': 'Tamamlandi',
    'iptal': 'Iptal',
}


def _get_actor_id():
    if getattr(current_user, 'is_authenticated', False):
        return current_user.id
    return None


def _collect_parts_from_form(form):
    stok_karti_ids = form.getlist('parca_stok_karti_id')
    malzeme_adlari = form.getlist('parca_malzeme_adi')
    kullanilan_adetler = form.getlist('parca_kullanilan_adet')
    birim_fiyatlar = form.getlist('parca_birim_fiyat')
    parts = []

    for stok_karti_id, malzeme_adi, kullanilan_adet, birim_fiyat in zip(stok_karti_ids, malzeme_adlari, kullanilan_adetler, birim_fiyatlar):
        stok_karti_id = (stok_karti_id or '').strip()
        malzeme_adi = (malzeme_adi or '').strip()
        kullanilan_adet = (kullanilan_adet or '').strip()
        birim_fiyat = (birim_fiyat or '').strip()
        if not stok_karti_id and not malzeme_adi and not kullanilan_adet and not birim_fiyat:
            continue

        parts.append({
            'stok_karti_id': stok_karti_id,
            'malzeme_adi': malzeme_adi,
            'kullanilan_adet': kullanilan_adet,
            'birim_fiyat': birim_fiyat,
        })

    return parts


def _common_form_context():
    return {
        'bakim_tipi_labels': BAKIM_TIPI_LABELS,
        'servis_tipi_labels': SERVIS_TIPI_LABELS,
        'durum_labels': DURUM_LABELS,
        'stok_kartlari': StokKarti.query.filter_by(is_deleted=False).order_by(StokKarti.parca_adi).all(),
        'servis_firmalari': Firma.query.filter_by(is_tedarikci=True).order_by(Firma.firma_adi).all(),
        'personel_listesi': Personel.query.filter(
            Personel.is_deleted == False,
            Personel.isten_cikis_tarihi == None
        ).order_by(Personel.ad, Personel.soyad).all(),
    }


def _resolve_part_unit_price(parca):
    if parca.birim_fiyat is not None:
        return Decimal(parca.birim_fiyat or 0)

    if not parca.stok_karti_id:
        return Decimal('0')

    hareket = StokHareket.query.filter(
        StokHareket.stok_karti_id == parca.stok_karti_id,
        StokHareket.hareket_tipi == 'giris',
    ).order_by(StokHareket.tarih.desc(), StokHareket.id.desc()).first()
    return Decimal(hareket.birim_fiyat or 0) if hareket else Decimal('0')


def _attach_cost_totals(kayitlar):
    for kayit in kayitlar:
        malzeme_maliyeti = Decimal('0')
        for parca in kayit.kullanilan_parcalar:
            malzeme_maliyeti += Decimal(parca.kullanilan_adet or 0) * _resolve_part_unit_price(parca)

        kayit.malzeme_maliyeti = malzeme_maliyeti
        kayit.toplam_servis_maliyeti = Decimal(kayit.toplam_iscilik_maliyeti or 0) + malzeme_maliyeti


@servis_bp.route('/')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    if per_page not in {10, 25, 50, 100}:
        per_page = 25

    q = request.args.get('q', '', type=str).strip()
    durum = request.args.get('durum', '', type=str).strip()

    query = BakimKaydi.query.filter(BakimKaydi.is_deleted == False).options(
        joinedload(BakimKaydi.ekipman).joinedload(Ekipman.sube),
        joinedload(BakimKaydi.servis_veren_firma),
        joinedload(BakimKaydi.kullanilan_parcalar).joinedload(KullanilanParca.stok_karti),
    )

    if q:
        term = f'%{q}%'
        query = query.join(BakimKaydi.ekipman).outerjoin(BakimKaydi.servis_veren_firma).filter(
            or_(
                Ekipman.kod.ilike(term),
                Ekipman.marka.ilike(term),
                Ekipman.model.ilike(term),
                BakimKaydi.aciklama.ilike(term),
                BakimKaydi.servis_veren_kisi.ilike(term),
                Firma.firma_adi.ilike(term),
            )
        ).distinct()

    if durum:
        query = query.filter(BakimKaydi.durum == durum)

    pagination = query.order_by(BakimKaydi.tarih.desc(), BakimKaydi.id.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )
    _attach_cost_totals(pagination.items)

    acik_servis_sayisi = BakimKaydi.query.filter(
        BakimKaydi.is_deleted == False,
        BakimKaydi.durum.in_(['acik', 'parca_bekliyor']),
    ).count()
    planli_bakim_sayisi = BakimKaydi.query.filter(
        BakimKaydi.is_deleted == False,
        BakimKaydi.bakim_tipi == 'periyodik',
    ).count()

    return render_template(
        'servis/index.html',
        kayitlar=pagination.items,
        pagination=pagination,
        per_page=per_page,
        q=q,
        durum=durum,
        acik_servis_sayisi=acik_servis_sayisi,
        planli_bakim_sayisi=planli_bakim_sayisi,
        bakim_tipi_labels=BAKIM_TIPI_LABELS,
        servis_tipi_labels=SERVIS_TIPI_LABELS,
        durum_labels=DURUM_LABELS,
    )


def _servis_filtered_query(q, durum, eager=True):
    opts = []
    if eager:
        opts = [
            joinedload(BakimKaydi.ekipman).joinedload(Ekipman.sube),
            joinedload(BakimKaydi.servis_veren_firma),
            joinedload(BakimKaydi.kullanilan_parcalar).joinedload(KullanilanParca.stok_karti),
        ]
    query = BakimKaydi.query.filter(BakimKaydi.is_deleted == False)
    if opts:
        query = query.options(*opts)
    if q:
        term = f'%{q}%'
        query = query.join(BakimKaydi.ekipman).outerjoin(BakimKaydi.servis_veren_firma).filter(
            or_(
                Ekipman.kod.ilike(term),
                Ekipman.marka.ilike(term),
                Ekipman.model.ilike(term),
                BakimKaydi.aciklama.ilike(term),
                BakimKaydi.servis_veren_kisi.ilike(term),
                Firma.firma_adi.ilike(term),
            )
        ).distinct()
    if durum:
        query = query.filter(BakimKaydi.durum == durum)
    return query


@servis_bp.route('/yazdir')
@login_required
def yazdir():
    q = request.args.get('q', '', type=str).strip()
    durum = request.args.get('durum', '', type=str).strip()

    kayitlar = _servis_filtered_query(q, durum).order_by(
        BakimKaydi.tarih.desc(), BakimKaydi.id.desc()
    ).all()
    _attach_cost_totals(kayitlar)

    return render_template(
        'servis/yazdir.html',
        kayitlar=kayitlar,
        q=q,
        durum=durum,
        bakim_tipi_labels=BAKIM_TIPI_LABELS,
        servis_tipi_labels=SERVIS_TIPI_LABELS,
        durum_labels=DURUM_LABELS,
        rapor_tarihi=date.today().strftime('%d.%m.%Y'),
    )


@servis_bp.route('/excel')
@login_required
def excel():
    q = request.args.get('q', '', type=str).strip()
    durum = request.args.get('durum', '', type=str).strip()

    kayitlar = _servis_filtered_query(q, durum).order_by(
        BakimKaydi.tarih.desc(), BakimKaydi.id.desc()
    ).all()
    _attach_cost_totals(kayitlar)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Servis Kayitlari'
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    title_font  = Font(name='Calibri', size=14, bold=True,  color='1F1F1F')
    meta_font   = Font(name='Calibri', size=10,              color='44546A')
    header_font = Font(name='Calibri', size=10, bold=True,  color='FFFFFF')
    body_font   = Font(name='Calibri', size=10,              color='1F1F1F')
    total_font  = Font(name='Calibri', size=10, bold=True,  color='1F1F1F')

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    total_fill  = PatternFill(fill_type='solid', fgColor='EDF3F8')

    thin_blue   = Side(style='thin',   color='D9E2F0')
    thin_header = Side(style='thin',   color='1F4E78')
    medium_tot  = Side(style='medium', color='9FBAD0')

    header_border = Border(top=thin_header, bottom=thin_header, left=thin_header, right=thin_header)
    cell_border   = Border(bottom=thin_blue)
    total_border  = Border(top=medium_tot)

    left_align   = Alignment(vertical='center', horizontal='left',   wrap_text=True)
    center_align = Alignment(vertical='center', horizontal='center', wrap_text=True)
    right_align  = Alignment(vertical='center', horizontal='right')

    NUM_COLS = 12
    last_col  = 'L'

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'Servis / Bakım Kayıtları'
    ws['A1'].font      = title_font
    ws['A1'].alignment = left_align

    ws.merge_cells('A2:F2')
    filter_text = ('Durum: ' + DURUM_LABELS.get(durum, durum)) if durum else 'Tümü'
    if q:
        filter_text += f' | Arama: {q}'
    ws['A2'] = f'Filtre: {filter_text}'
    ws['A2'].font      = meta_font
    ws['A2'].alignment = left_align

    ws.merge_cells('G2:L2')
    ws['G2'] = f"Rapor Tarihi: {date.today().strftime('%d.%m.%Y')}"
    ws['G2'].font      = meta_font
    ws['G2'].alignment = left_align

    headers = [
        '#', 'Tarih', 'Makine Kodu', 'Marka / Model',
        'Bakım Tipi', 'Servis Tipi', 'Çalışma Saati',
        'Kim Yaptı', 'Kullanılan Parçalar',
        'İşçilik', 'Malzeme', 'Toplam',
    ]
    center_cols = {1, 2, 7}
    right_cols  = {10, 11, 12}

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = header_border
        cell.alignment = center_align if col_idx in center_cols else (right_align if col_idx in right_cols else left_align)

    current_row  = 5
    toplam_iscilik = Decimal('0')
    toplam_malzeme = Decimal('0')
    toplam_toplam  = Decimal('0')

    for idx, kayit in enumerate(kayitlar, start=1):
        parcalar_text = ', '.join(
            (p.stok_karti.parca_adi if p.stok_karti else p.malzeme_adi or 'Parça') +
            f' x{p.kullanilan_adet}'
            for p in kayit.kullanilan_parcalar
        ) or '-'

        kim_yapti = ''
        if kayit.servis_veren_firma:
            kim_yapti = kayit.servis_veren_firma.firma_adi or ''
        if kayit.servis_veren_kisi:
            kim_yapti = (kim_yapti + ' / ' + kayit.servis_veren_kisi).strip(' /')

        iscilik  = Decimal(kayit.toplam_iscilik_maliyeti or 0)
        malzeme  = kayit.malzeme_maliyeti
        toplam   = kayit.toplam_servis_maliyeti
        toplam_iscilik += iscilik
        toplam_malzeme += malzeme
        toplam_toplam  += toplam

        row_data = [
            idx,
            kayit.tarih.strftime('%d.%m.%Y') if kayit.tarih else '',
            kayit.ekipman.kod if kayit.ekipman else '-',
            f"{kayit.ekipman.marka or ''} {kayit.ekipman.model or ''}".strip() if kayit.ekipman else '-',
            BAKIM_TIPI_LABELS.get(kayit.bakim_tipi, kayit.bakim_tipi or '-'),
            SERVIS_TIPI_LABELS.get(kayit.servis_tipi, kayit.servis_tipi or '-'),
            kayit.calisma_saati if kayit.calisma_saati is not None else '',
            kim_yapti,
            parcalar_text,
            float(iscilik),
            float(malzeme),
            float(toplam),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font      = body_font
            cell.border    = cell_border
            cell.alignment = center_align if col_idx in center_cols else (right_align if col_idx in right_cols else left_align)
            if col_idx in right_cols:
                cell.number_format = '#,##0.00'

        ws.row_dimensions[current_row].height = 30
        current_row += 1

    if kayitlar:
        total_values = {
            10: float(toplam_iscilik),
            11: float(toplam_malzeme),
            12: float(toplam_toplam),
        }
        for col_idx in range(1, NUM_COLS + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font   = total_font
            cell.fill   = total_fill
            cell.border = total_border
            if col_idx == 1:
                cell.value     = 'TOPLAM'
                cell.alignment = left_align
            elif col_idx in total_values:
                cell.value        = total_values[col_idx]
                cell.alignment    = right_align
                cell.number_format = '#,##0.00'

    col_widths = {'A': 5, 'B': 13, 'C': 14, 'D': 22, 'E': 18, 'F': 16,
                  'G': 14, 'H': 22, 'I': 36, 'J': 14, 'K': 14, 'L': 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"servis_kayitlari_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@servis_bp.route('/duzenle/<int:id>', methods=['GET', 'POST'])
@login_required
def duzenle(id):
    kayit = BakimKaydi.query.filter(BakimKaydi.id == id, BakimKaydi.is_deleted == False).options(
        joinedload(BakimKaydi.ekipman).joinedload(Ekipman.sube),
        joinedload(BakimKaydi.kullanilan_parcalar).joinedload(KullanilanParca.stok_karti),
        joinedload(BakimKaydi.servis_veren_firma),
    ).first_or_404()

    if request.method == 'POST':
        bakim_verileri = {
            'tarih': request.form.get('tarih'),
            'calisma_saati': request.form.get('calisma_saati', type=int),
            'bakim_tipi': request.form.get('bakim_tipi') or 'ariza',
            'servis_tipi': request.form.get('servis_tipi') or 'ic_servis',
            'durum': request.form.get('durum') or 'acik',
            'servis_veren_firma_id': request.form.get('servis_veren_firma_id', type=int) or None,
            'servis_veren_kisi': (request.form.get('servis_veren_kisi') or '').strip() or None,
            'aciklama': (request.form.get('aciklama') or '').strip() or None,
            'sonraki_bakim_tarihi': request.form.get('sonraki_bakim_tarihi') or None,
            'toplam_iscilik_maliyeti': request.form.get('toplam_iscilik_maliyeti') or 0,
        }
        parts_data = _collect_parts_from_form(request.form)

        try:
            BakimService.bakim_guncelle(id, bakim_verileri, parts_data=parts_data, actor_id=_get_actor_id())
            flash('Servis kaydı güncellendi.', 'success')
            return redirect(url_for('servis.index', q=kayit.ekipman.kod if kayit.ekipman else ''))
        except ValidationError as exc:
            flash(str(exc), 'warning')
        except Exception as exc:
            flash(f'Hata: {exc}', 'danger')

    return render_template(
        'servis/duzenle.html',
        kayit=kayit,
        aktif_kiralama_var=BakimService.has_active_rental(kayit.ekipman_id) if kayit.ekipman_id else False,
        **_common_form_context(),
    )


@servis_bp.route('/sil/<int:id>', methods=['POST'])
@login_required
def sil(id):
    try:
        BakimService.bakim_sil(id, actor_id=_get_actor_id())
        flash('Servis kaydı silindi.', 'success')
    except ValidationError as exc:
        flash(str(exc), 'warning')
    except Exception as exc:
        flash(f'Hata: {exc}', 'danger')

    return redirect(url_for('servis.index'))


# -------------------------------------------------------------------------
# Bakımda olan makineler sayfası
# -------------------------------------------------------------------------
@servis_bp.route('/bakimda')
@login_required
def bakimda():
    """Serviste olan (bakımda) makinelerin listesi ve aktif makineler"""
    try:
        page = request.args.get('page', 1, type=int)
        q = request.args.get('q', '', type=str)

        # Serviste olan makineler + açık yerinde servis kaydı olan kirada makineler
        acik_yerinde_servis_ekipman_ids = [
            row[0] for row in db.session.query(BakimKaydi.ekipman_id).filter(
                BakimKaydi.is_deleted == False,
                BakimKaydi.durum.in_(('acik', 'parca_bekliyor')),
            ).distinct().all()
        ]

        serviste_query = Ekipman.query.filter(
            Ekipman.firma_tedarikci_id.is_(None),
            Ekipman.is_active == True,
            or_(
                Ekipman.calisma_durumu == 'serviste',
                Ekipman.id.in_(acik_yerinde_servis_ekipman_ids) if acik_yerinde_servis_ekipman_ids else False,
            )
        ).options(subqueryload(Ekipman.bakim_kayitlari))

        if q:
            serviste_query = serviste_query.filter(Ekipman.kod.ilike(f'%{q}%'))

        pagination = serviste_query.order_by(Ekipman.kod).paginate(page=page, per_page=25, error_out=False)
        for ekipman in pagination.items:
            acik_bakimlar = [
                kayit for kayit in ekipman.bakim_kayitlari
                if not kayit.is_deleted and kayit.durum in {'acik', 'parca_bekliyor'}
            ]
            ekipman.aktif_bakim_kaydi = max(acik_bakimlar, key=lambda kayit: (kayit.tarih or date.min, kayit.id)) if acik_bakimlar else None
            ekipman.aktif_kiralama_var = BakimService.has_active_rental(ekipman.id)

        # Aktif makineler (sahadaki, kullanımda olan) - sayfalanmaz
        aktif_query = Ekipman.query.filter(
            Ekipman.firma_tedarikci_id.is_(None),
            Ekipman.is_active == True,
            Ekipman.calisma_durumu != 'serviste'
        ).order_by(Ekipman.kod)

        if q:
            aktif_query = aktif_query.filter(Ekipman.kod.ilike(f'%{q}%'))

        aktif_makineler = aktif_query.all()

        return render_template('servis/bakimda.html',
                             ekipmanlar=pagination.items,
                             pagination=pagination,
                             aktif_makineler=aktif_makineler,
                             q=q)
    except Exception as e:
        flash(f"Hata: {str(e)}", "danger")
        return render_template('servis/bakimda.html',
                             ekipmanlar=[],
                             pagination=None,
                             aktif_makineler=[],
                             q='')


@servis_bp.route('/hizli_servis_ac', methods=['POST'])
@login_required
def hizli_servis_ac():
    """Aktif makineler için hızlı servis kaydı aç (servis türüne göre makine durumunu ayarla)"""
    try:
        ekipman_id = request.form.get('ekipman_id', type=int)
        servis_tipi = request.form.get('servis_tipi', 'ic_servis')

        if not ekipman_id:
            flash('Makine seçiniz.', 'warning')
            return redirect(url_for('servis.bakimda'))

        ekipman = EkipmanService.get_by_id(ekipman_id)
        if not ekipman:
            flash('Makine bulunamadı.', 'danger')
            return redirect(url_for('servis.bakimda'))

        # Serviste olan (kirada olan) makineler için servis kaydı açılamaz
        if ekipman.calisma_durumu == 'serviste':
            flash(f"'{ekipman.kod}' zaten serviste (kirada). Servis kaydı açılamazsınız.", 'warning')
            return redirect(url_for('servis.bakimda'))

        # Kirada (sahada çalışan) makine için sadece yerinde servis açılabilir.
        # Aksi halde makine durumu 'serviste' olur ve kiralama kaydı bozulur.
        if ekipman.calisma_durumu == 'kirada' and servis_tipi != 'yerinde_servis':
            flash(f"'{ekipman.kod}' sahada kirada olduğu için yalnızca 'Yerinde Servis' açılabilir.", 'warning')
            return redirect(url_for('servis.bakimda'))

        # Servis kaydı oluştur
        bakim_verileri = {
            'tarih': date.today(),
            'bakim_tipi': 'ariza',
            'servis_tipi': servis_tipi,
            'servis_veren_firma_id': None,
            'servis_veren_kisi': None,
            'aciklama': 'Hızlı servis kaydı açıldı.',
            'calisma_saati': None,
            'sonraki_bakim_tarihi': None,
            'toplam_iscilik_maliyeti': Decimal('0'),
            'durum': 'acik',
        }

        bakim_id = BakimService.bakim_kaydet(ekipman_id, bakim_verileri, actor_id=get_actor_id())

        # Servis türüne göre makine durumunu ayarla
        # Yerinde servis = makine durumu değişmez (aktif kalır)
        # İç/Dış servis = makine serviste geçer
        if servis_tipi != 'yerinde_servis':
            ekipman.calisma_durumu = 'serviste'
            from app.extensions import db
            db.session.commit()

        OperationLogService.log(
            module='servis', action='hizli_servis_ac',
            user_id=get_actor_id(),
            username=getattr(current_user, 'username', None),
            entity_type='Ekipman', entity_id=ekipman_id,
            description=f"{ekipman.kod} için servis kaydı açıldı ({servis_tipi}).",
            success=True
        )

        flash(f"'{ekipman.kod}' için servis kaydı açıldı.", "success")
        return redirect(url_for('servis.duzenle', id=bakim_id.id))
    except ValidationError as e:
        flash(str(e), "warning")
        return redirect(url_for('servis.bakimda'))
    except Exception as e:
        flash(f"Hata: {str(e)}", "danger")
        return redirect(url_for('servis.bakimda'))


@servis_bp.route('/bakim_bitir/<int:id>', methods=['POST'])
@login_required
def bakim_bitir(id):
    """Makinenin bakımını tamamla ve servisten çıkar"""
    try:
        ekipman = EkipmanService.get_by_id(id)
        if ekipman:
            kapatilan = BakimService.ekipman_bakimini_tamamla(id, actor_id=get_actor_id())
            if kapatilan:
                OperationLogService.log(
                    module='servis', action='bakim_bitir',
                    user_id=get_actor_id(),
                    username=getattr(current_user, 'username', None),
                    entity_type='Ekipman', entity_id=id,
                    description=f"{ekipman.kod} servisten çıkarıldı.",
                    success=True
                )
                flash(f"'{ekipman.kod}' servisi tamamlandı.", "success")
            else:
                flash(f"'{ekipman.kod}' için açık servis kaydı bulunamadı.", "warning")
    except ValidationError as e:
        flash(str(e), "danger")
    except Exception as e:
        flash(f"Hata: {str(e)}", "danger")

    return redirect(url_for('servis.bakimda'))