import traceback
import threading
import json
from datetime import datetime, date, timedelta, timezone
from io import BytesIO
from flask import render_template, redirect, url_for, flash, request, current_app, jsonify, send_file
from flask_login import current_user, login_required
from sqlalchemy import or_
from sqlalchemy.orm import joinedload
from openpyxl import Workbook, load_workbook

from app import db
from app.kiralama import kiralama_bp

# Modeller
from app.firmalar.models import Firma
from app.kiralama.models import Kiralama, KiralamaKalemi
from app.filo.models import Ekipman
from app.filo.forms import EKIPMAN_TIPI_SECENEKLERI
from app.subeler.models import Sube
from app.araclar.models import Arac as NakliyeAraci
from app.kiralama.forms import KiralamaForm

# Servis Katmanı ve Hata Yönetimi
from app.services.kiralama_services import KiralamaService, KiralamaKalemiService
from app.services.base import ValidationError
from app.services.operation_log_service import OperationLogService
from app.utils import ensure_active_sube_exists

# --- BELLEK İÇİ ÖNBELLEKLEME (IN-MEMORY CACHE) ---
_CACHE_DATA = {
    'subeler': {'data': None, 'last_update': None},
    'aktif_araclar': {'data': None, 'last_update': None}
}

# DARBOĞAZ ÖNLEME: Her veri kümesi için ayrı kilit (lock) tanımlandı.
# Böylece şube güncellenirken araç listesi etkilenmez.
_SUBE_CACHE_LOCK = threading.Lock()
_ARAC_CACHE_LOCK = threading.Lock()

_CACHE_TIMEOUT_MINUTES = 60 

def get_cached_subeler():
    """Şubeleri thread-safe ve bağımsız bir kilitle bellekten getirir."""
    now = datetime.now()
    cache = _CACHE_DATA['subeler']
    
    # 1. Hızlı Okuma (last_update None kontrolü: thread race condition'ı önler)
    if cache['data'] is not None and cache['last_update'] is not None and (now - cache['last_update']) < timedelta(minutes=_CACHE_TIMEOUT_MINUTES):
        return cache['data']

    # 2. Kaynağa Özel Kilit
    with _SUBE_CACHE_LOCK:
        if cache['data'] is None or cache['last_update'] is None or (datetime.now() - cache['last_update']) > timedelta(minutes=_CACHE_TIMEOUT_MINUTES):
            try:
                cache['data'] = Sube.query.all()
                cache['last_update'] = datetime.now()
            except Exception as e:
                current_app.logger.error(f"Sube Cache Hatası: {e}")
                return cache['data'] or []
    return cache['data']

def get_cached_aktif_araclar():
    """Aktif nakliye araçlarını thread-safe ve bağımsız bir kilitle bellekten getirir."""
    now = datetime.now()
    cache = _CACHE_DATA['aktif_araclar']
    
    if cache['data'] is not None and cache['last_update'] is not None and (now - cache['last_update']) < timedelta(minutes=_CACHE_TIMEOUT_MINUTES):
        return cache['data']

    with _ARAC_CACHE_LOCK:
        if cache['data'] is None or cache['last_update'] is None or (datetime.now() - cache['last_update']) > timedelta(minutes=_CACHE_TIMEOUT_MINUTES):
            try:
                cache['data'] = NakliyeAraci.aktif_nakliye_query().order_by(NakliyeAraci.plaka).all()
                cache['last_update'] = datetime.now()
            except Exception as e:
                current_app.logger.error(f"Arac Cache Hatası: {e}")
                return cache['data'] or []
    return cache['data']

def populate_kiralama_form_choices(form, include_ids=None):
    """
    Formdaki tüm SelectField alanlarını veritabanından dinamik olarak doldurur.
    """
    if include_ids is None: include_ids = []
    
    # 1. Ana Müşteri ve Tedarikçi Listeleri (Dahili firmalar gizlendi)
    musteriler = Firma.query.filter(
        Firma.is_musteri == True, 
        Firma.is_active == True,
        Firma.firma_adi.notin_(['DAHİLİ İŞLEMLER', 'Dahili Kasa İşlemleri'])
    ).order_by(Firma.firma_adi).all()
    form.firma_musteri_id.choices = [(0, '--- Müşteri Seçiniz ---')] + [(f.id, f.firma_adi) for f in musteriler]

    tedarikciler = Firma.query.filter(
        Firma.is_tedarikci == True, 
        Firma.is_active == True,
        Firma.firma_adi.notin_(['DAHİLİ İŞLEMLER', 'Dahili Kasa İşlemleri'])
    ).order_by(Firma.firma_adi).all()
    ted_choices = [(0, '--- Tedarikçi Seçiniz ---')] + [(f.id, f.firma_adi) for f in tedarikciler]
    
    # 2. Makine Parkı (Pimaks Filosu) - Detaylandırılmış Etiketler Eklendi
    filo_query = Ekipman.query.filter(
        Ekipman.firma_tedarikci_id.is_(None),
        or_(Ekipman.calisma_durumu == 'bosta', Ekipman.id.in_(include_ids))
    ).order_by(Ekipman.kod).all()
    
    # KOD | TİP (Marka-Yükseklikm - Kapasitekg) formatı uygulandı
    pimaks_choices = [(0, '--- Seçiniz ---')] + [
        (
            e.id,
            f"{(e.kod or '').strip()} | {(e.tipi or '').strip()} ({(e.marka or 'Bilinmiyor').strip()}-{e.calisma_yuksekligi or 0}m - {e.kaldirma_kapasitesi or 0}kg)"
        )
        for e in filo_query
    ]

    # 3. Nakliye Araçları (Cache destekli)
    arac_choices = [(0, '--- Araç Seçiniz ---')] + [(a.id, f"{a.plaka} - {a.arac_tipi}") for a in get_cached_aktif_araclar()]

    # 4. Kalemler Listesi Doldurma
    if not form.kalemler.entries:
        form.kalemler.append_entry()

    for entry in form.kalemler:
        f = entry.form
        f.ekipman_id.choices = pimaks_choices
        f.harici_ekipman_tedarikci_id.choices = ted_choices
        f.nakliye_tedarikci_id.choices = ted_choices
        f.nakliye_araci_id.choices = arac_choices

@kiralama_bp.route('/')
@kiralama_bp.route('/index')
@login_required
def index():
    """Kiralama ana listesi ve arama."""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        if per_page not in {10, 25, 50, 100}:
            per_page = 25
        q = request.args.get('q', '', type=str)

        query = Kiralama.query.options(
            joinedload(Kiralama.firma_musteri),
            joinedload(Kiralama.kalemler).joinedload(KiralamaKalemi.ekipman)
        )

        if q:
            search = f"%{q}%"
            query = query.join(Firma, Kiralama.firma_musteri_id == Firma.id)\
                         .filter(or_(Kiralama.kiralama_form_no.ilike(search), Firma.firma_adi.ilike(search)))

        pagination = query.order_by(Kiralama.kiralama_form_no.desc()).paginate(page=page, per_page=per_page)

        # ÖNCELİK: Kalem verilerini rollback'ten etkilenmeden önce hesapla.
        # (guncelle_cari_toplam rollback yaparsa session expire olur ve objelere erişim patlar)
        recent_threshold = datetime.now(timezone.utc) - timedelta(days=1)
        recently_returned_kalem_ids = {
            kalem.id
            for kiralama in pagination.items
            for kalem in kiralama.kalemler
            if kalem.sonlandirildi
            and kalem.is_active
            and kalem.updated_at
            and (kalem.updated_at if kalem.updated_at.tzinfo else kalem.updated_at.replace(tzinfo=timezone.utc)) >= recent_threshold
        }

        # Liste ekranında görünen kiralamalar için bekleyen cari tutarı güncel tut.
        # Bu işlem bağımsız bir try bloğunda çalışır; başarısız olursa
        # liste yine de gösterilir, sadece cari tutarlar güncellenmemiş olur.
        try:
            for kiralama in pagination.items:
                KiralamaService.guncelle_cari_toplam(kiralama.id, auto_commit=False)
            if pagination.items:
                db.session.commit()
        except Exception as sync_err:
            db.session.rollback()
            current_app.logger.warning(f"Kiralama cari senkronizasyon uyarısı: {sync_err}")
            # Rollback sonrası objeler expire oldu, listeyi yeniden yükle
            pagination = query.order_by(Kiralama.kiralama_form_no.desc()).paginate(page=page, per_page=per_page)

        return render_template(
            'kiralama/index.html',
            kiralamalar=pagination.items,
            pagination=pagination,
            per_page=per_page,
            q=q,
            kurlar=KiralamaService.get_tcmb_kurlari(),
            today=date.today(),
            subeler=get_cached_subeler(),
            nakliye_araclari=get_cached_aktif_araclar(),
            nakliye_tedarikci_listesi=Firma.query.filter_by(is_tedarikci=True).order_by(Firma.firma_adi).all(),
            recently_returned_kalem_ids=recently_returned_kalem_ids
        )
    except Exception as e:
        current_app.logger.error(f"Kiralama Liste Yükleme Hatası: {str(e)}\n{traceback.format_exc()}")
        flash(f"Liste yüklenirken bir hata oluştu.", "danger")
        return render_template('kiralama/index.html', kiralamalar=[], pagination=None, per_page=25, q='', kurlar={}, today=date.today(), subeler=[], nakliye_araclari=[], nakliye_tedarikci_listesi=[], recently_returned_kalem_ids=set())

@kiralama_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
def ekle():
    """Yeni kiralama kaydı oluşturma."""
    guard_response = ensure_active_sube_exists(
        warning_message="Kiralama oluşturmadan önce en az bir aktif şube / depo tanımlamalısınız."
    )
    if guard_response:
        return guard_response

    form = KiralamaForm()
    
    try:
        populate_kiralama_form_choices(form)
    except Exception as e:
        current_app.logger.error(f"Seçenek doldurma hatası (Ekle): {str(e)}")
        flash("Seçenek listeleri yüklenemedi. Lütfen sistem yöneticisine başvurun.", "danger")
        return redirect(url_for('kiralama.index'))

    # --- TCMB KURU VE FORM NUMARASI OTOMATİK DOLDURMA ---
    if request.method == 'GET':
        # Kur bilgisini al
        try:
            kurlar = KiralamaService.get_tcmb_kurlari()
            form.doviz_kuru_usd.data = kurlar.get('USD', 1.0)
            if hasattr(form, 'doviz_kuru_eur'):
                form.doviz_kuru_eur.data = kurlar.get('EUR', 1.0)
        except Exception as e:
            current_app.logger.warning(f"TCMB kur bilgisi alınamadı: {str(e)}")
        
        # Form numarasını otomatik al, hata durumunda manuel girilmesine izin ver
        try:
            form.kiralama_form_no.data = KiralamaService.get_next_form_no()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Form numarası otomatik alınamadı: {str(e)}")
            # Son form numarasını al ve kullanıcıya bilgi ver
            try:
                last_kiralama = Kiralama.query.order_by(Kiralama.id.desc()).first()
                last_form_no = last_kiralama.kiralama_form_no if last_kiralama else "Kayıt bulunamadı"
                flash(
                    f"Uyarı: Form numarası otomatik alınamadı. Son form numarası: {last_form_no}. "
                    f"Lütfen manuel olarak giriniz.",
                    "warning"
                )
            except Exception as e2:
                flash(
                    "Uyarı: Form numarası otomatik alınamadı. Lütfen manuel olarak giriniz. "
                    "Örnek format: PF-2026/0001",
                    "warning"
                )
    # ----------------------------------------------------

    if form.validate_on_submit():
        try:
            import traceback
            actor_id = getattr(current_user, 'id', None)
            kiralama_data = {
                'kiralama_form_no': form.kiralama_form_no.data,
                'makine_calisma_adresi': form.makine_calisma_adresi.data,
                'firma_musteri_id': form.firma_musteri_id.data,
                'kdv_orani': form.kdv_orani.data,
                'doviz_kuru_usd': form.doviz_kuru_usd.data,
                'doviz_kuru_eur': getattr(form, 'doviz_kuru_eur', form.doviz_kuru_usd).data
            }
            kalemler_data = [k_form.data for k_form in form.kalemler]
            current_app.logger.debug(f"[EKLE] kiralama_data: {kiralama_data}")
            current_app.logger.debug(f"[EKLE] kalemler_data: {kalemler_data}")

            created_kiralama = KiralamaService.create_kiralama_with_relations(kiralama_data, kalemler_data, actor_id=actor_id)
            current_app.logger.debug(f"[EKLE] created_kiralama: {created_kiralama}")
            OperationLogService.log(
                module='kiralama',
                action='create',
                user_id=actor_id,
                username=getattr(current_user, 'username', None),
                entity_type='Kiralama',
                entity_id=getattr(created_kiralama, 'id', None),
                description=f"Kiralama oluşturuldu: {created_kiralama.kiralama_form_no}",
                success=True,
            )
            # İleri tarihli kalem uyarısı
            from datetime import date as _date
            ileri_tarihli = any(
                k.get('kiralama_baslangici') and str(k.get('kiralama_baslangici')) > _date.today().isoformat()
                for k in kalemler_data
            )
            if ileri_tarihli:
                flash('Kiralama kaydı başarıyla oluşturuldu. ⚠️ Bir veya daha fazla kalemin başlangıç tarihi ileridedir — cari tahakkuk başlangıç tarihine geldiğinde otomatik yansıyacaktır.', 'info')
            else:
                flash('Kiralama kaydı başarıyla oluşturuldu.', 'success')
            return redirect(url_for('kiralama.index'))

        except ValidationError as e:
            db.session.rollback()
            OperationLogService.log(
                module='kiralama',
                action='create',
                user_id=getattr(current_user, 'id', None),
                username=getattr(current_user, 'username', None),
                entity_type='Kiralama',
                description=f"Kiralama oluşturma doğrulama hatası: {str(e)}",
                success=False,
            )
            flash(f"Doğrulama Hatası: {str(e)}", "warning")
        except ValueError as e:
            db.session.rollback()
            OperationLogService.log(
                module='kiralama',
                action='create',
                user_id=getattr(current_user, 'id', None),
                username=getattr(current_user, 'username', None),
                entity_type='Kiralama',
                description=f"Kiralama oluşturma veri hatası: {str(e)}",
                success=False,
            )
            flash(f"Veri Hatası: {str(e)}", "danger")
        except Exception as e:
            db.session.rollback()
            import traceback
            tb_str = traceback.format_exc()
            current_app.logger.error(f"Kiralama Kayıt Hatası: {str(e)}\nTraceback:\n{tb_str}")
            OperationLogService.log(
                module='kiralama',
                action='create',
                user_id=getattr(current_user, 'id', None),
                username=getattr(current_user, 'username', None),
                entity_type='Kiralama',
                description=f"Kiralama oluşturma sistem hatası: {str(e)}\nTraceback:\n{tb_str}",
                success=False,
            )
            flash(f"Sistemsel bir hata oluştu. Lütfen tekrar deneyin.", "danger")
    
    elif request.method == 'POST':
        for field, errors in form.errors.items():
            if field == 'kalemler':
                for idx, kalem_errors in enumerate(errors):
                    for k_field, k_msg in kalem_errors.items():
                        flash(f"Satır {idx+1} - {k_field}: {k_msg}", "warning")
            else:
                flash(f"{field}: {errors}", "warning")

    # Rollback sonrası session bozulmuş olabilir; form render için gereken
    # sorguları korumalı blokta çalıştır
    try:
        ekipman_sube_map = {
            e.id: (e.sube.isim if e.sube else 'Şube Tanımsız')
            for e in Ekipman.query.options(joinedload(Ekipman.sube)).all()
        }
        subeler = Sube.query.all()
        markalar = [m[0] for m in db.session.query(Ekipman.marka).filter(Ekipman.marka.isnot(None)).distinct().all()]
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Ekle formu yardımcı veri hatası: {str(e)}")
        ekipman_sube_map = {}
        subeler = []
        markalar = []
    tipler = [tip for tip, _ in EKIPMAN_TIPI_SECENEKLERI]
    return render_template('kiralama/form.html', form=form, subeler=subeler, markalar=markalar, tipler=tipler, is_edit=False, ekipman_sube_map=ekipman_sube_map, ekipman_map_json='{}')

@kiralama_bp.route('/duzenle/<int:kiralama_id>', methods=['GET', 'POST'])
@login_required
def duzenle(kiralama_id):
    """Mevcut bir kiralama kaydını düzenleme."""
    guard_response = ensure_active_sube_exists(
        warning_message="Kiralama düzenlemeden önce en az bir aktif şube / depo tanımlamalısınız."
    )
    if guard_response:
        return guard_response

    kiralama = db.get_or_404(Kiralama, kiralama_id)
    form = KiralamaForm(obj=kiralama)
    form.current_kiralama_id = kiralama.id

    # Düzenleme ekranında form numarası değiştirilemez.
    if request.method == 'POST':
        form.kiralama_form_no.data = kiralama.kiralama_form_no

    if request.method == 'GET':
        form.makine_calisma_adresi.data = kiralama.makine_calisma_adresi
    

    try:
        include_ids = [k.ekipman_id for k in kiralama.kalemler if k.ekipman_id]
        populate_kiralama_form_choices(form, include_ids=include_ids)
        # Güvenlik: Eğer choices None ise boş listeye çevir
        for entry in form.kalemler.entries:
            for field_name in ['ekipman_id', 'harici_ekipman_tedarikci_id', 'nakliye_tedarikci_id', 'nakliye_araci_id']:
                field = getattr(entry.form, field_name, None)
                if field is not None and getattr(field, 'choices', None) is None:
                    field.choices = []
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        current_app.logger.error(f"Seçenek doldurma hatası (Düzenle): {str(e)}\n{tb}")
        flash(f"Form seçenekleri yüklenirken hata oluştu: {str(e)}", "danger")
        return redirect(url_for('kiralama.index'))

    if request.method == 'GET':
        # Kalemler listesini sıfırla ve modelden doldur
        form.kalemler.entries = []
        if kiralama.kalemler is None:
            current_app.logger.error(f"[DUZENLE] kiralama.kalemler None! kiralama_id={kiralama.id}")
        kalemler = kiralama.kalemler if kiralama.kalemler is not None else []
        for idx, kalem in enumerate(kalemler):
            entry = form.kalemler.append_entry()
            k_form = entry.form
            current_app.logger.info(f"[DUZENLE] kalem[{idx}].nakliye_alis_kdv DB'den: {kalem.nakliye_alis_kdv} -> Form: {getattr(k_form, 'nakliye_alis_kdv', None)}")
            # SelectField alanlarının choices'ı None ise boş listeye set et
            for field_name in ['ekipman_id', 'harici_ekipman_tedarikci_id', 'nakliye_tedarikci_id', 'nakliye_araci_id']:
                field = getattr(k_form, field_name, None)
                if field is not None and getattr(field, 'choices', None) is None:
                    field.choices = []
            k_form.id.data = kalem.id
            k_form.dis_tedarik_ekipman.data = 1 if getattr(kalem, 'is_dis_tedarik_ekipman', False) else 0
            k_form.ekipman_id.data = kalem.ekipman_id
            k_form.harici_ekipman_tedarikci_id.data = kalem.harici_ekipman_tedarikci_id
            k_form.harici_ekipman_tipi.data = kalem.harici_ekipman_tipi
            k_form.harici_ekipman_marka.data = kalem.harici_ekipman_marka
            k_form.harici_ekipman_model.data = kalem.harici_ekipman_model
            k_form.harici_ekipman_seri_no.data = kalem.harici_ekipman_seri_no
            k_form.harici_ekipman_kaldirma_kapasitesi.data = kalem.harici_ekipman_kapasite
            k_form.harici_ekipman_calisma_yuksekligi.data = kalem.harici_ekipman_yukseklik
            k_form.harici_ekipman_uretim_tarihi.data = kalem.harici_ekipman_uretim_yili
            k_form.kiralama_baslangici.data = kalem.kiralama_baslangici
            k_form.kiralama_bitis.data = kalem.kiralama_bitis
            k_form.kiralama_brm_fiyat.data = kalem.kiralama_brm_fiyat
            k_form.kiralama_alis_fiyat.data = kalem.kiralama_alis_fiyat
            k_form.kiralama_alis_kdv.data = kalem.kiralama_alis_kdv
            k_form.nakliye_alis_kdv.data = kalem.nakliye_alis_kdv
            k_form.nakliye_satis_kdv.data = kalem.nakliye_satis_kdv
            k_form.dis_tedarik_nakliye.data = 1 if getattr(kalem, 'is_harici_nakliye', False) else 0
            k_form.nakliye_satis_fiyat.data = kalem.nakliye_satis_fiyat
            k_form.donus_nakliye_fatura_et.data = 1 if getattr(kalem, 'donus_nakliye_fatura_et', False) else 0
            k_form.nakliye_alis_fiyat.data = kalem.nakliye_alis_fiyat
            k_form.nakliye_tedarikci_id.data = kalem.nakliye_tedarikci_id
            k_form.nakliye_araci_id.data = kalem.nakliye_araci_id
            # Ek alanlar (varsa formda karşılığı olanlar)
            if hasattr(k_form, 'sonlandirildi'):
                k_form.sonlandirildi.data = kalem.sonlandirildi
            if hasattr(k_form, 'is_active'):
                k_form.is_active.data = kalem.is_active
            if hasattr(k_form, 'parent_id'):
                k_form.parent_id.data = kalem.parent_id
            if hasattr(k_form, 'versiyon_no'):
                k_form.versiyon_no.data = kalem.versiyon_no
            if hasattr(k_form, 'degisim_nedeni'):
                k_form.degisim_nedeni.data = kalem.degisim_nedeni
            if hasattr(k_form, 'degisim_tarihi'):
                k_form.degisim_tarihi.data = kalem.degisim_tarihi
            if hasattr(k_form, 'cikis_saati'):
                k_form.cikis_saati.data = kalem.cikis_saati
            if hasattr(k_form, 'donus_saati'):
                k_form.donus_saati.data = kalem.donus_saati
            if hasattr(k_form, 'degisim_aciklama'):
                k_form.degisim_aciklama.data = kalem.degisim_aciklama
            if hasattr(k_form, 'chain_id'):
                k_form.chain_id.data = kalem.chain_id
        # Kalemler doldurulduktan sonra tekrar choices'ları ata
        try:
            populate_kiralama_form_choices(form, include_ids=include_ids)
        except Exception as e:
            current_app.logger.error(f"Seçenek doldurma hatası (Düzenle/GET tekrar): {str(e)}")

    # Tarih field'larını manual parse et (WTForms validation hatasından kaçınmak için)
    if request.method == 'POST':
        from datetime import datetime as dt
        for idx, k_form in enumerate(form.kalemler):
            try:
                bas_val = request.form.get(f'kalemler-{idx}-kiralama_baslangici', '')
                bit_val = request.form.get(f'kalemler-{idx}-kiralama_bitis', '')
                if bas_val:
                    k_form.kiralama_baslangici.data = dt.strptime(bas_val, '%Y-%m-%d').date()
                if bit_val:
                    k_form.kiralama_bitis.data = dt.strptime(bit_val, '%Y-%m-%d').date()
            except ValueError:
                pass

    if form.validate_on_submit():
        try:
            # Tamamlanmış kalemler: UI'da butonlar pasif, service'te sonlandirildi korunuyor.
            # Form POST'u engellemeye gerek yok, tamamlanmış kalemler olduğu gibi geçer.

            actor_id = getattr(current_user, 'id', None)
            kiralama_data = {
                'makine_calisma_adresi': form.makine_calisma_adresi.data,
                'firma_musteri_id': form.firma_musteri_id.data,
                'kdv_orani': form.kdv_orani.data,
                'doviz_kuru_usd': form.doviz_kuru_usd.data,
                'doviz_kuru_eur': getattr(form, 'doviz_kuru_eur', form.doviz_kuru_usd).data
            }
            kalemler_data = [k_form.data for k_form in form.kalemler]

            KiralamaService.update_kiralama_with_relations(kiralama.id, kiralama_data, kalemler_data, actor_id=actor_id)
            OperationLogService.log(
                module='kiralama',
                action='update',
                user_id=actor_id,
                username=getattr(current_user, 'username', None),
                entity_type='Kiralama',
                entity_id=kiralama.id,
                description=f"Kiralama güncellendi: {kiralama.kiralama_form_no}",
                success=True,
            )
            # İleri tarihli kalem uyarısı
            from datetime import date as _date
            ileri_tarihli = any(
                k.get('kiralama_baslangici') and str(k.get('kiralama_baslangici')) > _date.today().isoformat()
                for k in kalemler_data
            )
            if ileri_tarihli:
                flash('Kiralama başarıyla güncellendi. ⚠️ Bir veya daha fazla kalemin başlangıç tarihi ileridedir — cari tahakkuk başlangıç tarihine geldiğinde otomatik yansıyacaktır.', 'info')
            else:
                flash('Kiralama başarıyla güncellendi.', 'success')
            return redirect(url_for('kiralama.index'))
        except ValidationError as e:
            db.session.rollback()
            OperationLogService.log(
                module='kiralama',
                action='update',
                user_id=getattr(current_user, 'id', None),
                username=getattr(current_user, 'username', None),
                entity_type='Kiralama',
                entity_id=kiralama_id,
                description=f"Kiralama güncelleme doğrulama hatası: {str(e)}",
                success=False,
            )
            flash(f"Hata: {str(e)}", "warning")
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Kiralama Güncelleme Hatası (ID: {kiralama_id}): {str(e)}")
            OperationLogService.log(
                module='kiralama',
                action='update',
                user_id=getattr(current_user, 'id', None),
                username=getattr(current_user, 'username', None),
                entity_type='Kiralama',
                entity_id=kiralama_id,
                description=f"Kiralama güncelleme sistem hatası: {str(e)}",
                success=False,
            )
            flash(f"Güncelleme sırasında sistemsel bir hata oluştu.", "danger")
    else:
        if request.method == 'POST':
            # Form validasyon hatalarını ve eksik alanları logla
            current_app.logger.warning(f"Kiralama düzenle formu validasyon hatası: {form.errors}")
            for idx, kalem in enumerate(form.kalemler):
                if kalem.errors:
                    current_app.logger.warning(f"Kiralama kalemi {idx} errors: {kalem.errors}")
            # Hataları kullanıcıya göster
            for field, errors in form.errors.items():
                if field == 'kalemler':
                    for idx, kalem_errors in enumerate(errors):
                        for k_field, k_msg in kalem_errors.items():
                            flash(f"Satır {idx+1} - {k_field}: {k_msg}", "warning")
                else:
                    flash(f"{field}: {errors}", "warning")
    # Rollback sonrası session bozulmuş olabilir; form render için gereken
    # sorguları korumalı blokta çalıştır
    try:
        # Rollback sonrası kiralama objesi expire olmuş olabilir, yeniden yükle
        db.session.refresh(kiralama)
        include_ids_for_map = [k.ekipman_id for k in kiralama.kalemler if k.ekipman_id]
        filo_query_for_map = Ekipman.query.filter(
            Ekipman.firma_tedarikci_id.is_(None),
            or_(Ekipman.calisma_durumu == 'bosta', Ekipman.id.in_(include_ids_for_map))
        ).order_by(Ekipman.kod).all()

        ekipman_map = {
            e.id: f"{(e.kod or '').strip()} | {(e.tipi or '').strip()} ({(e.marka or 'Bilinmiyor').strip()}-{e.calisma_yuksekligi or 0}m - {e.kaldirma_kapasitesi or 0}kg)"
            for e in filo_query_for_map
        }

        ekipman_sube_map = {
            e.id: (e.sube.isim if e.sube else 'Şube Tanımsız')
            for e in Ekipman.query.options(joinedload(Ekipman.sube)).all()
        }
        subeler = Sube.query.all()
        markalar = [m[0] for m in db.session.query(Ekipman.marka).filter(Ekipman.marka.isnot(None)).distinct().all()]
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Düzenle formu yardımcı veri hatası: {str(e)}")
        ekipman_map = {}
        ekipman_sube_map = {}
        subeler = []
        markalar = []
    tipler = [tip for tip, _ in EKIPMAN_TIPI_SECENEKLERI]
    return render_template('kiralama/form.html', form=form, kiralama=kiralama, markalar=markalar, subeler=subeler, tipler=tipler, is_edit=True, ekipman_sube_map=ekipman_sube_map, ekipman_map_json=json.dumps(ekipman_map))

@kiralama_bp.route('/sil/<int:kiralama_id>', methods=['POST'])
@login_required
def sil(kiralama_id):
    """Kiralama ve bağlı finansal kayıtları siler."""
    try:
        actor_id = getattr(current_user, 'id', None)
        KiralamaService.delete_with_relations(kiralama_id, actor_id=actor_id)
        OperationLogService.log(
            module='kiralama',
            action='delete',
            user_id=actor_id,
            username=getattr(current_user, 'username', None),
            entity_type='Kiralama',
            entity_id=kiralama_id,
            description=f"Kiralama silindi (soft/hard): ID={kiralama_id}",
            success=True,
        )
        flash('Kiralama kaydı ve bağlı tüm hareketler silindi.', 'success')
    except ValidationError as e:
        OperationLogService.log(
            module='kiralama',
            action='delete',
            user_id=getattr(current_user, 'id', None),
            username=getattr(current_user, 'username', None),
            entity_type='Kiralama',
            entity_id=kiralama_id,
            description=f"Kiralama silme doğrulama hatası: {str(e)}",
            success=False,
        )
        flash(str(e), "warning")
    except Exception as e:
        current_app.logger.error(f"Kiralama Silme Hatası (ID: {kiralama_id}): {str(e)}")
        OperationLogService.log(
            module='kiralama',
            action='delete',
            user_id=getattr(current_user, 'id', None),
            username=getattr(current_user, 'username', None),
            entity_type='Kiralama',
            entity_id=kiralama_id,
            description=f"Kiralama silme sistem hatası: {str(e)}",
            success=False,
        )
        flash(f'Silme işlemi başarısız oldu.', 'danger')
    return redirect(url_for('kiralama.index'))

@kiralama_bp.route('/kalem/sonlandir', methods=['POST'])
@login_required
def sonlandir_kalem():
    """Kiralama kalemini kapatır ve makineyi boşa çıkarır."""
    try:
        kalem_id = request.form.get('kalem_id', type=int)
        if not kalem_id:
            flash("İşlem yapılacak kiralama kalemi seçilmedi.", "warning")
            return redirect(url_for('kiralama.index'))

        actor_id = getattr(current_user, 'id', None)
        bitis_str = request.form.get('bitis_tarihi')
        donus_sube_id = request.form.get('donus_sube_id')
        is_harici_nakliye = request.form.get('is_harici_nakliye') in ('on', '1', 'true', 'True')
        nakliye_tedarikci_id = request.form.get('nakliye_tedarikci_id', type=int)
        nakliye_araci_id = request.form.get('nakliye_araci_id', type=int)
        nakliye_alis_fiyat = request.form.get('nakliye_alis_fiyat')
        donus_nakliye_satis_fiyat = request.form.get('donus_nakliye_satis_fiyat')
        
        KiralamaKalemiService.sonlandir(
            kalem_id,
            bitis_str,
            donus_sube_id,
            actor_id=actor_id,
            is_harici_nakliye=is_harici_nakliye,
            nakliye_tedarikci_id=nakliye_tedarikci_id,
            nakliye_araci_id=nakliye_araci_id,
            nakliye_alis_fiyat=nakliye_alis_fiyat,
            donus_nakliye_satis_fiyat=donus_nakliye_satis_fiyat,
        )
        OperationLogService.log(
            module='kiralama',
            action='sonlandir_kalem',
            user_id=actor_id,
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=kalem_id,
            description=f"Kiralama kalemi sonlandırıldı: kalem_id={kalem_id}",
            success=True,
        )
        flash("Kiralama başarıyla sonlandırıldı.", "success")
    except ValidationError as e:
        OperationLogService.log(
            module='kiralama',
            action='sonlandir_kalem',
            user_id=getattr(current_user, 'id', None),
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=request.form.get('kalem_id', type=int),
            description=f"Kalem sonlandırma doğrulama hatası: {str(e)}",
            success=False,
        )
        flash(f"Hata: {str(e)}", "warning")
    except Exception as e:
        current_app.logger.error(f"Kalem Sonlandırma Hatası: {str(e)}")
        OperationLogService.log(
            module='kiralama',
            action='sonlandir_kalem',
            user_id=getattr(current_user, 'id', None),
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=request.form.get('kalem_id', type=int),
            description=f"Kalem sonlandırma sistem hatası: {str(e)}",
            success=False,
        )
        flash(f"İşlem sırasında bir hata oluştu.", "danger")
    return redirect(url_for('kiralama.index'))

@kiralama_bp.route('/kalem/tarih_guncelle', methods=['POST'])
@login_required
def tarih_guncelle_kalem():
    """Aktif kiralama kaleminin planlanan bitiş tarihini günceller."""
    try:
        kalem_id = request.form.get('kalem_id', type=int)
        yeni_bitis_str = request.form.get('yeni_bitis_tarihi')

        if not kalem_id or not yeni_bitis_str:
            flash("Kalem ID veya tarih eksik.", "warning")
            return redirect(url_for('kiralama.index'))

        # Kalem'i bul
        kalem = db.session.get(KiralamaKalemi, kalem_id)
        if not kalem:
            flash("Kalem bulunamadı.", "danger")
            return redirect(url_for('kiralama.index'))

        # Aktif ve sonlandırılmamış olmalı
        if kalem.sonlandirildi or not kalem.is_active:
            flash("Sadece aktif ve sonlandırılmamış kalemler güncellenebilir.", "warning")
            return redirect(url_for('kiralama.index'))

        # Tarih validasyonu
        from app.services.kiralama_services import to_date
        yeni_bitis_date = to_date(yeni_bitis_str)
        if not yeni_bitis_date:
            flash("Geçersiz tarih formatı.", "warning")
            return redirect(url_for('kiralama.index'))

        # Yeni bitiş tarihi başlangıçtan sonra olmalı
        bas_date = to_date(kalem.kiralama_baslangici)
        if yeni_bitis_date < bas_date:
            flash("Bitiş tarihi başlangıç tarihinden sonra olmalıdır.", "warning")
            return redirect(url_for('kiralama.index'))

        # Eski tarihi kaydet
        eski_bitis = kalem.kiralama_bitis

        # Yeni tarihi set et
        kalem.kiralama_bitis = yeni_bitis_date

        # DB'ye kaydet
        db.session.add(kalem)
        db.session.commit()

        actor_id = getattr(current_user, 'id', None)

        # Log kaydı oluştur
        OperationLogService.log(
            module='kiralama',
            action='tarih_guncelle',
            user_id=actor_id,
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=kalem_id,
            description=f"Kiralama kalemi bitiş tarihi güncellendi: {eski_bitis} → {yeni_bitis_date}",
            success=True,
        )

        flash(f"Bitiş tarihi başarıyla güncellendi: {yeni_bitis_date.strftime('%d.%m.%Y')}", "success")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Kalem Tarih Güncelleme Hatası: {str(e)}", exc_info=True)
        OperationLogService.log(
            module='kiralama',
            action='tarih_guncelle',
            user_id=getattr(current_user, 'id', None),
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=request.form.get('kalem_id', type=int),
            description=f"Kalem tarih güncelleme hatası: {str(e)}",
            success=False,
        )
        flash(f"İşlem sırasında bir hata oluştu: {str(e)}", "danger")

    return redirect(url_for('kiralama.index'))

@kiralama_bp.route('/kalem/iptal_et', methods=['POST'])
@login_required
def iptal_et_kalem():
    """Sonlandırma işlemini geri alır."""
    try:
        kalem_id = request.form.get('kalem_id', type=int)
        if not kalem_id:
            flash("Hatalı kalem seçimi.", "warning")
            return redirect(url_for('kiralama.index'))

        actor_id = getattr(current_user, 'id', None)
        KiralamaKalemiService.iptal_et_sonlandirma(kalem_id, actor_id=actor_id)
        OperationLogService.log(
            module='kiralama',
            action='iptal_sonlandirma',
            user_id=actor_id,
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=kalem_id,
            description=f"Kalem sonlandırma iptal edildi: kalem_id={kalem_id}",
            success=True,
        )
        flash("İşlem başarıyla geri alındı.", "success")
    except ValidationError as e:
        OperationLogService.log(
            module='kiralama',
            action='iptal_sonlandirma',
            user_id=getattr(current_user, 'id', None),
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=request.form.get('kalem_id', type=int),
            description=f"Kalem sonlandırma iptal doğrulama hatası: {str(e)}",
            success=False,
        )
        flash(f"Hata: {str(e)}", "warning")
    except Exception as e:
        current_app.logger.error(f"Sonlandırma İptal Hatası: {str(e)}")
        OperationLogService.log(
            module='kiralama',
            action='iptal_sonlandirma',
            user_id=getattr(current_user, 'id', None),
            username=getattr(current_user, 'username', None),
            entity_type='KiralamaKalemi',
            entity_id=request.form.get('kalem_id', type=int),
            description=f"Kalem sonlandırma iptal sistem hatası: {str(e)}",
            success=False,
        )
        flash(f"İşlem geri alınamadı.", "danger")
    return redirect(url_for('kiralama.index'))
@kiralama_bp.route('/api/ekipman-filtrele')
def api_ekipman_filtrele():
    try:
        # Sadece bizim olan (Tedarikçi olmayan), aktif ve boşta olan makineler
        query = Ekipman.query.filter_by(is_active=True, firma_tedarikci_id=None, calisma_durumu='bosta')
        
        # Filtreleri yakala
        sube_id = request.args.get('sube_id', type=int)
        tip = request.args.get('tip')
        marka = request.args.get('marka')
        enerji = request.args.get('enerji')
        ortam = request.args.get('ortam')
        
        y_min = request.args.get('y_min', type=float)
        y_max = request.args.get('y_max', type=float)
        k_min = request.args.get('k_min', type=float)
        
        agirlik_max = request.args.get('agirlik_max', type=float)
        genislik_max = request.args.get('genislik_max', type=float)
        uzunluk_max = request.args.get('uzunluk_max', type=float)
        ky_max = request.args.get('ky_max', type=float)
        
        # Sorguları uygula
        if sube_id: query = query.filter(Ekipman.sube_id == sube_id)
        if tip: query = query.filter(Ekipman.tipi == tip)
        if marka: query = query.filter(Ekipman.marka == marka)
        if enerji: query = query.filter(Ekipman.yakit == enerji)
        if ortam == 'ic': query = query.filter(Ekipman.ic_mekan_uygun == True)
        
        if y_min: query = query.filter(Ekipman.calisma_yuksekligi >= y_min)
        if y_max: query = query.filter(Ekipman.calisma_yuksekligi <= y_max)
        if k_min: query = query.filter(Ekipman.kaldirma_kapasitesi >= k_min)
        if agirlik_max: query = query.filter(Ekipman.agirlik <= agirlik_max)
        if genislik_max: query = query.filter(Ekipman.genislik <= genislik_max)
        if uzunluk_max: query = query.filter(Ekipman.uzunluk <= uzunluk_max)
        if ky_max: query = query.filter(Ekipman.kapali_yukseklik <= ky_max)
        
        ekipmanlar = query.order_by(Ekipman.kod).all()
        
        data = []
        for e in ekipmanlar:
            # Şube varsa adını, yoksa 'Şubesiz' yazar. Değişken adı temizlendi.
            gecici_sube_adi = e.sube.isim if e.sube else 'Merkez / Şubesiz'
            
            data.append({
                'id': e.id,
                'label': f"{e.kod} | {e.tipi} ({e.calisma_yuksekligi}m) - {gecici_sube_adi}"
            })
            
        return jsonify({
            'success': True,
            'count': len(data),
            'data': data
        })
    except Exception as e:
        # Gerçek hatayı ekrana basması için 'error' anahtarı eklendi
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== EXCEL EXPORT/IMPORT ====================

@kiralama_bp.route('/excel-disari-aktar', methods=['GET'])
@login_required
def excel_disari_aktar():
    """Tüm kiralama kaydını Excel'e aktar."""
    try:
        kiralamalar = Kiralama.query.filter(Kiralama.is_active == True).options(
            joinedload(Kiralama.firma_musteri),
            joinedload(Kiralama.kalemler)
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Kiralamalar'

        headers = [
            'Form No', 'Müşteri', 'Başlangıç Tarihi', 'Bitiş Tarihi', 
            'Birim Fiyat', 'Satış Fiyat', 'KDV Oranı (%)', 'Makine Kodu',
            'Makine Tipi', 'Makine Marka', 'Dış Tedarik Mi', 'Sonlanmış mı'
        ]
        ws.append(headers)

        for kiralama in kiralamalar:
            for kalem in kiralama.kalemler:
                makine_kodu = kalem.ekipman.kod if kalem.ekipman else (kalem.harici_ekipman_tipi or '')
                makine_tipi = kalem.ekipman.tipi if kalem.ekipman else (kalem.harici_ekipman_tipi or '')
                makine_marka = kalem.ekipman.marka if kalem.ekipman else (kalem.harici_ekipman_marka or '')
                
                ws.append([
                    kiralama.kiralama_form_no,
                    kiralama.firma_musteri.firma_adi if kiralama.firma_musteri else '',
                    kalem.kiralama_baslangici,
                    kalem.kiralama_bitis,
                    float(kalem.kiralama_brm_fiyat or 0),
                    float(kalem.kiralama_alis_fiyat or 0),
                    kiralama.kdv_orani,
                    makine_kodu,
                    makine_tipi,
                    makine_marka,
                    'Evet' if kalem.is_dis_tedarik_ekipman else 'Hayir',
                    'Evet' if kalem.sonlandirildi else 'Hayir',
                ])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"kiralamalar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        flash(f'Excel dışa aktarım başarısız: {e}', 'danger')
        return redirect(url_for('kiralama.index'))


@kiralama_bp.route('/excel-ice-yukle', methods=['POST'])
@login_required
def excel_ice_yukle():
    """Excel'den kiralama kaydını yükle (güncelle)."""
    confirm_password = (
        request.form.get('confirm_password')
        or request.form.get('password')
        or request.form.get('sifre')
        or ''
    ).strip()
    if not (
        getattr(current_user, 'is_authenticated', False)
        and hasattr(current_user, 'check_password')
        and current_user.check_password(confirm_password)
    ):
        flash('Excel içe aktarım için kullanıcı şifrenizi doğru girmeniz gerekiyor.', 'danger')
        return redirect(url_for('kiralama.index'))

    file = request.files.get('excel_file')
    if not file or not file.filename:
        flash('Lütfen bir Excel dosyası seçiniz.', 'warning')
        return redirect(url_for('kiralama.index'))

    if not file.filename.lower().endswith('.xlsx'):
        flash('Sadece .xlsx uzantılı dosyalar destekleniyor.', 'danger')
        return redirect(url_for('kiralama.index'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as exc:
        flash(f'Excel dosyası okunamadı: {exc}', 'danger')
        return redirect(url_for('kiralama.index'))

    updated = 0
    skipped = 0
    errors = []

    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_idx, row in enumerate(rows, start=2):
        try:
            form_no = (row[0] or '').strip() if row[0] else ''
            
            if not form_no:
                skipped += 1
                continue

            mevcut_kiralama = Kiralama.query.filter_by(kiralama_form_no=form_no).first()
            if not mevcut_kiralama:
                errors.append(f'Satır {row_idx}: Form No {form_no} bulunamadı.')
                continue

            # KDV Oranı güncelle
            try:
                mevcut_kiralama.kdv_orani = int(row[6]) if row[6] else 20
            except (ValueError, TypeError):
                mevcut_kiralama.kdv_orani = 20

            updated += 1

        except Exception as exc:
            errors.append(f'Satır {row_idx}: {exc}')

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f'Excel içe aktarma başarısız: {exc}', 'danger')
        return redirect(url_for('kiralama.index'))

    message = f'Excel içe aktarım tamamlandı. Güncellenen: {updated}, Atlanan: {skipped}.'
    if errors:
        message += f' Hatalar: {", ".join(errors[:5])}'
    
    flash(message, 'success' if not errors else 'warning')
    return redirect(url_for('kiralama.index'))

