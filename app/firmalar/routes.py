from app.firmalar import firmalar_bp
from flask_login import login_required
from app.utils import admin_required, bugun
# -------------------------------------------------------------------------
# Sözleşme No Düzelt (Admin)
# -------------------------------------------------------------------------
import os
from io import BytesIO
from flask import render_template, url_for, redirect, flash, request, current_app, send_file
from datetime import date
from flask_login import current_user, login_required
from decimal import Decimal
from sqlalchemy import asc, desc
from sqlalchemy.orm import joinedload, subqueryload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.firmalar import firmalar_bp
from app.firmalar.forms import FirmaForm
from app.firmalar.models import Firma
from app.filo.models import Ekipman
from app.kiralama.models import Kiralama, KiralamaKalemi
from app.extensions import db
from app.services.firma_services import FirmaService
from app.services.base import ValidationError
from app.services.cari_window import build_period_filtered_cari

# --- GÜVENLİK YARDIMCISI ---
def get_actor_id():
    """Kullanıcı giriş yapmışsa ID'sini döner, aksi halde None."""
    return getattr(current_user, 'id', None)


def _build_cari_rows(firma, today_date):
    """FirmaService.build_cari_rows için ince sarmalayıcı — geriye dönük uyumluluk."""
    return FirmaService.build_cari_rows(firma, today_date)


def _durum_metni_ve_rengi_bakiye(bakiye_val):
    b = Decimal(str(bakiye_val or 0))
    if b > 0:
        return 'Borçlu', 'text-danger'
    if b < 0:
        return 'Alacaklı', 'text-success'
    return 'Hesap Kapalı', 'text-muted'


def _resolve_bilgi_start_end_dates(request_args, firma_pre, today_d):
    """bilgi/yazdır/excel tarih parametreleri (bilgi görünümü ile aynı kurallar)."""
    from datetime import datetime as _dt_mod
    end_date_str = request_args.get('end_date')
    start_date_str = request_args.get('start_date')
    if end_date_str:
        try:
            end_date = _dt_mod.strptime(end_date_str, '%Y-%m-%d').date()
        except Exception:
            end_date = today_d
    else:
        end_date = today_d

    if start_date_str:
        try:
            start_date = _dt_mod.strptime(start_date_str, '%Y-%m-%d').date()
        except Exception:
            start_date = today_d
    else:
        start_date = FirmaService.firma_en_erken_islem_gunu(firma_pre.id)
        if start_date is None and firma_pre and getattr(firma_pre, 'created_at', None):
            ca = firma_pre.created_at
            start_date = ca.date() if hasattr(ca, 'date') else None
        if start_date is None:
            start_date = today_d

    return start_date, end_date


def _cari_period_mode_active(start_date_str, end_date_str):
    """URL'de iki tarih + config ile dönem modu."""
    if not current_app.config.get('CARI_DONEM_FILTRESI_ENABLED', True):
        return False
    s = (start_date_str or '').strip()
    e = (end_date_str or '').strip()
    return bool(s and e)


def _parse_kiralama_tab_iso_dates(start_date_str, end_date_str):
    """Kiralama sekmesi tarih süzgeci — cariden bağımsız: iki parametre dolu ve geçerliyse (start, end), değilse None."""
    from datetime import datetime as _dt
    if not start_date_str or not end_date_str:
        return None
    sa = str(start_date_str).strip()
    eb = str(end_date_str).strip()
    if not sa or not eb:
        return None
    try:
        start_date = _dt.strptime(sa, '%Y-%m-%d').date()
        end_date = _dt.strptime(eb, '%Y-%m-%d').date()
    except ValueError:
        return None
    if start_date > end_date:
        return None
    return start_date, end_date


def _sum_cari_rows_net(rows):
    net = Decimal('0')
    for row in rows:
        net += Decimal(str(row.get('toplam') or 0))
    return net


def _cari_yazdir_footer_totals(cari_rows, devreden_bakiye=None):
    """Yazdır/PDF Cari tfoot: borç, alacak (pozitif), kapanış bakiyesi (devreden dahil)."""
    borc = Decimal('0')
    alacak = Decimal('0')
    for row in cari_rows or []:
        t = Decimal(str(row.get('toplam') or 0))
        if t > 0:
            borc += t
        elif t < 0:
            alacak += abs(t)
    bakiye = borc - alacak
    if devreden_bakiye is not None:
        bakiye += Decimal(str(devreden_bakiye))
    return float(borc), float(alacak), float(bakiye)


def _build_kiralama_bilgi_tab(firma, start_date_str, end_date_str, kiralama_page, kiralama_per_page):
    """Firma bilgi > Kiralamalar sekmesi: liste + sayfalama (oluşturma tarihi yaklaşımı A)."""
    allowed_pp = {10, 25, 50, 100}
    if kiralama_per_page not in allowed_pp:
        kiralama_per_page = 25
    kiralamalar_filtered, filt_aktif, k_f_bas, k_f_bit = _filter_kiralamalar_bilgi(
        firma,
        start_date_str,
        end_date_str,
        key_fn=lambda k: k.kiralama_form_no or '',
    )

    k_pag = ListPagination(total=len(kiralamalar_filtered), page=kiralama_page, per_page=kiralama_per_page)
    kb = (k_pag.page - 1) * k_pag.per_page
    kiralamalar_sayfa = kiralamalar_filtered[kb : kb + k_pag.per_page]
    return {
        'kiralama_tarih_filtre_aktif': filt_aktif,
        'kiralama_filt_baslangic': k_f_bas,
        'kiralama_filt_bitis': k_f_bit,
        'kiralama_pagination': k_pag,
        'kiralamalar_sayfa': kiralamalar_sayfa,
        'kiralama_per_page': kiralama_per_page,
    }


def _filter_kiralamalar_bilgi(firma, start_date_str, end_date_str, key_fn=None):
    key_fn = key_fn or (lambda k: k.id or 0)
    kiralamalar_all = sorted(firma.kiralamalar, key=key_fn, reverse=True)
    rng = _parse_kiralama_tab_iso_dates(start_date_str, end_date_str)
    if rng:
        kiralamalar_filtered = _kiralamalar_filtered_by_olusturma_tarihi(kiralamalar_all, rng[0], rng[1])
        filt_aktif = True
        k_f_bas, k_f_bit = rng
    else:
        kiralamalar_filtered = list(kiralamalar_all)
        filt_aktif = False
        k_f_bas = k_f_bit = None
    return kiralamalar_filtered, filt_aktif, k_f_bas, k_f_bit


def _effective_kiralama_bilgi_tarih(k):
    """
    Kiralamalar bilgi tarih süzgecinde kullanılacak tek bir karşılık günü.
    Öncelik: form oluşturma tarihi → kalemlerdeki en erken kiralama başlangıcı → kayıt created_at.
    """
    from datetime import datetime as _dt

    olu = getattr(k, 'kiralama_olusturma_tarihi', None)
    if olu is not None:
        if isinstance(olu, date):
            return olu
        try:
            return _dt.strptime(str(olu), '%Y-%m-%d').date()
        except ValueError:
            pass

    kalemler = getattr(k, 'kalemler', None) or []
    starts = []
    for km in kalemler:
        b = getattr(km, 'kiralama_baslangici', None)
        if b is None:
            continue
        if isinstance(b, date):
            starts.append(b)
        else:
            try:
                starts.append(_dt.strptime(str(b), '%Y-%m-%d').date())
            except ValueError:
                pass
    if starts:
        return min(starts)

    ca = getattr(k, 'created_at', None)
    if ca is not None:
        try:
            if hasattr(ca, 'date'):
                d = ca.date()
                return d if isinstance(d, date) else None
        except Exception:
            return None
    return None


def _kiralamalar_filtered_by_olusturma_tarihi(kiralamalar_all, start_date, end_date):
    """
    Tarih aralığı süzümü — _effective_kiralama_bilgi_tarih ile karşılık günü [start_date, end_date].
    Çıkarılamayan tarihli kayıt filtreliyken gösterilmez.
    """
    out = []
    for k in kiralamalar_all:
        t = _effective_kiralama_bilgi_tarih(k)
        if t is None:
            continue
        if start_date <= t <= end_date:
            out.append(k)
    return out


def _get_firma_bilgi_context(firma_id):
    """Firma bilgi ekrani icin cari satir kaynagini onden yukler."""
    firma = Firma.query.options(
        subqueryload(Firma.kiralamalar).options(
            joinedload(Kiralama.firma_musteri),
            subqueryload(Kiralama.kalemler).options(
                joinedload(KiralamaKalemi.ekipman).joinedload(Ekipman.sube),
                joinedload(KiralamaKalemi.dondurmalar),
            ),
            subqueryload(Kiralama.nakliyeler),
        ),
        subqueryload(Firma.hizmet_kayitlari),
    ).filter_by(id=firma_id).first()
    if not firma:
        raise ValidationError("Firma bulunamadi.")
    return {
        'firma': firma,
        'bakiye': Decimal('0'),
        'toplam_borc': Decimal('0'),
        'toplam_alacak': Decimal('0'),
        'guncel_bakiye': Decimal('0'),
        'durum_metni': '',
        'durum_rengi': '',
    }


class ListPagination:
    """In-memory list için sayfalama yardımcısı."""
    def __init__(self, total, page, per_page):
        self.total = max(0, int(total or 0))
        self.per_page = max(1, int(per_page or 1))
        self.pages = (self.total + self.per_page - 1) // self.per_page if self.total else 0
        self.page = max(1, min(int(page or 1), self.pages or 1))

    @property
    def has_prev(self): return self.page > 1
    @property
    def has_next(self): return self.page < self.pages
    @property
    def prev_num(self): return self.page - 1
    @property
    def next_num(self): return self.page + 1

    def iter_pages(self, left_edge=1, left_current=1, right_current=2, right_edge=1):
        last = 0
        for num in range(1, self.pages + 1):
            if (
                num <= left_edge
                or (self.page - left_current - 1 < num < self.page + right_current)
                or num > self.pages - right_edge
            ):
                if last + 1 != num:
                    yield None
                yield num
                last = num


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {'1', 'true', 'evet', 'yes', 'x'}:
        return True
    if text in {'0', 'false', 'hayir', 'hayır', 'no', ''}:
        return False
    return default

# -------------------------------------------------------------------------
# 1. Aktif Firma Listeleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/')
@firmalar_bp.route('/index')
@login_required
def index():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        if per_page not in {10, 25, 50, 100}:
            per_page = 25
        q = request.args.get('q', '', type=str)
        sort_by = request.args.get('sort_by', 'firma_adi', type=str)
        sort_dir = request.args.get('sort_dir', 'asc', type=str)

        allowed_sort_fields = {
            'firma_adi': Firma.firma_adi,
            'yetkili_adi': Firma.yetkili_adi,
            'vergi_no': Firma.vergi_no
        }
        if sort_by not in allowed_sort_fields:
            sort_by = 'firma_adi'
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'asc'
        
        query = FirmaService.get_active_firms(search_query=q)
        sort_column = allowed_sort_fields[sort_by]
        sort_expression = asc(sort_column) if sort_dir == 'asc' else desc(sort_column)
        query = query.order_by(None).order_by(sort_expression, desc(Firma.id))
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        from app.firmalar.forms import SozlesmeNoDuzeltForm
        return render_template(
            'firmalar/index.html',
            firmalar=pagination.items,
            pagination=pagination,
            per_page=per_page,
            q=q,
            sort_by=sort_by,
            sort_dir=sort_dir,
            form=SozlesmeNoDuzeltForm()
        )
    except Exception as e:
        current_app.logger.error(f"Firma listesi yüklenirken hata: {str(e)}")
        flash("Firma listesi şu an yüklenemiyor.", "danger")
        return redirect(url_for('main.index')) # Ana sayfaya yönlendir

# -------------------------------------------------------------------------
# 2. Pasif (Arşivlenmiş) Firma Listeleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/pasif')
@login_required
def pasif_index():
    try:
        page = request.args.get('page', 1, type=int)
        q = request.args.get('q', '', type=str)
        
        query = FirmaService.get_inactive_firms(search_query=q)
        pagination = query.paginate(page=page, per_page=50, error_out=False)
        
        return render_template('firmalar/pasif_index.html', firmalar=pagination.items, pagination=pagination, q=q)
    except Exception as e:
        current_app.logger.error(f"Pasif firma listesi hatası: {str(e)}")
        flash("Arşivlenmiş firmalar yüklenemedi.", "danger")
        return redirect(url_for('firmalar.index'))

# -------------------------------------------------------------------------
# 3. Yeni Firma Ekleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
def ekle():
    form = FirmaForm()
    
    if form.validate_on_submit():
        try:
            # Servis katmanında tanımlı güncellenebilir alanları süzerek al
            data = {k: v for k, v in form.data.items() if k in FirmaService.updatable_fields}
            
            # Başlangıç değerlerini set et
            data.update({
                'bakiye': Decimal('0'),
                'sozlesme_rev_no': 0,
                'sozlesme_no': None
            })
            
            yeni_firma = Firma(**data)
            FirmaService.save(yeni_firma, actor_id=get_actor_id())
            
            flash(f"'{yeni_firma.firma_adi}' başarıyla sisteme kaydedildi.", "success")
            return redirect(url_for('firmalar.index'))
        except ValidationError as e:
            flash(str(e), "warning")
        except Exception as e:
            current_app.logger.error(f"Firma ekleme hatası: {str(e)}")
            flash("Kayıt sırasında sistemsel bir hata oluştu.", "danger")
            
    referrer = request.referrer or url_for('firmalar.index')
    next_ps_no = FirmaService.get_next_sozlesme_no()
    return render_template('firmalar/ekle.html', form=form, today_date=date.today().strftime('%d.%m.%Y'),
                           referrer=referrer, next_ps_no=next_ps_no)

# -------------------------------------------------------------------------
# 4. Sözleşme Hazırla
# -------------------------------------------------------------------------
@firmalar_bp.route('/sozlesme-hazirla/<int:id>', methods=['POST'])
@login_required
def sozlesme_hazirla(id):
    try:
        app_path = os.path.join(os.getcwd(), 'app')
        FirmaService.sozlesme_hazirla(firma_id=id, base_app_path=app_path, actor_id=get_actor_id())
        flash("Sözleşme numarası başarıyla atandı ve arşiv klasörleri oluşturuldu.", "success")
    except ValidationError as e:
        flash(str(e), "warning")
    except Exception as e:
        current_app.logger.error(f"Sözleşme hazırlama hatası (Firma ID: {id}): {str(e)}")
        flash("Klasör yapısı oluşturulurken bir hata meydana geldi.", "danger")
    return redirect(url_for('firmalar.index'))

# -------------------------------------------------------------------------
# 5. Firma Düzenleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/duzelt/<int:id>', methods=['GET', 'POST'])
@login_required
def duzelt(id):
    firma = FirmaService.get_by_id(id)
    if not firma:
        flash("Düzenlenmek istenen firma bulunamadı!", "danger")
        return redirect(url_for('firmalar.index'))
        
    form = FirmaForm(obj=firma)
    
    # Özel alanları form verisine elle eşle (Form yapısına göre)
    if request.method == 'GET':
        form.genel_sozlesme_no.data = firma.sozlesme_no
        form.sozlesme_rev_no.data = firma.sozlesme_rev_no
        form.sozlesme_tarihi.data = firma.sozlesme_tarihi

    if form.validate_on_submit():
        try:
            data = {k: v for k, v in form.data.items() if k in FirmaService.updatable_fields}
            
            # Formdaki özel alanları veritabanı alanlarıyla eşleştir
            data.update({
                'sozlesme_no': form.genel_sozlesme_no.data,
                'sozlesme_rev_no': form.sozlesme_rev_no.data,
                'sozlesme_tarihi': form.sozlesme_tarihi.data
            })
            
            FirmaService.update(id, data, actor_id=get_actor_id())
            flash(f"'{firma.firma_adi}' bilgileri başarıyla güncellendi.", 'success')
            return redirect(url_for('firmalar.index'))
        except ValidationError as e:
            flash(str(e), "danger")
        except Exception as e:
            current_app.logger.error(f"Firma güncelleme hatası (ID: {id}): {str(e)}")
            flash("Bilgiler güncellenirken bir hata oluştu.", "danger")
            
    return render_template('firmalar/duzelt.html', form=form, firma=firma)

# -------------------------------------------------------------------------
# 6. Firma Bilgi / Cari Ekstre
# -------------------------------------------------------------------------
@firmalar_bp.route('/bilgi/<int:id>', methods=['GET'])
@login_required
def bilgi(id):
    try:
        tab = request.args.get('tab', 'cari')
        if tab in ('genel', 'hareket'):
            tab = 'cari'

        hareket_per_page = request.args.get('hareket_per_page', 25, type=int)
        hareket_page     = request.args.get('hareket_page', 1, type=int)
        kiralama_per_page = request.args.get('kiralama_per_page', 25, type=int)
        kiralama_page     = request.args.get('kiralama_page', 1, type=int)
        cari_per_page    = request.args.get('cari_per_page', 25, type=int)
        cari_page        = request.args.get('cari_page', 1, type=int)

        allowed_pp = {10, 25, 50, 100}
        if hareket_per_page not in allowed_pp:   hareket_per_page = 25
        if kiralama_per_page not in allowed_pp:  kiralama_per_page = 25
        if cari_per_page not in allowed_pp:      cari_per_page = 25

        # --- Tarih filtresi ---
        from datetime import datetime
        today_date = bugun()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        finans_verileri = _get_firma_bilgi_context(id)
        firma_pre = finans_verileri.get('firma')

        start_date, end_date = _resolve_bilgi_start_end_dates(request.args, firma_pre, today_date)

        bilgi_sd_input = start_date_str or (start_date.strftime('%Y-%m-%d') if start_date else '')
        bilgi_ed_input = end_date_str or (end_date.strftime('%Y-%m-%d') if end_date else '')

        cari_period_active = _cari_period_mode_active(start_date_str, end_date_str)

        # Form Bazlı Cari sekmesi kaldırıldı: yalnızca mevcut cari/kiralama akışları kullanılıyor.
        finans_verileri.pop('hareketler', None)
        hareketler_filtered = []
        h_pag = ListPagination(total=0, page=1, per_page=hareket_per_page)
        hareketler_sayfa = []

        # --- LOG: Taşeron Nakliye Bedeli ---
        try:
            nakliye_hareketler = [h for h in hareketler_filtered if (getattr(h, 'tur', None) == 'Nakliye') or (isinstance(h, dict) and h.get('tur') == 'Nakliye')]
            for i, islem in enumerate(nakliye_hareketler):
                if isinstance(islem, dict):
                    log_str = f"[NAKLIYE] Hareket {i+1}: id={islem.get('id')}, aciklama={islem.get('aciklama')}, kdv_orani={islem.get('kdv_orani')}, tutar={islem.get('tutar')}"
                else:
                    log_str = f"[NAKLIYE] Hareket {i+1}: id={getattr(islem, 'id', None)}, aciklama={getattr(islem, 'aciklama', None)}, kdv_orani={getattr(islem, 'kdv_orani', None)}, tutar={getattr(islem, 'tutar', None)}"
                current_app.logger.debug(log_str)
        except Exception as log_ex:
            current_app.logger.warning(f"[NAKLIYE] LOG sırasında hata: {log_ex}")

        # --- Kiralamalar (cariden bağımsız süzüm; yalnızca URL'de iki ISO tarih geçilirse süzülür) ---
        firma = finans_verileri['firma']
        kiralama_tab_ctx = _build_kiralama_bilgi_tab(
            firma, start_date_str, end_date_str, kiralama_page, kiralama_per_page
        )
        k_pag = kiralama_tab_ctx['kiralama_pagination']
        kiralamalar_sayfa = kiralama_tab_ctx['kiralamalar_sayfa']
        kiralama_per_page = kiralama_tab_ctx['kiralama_per_page']

        # --- Cari Tab ---
        cari_donem_modu = False
        cari_devreden_bakiye = None
        firma_genel_bakiye = Decimal(str(firma.cari_bakiye_kdvli or 0))
        try:
            cari_rows_raw = _build_cari_rows(firma, end_date)
            if cari_period_active and end_date != today_date:
                firma_genel_bakiye = _sum_cari_rows_net(_build_cari_rows(firma, today_date))
            else:
                firma_genel_bakiye = _sum_cari_rows_net(cari_rows_raw)

            def _row_date(row):
                return row.get('sort_date') or row.get('baslangic') or row.get('form_tarihi')

            def _row_sort_key_asc(row):
                d = _row_date(row)
                dd = d or date.min
                try:
                    if hasattr(dd, 'date') and callable(getattr(dd, 'date')):
                        dd = dd.date()
                except Exception:
                    pass
                rid = row.get('id')
                try:
                    ri = int(rid) if rid is not None else 0
                except (TypeError, ValueError):
                    ri = 0
                return (dd, ri)

            if cari_period_active:
                (
                    cari_rows_all,
                    _opening_snap,
                    cari_toplam_borc,
                    cari_toplam_alacak,
                    cari_guncel_bakiye,
                ) = build_period_filtered_cari(cari_rows_raw, start_date, end_date)
                cari_donem_modu = True
                cari_devreden_bakiye = _opening_snap
            else:
                opening_balance = Decimal('0')
                filtered_running = opening_balance
                rows_old_to_new = sorted(cari_rows_raw, key=_row_sort_key_asc)
                for row in rows_old_to_new:
                    filtered_running += Decimal(str(row.get('toplam') or 0))
                    row['bakiye'] = float(filtered_running)

                cari_rows_all = list(reversed(rows_old_to_new))

                cari_toplam_borc = Decimal('0')
                cari_toplam_alacak = Decimal('0')
                for row in cari_rows_all:
                    t = Decimal(str(row.get('toplam') or 0))
                    if t > 0:
                        cari_toplam_borc += t
                    elif t < 0:
                        cari_toplam_alacak += t
                cari_guncel_bakiye = cari_toplam_borc + cari_toplam_alacak

            dm, dr = _durum_metni_ve_rengi_bakiye(cari_guncel_bakiye)
            finans_verileri['durum_metni'] = dm
            finans_verileri['durum_rengi'] = dr
            finans_verileri['toplam_borc'] = cari_toplam_borc
            finans_verileri['toplam_alacak'] = cari_toplam_alacak
            finans_verileri['guncel_bakiye'] = cari_guncel_bakiye
            finans_verileri['bakiye'] = cari_guncel_bakiye

            # Not: GET ekraninda cache yazma/commit yapmayiz.
            # Cache, veri degistiren akislar (POST/PUT/DELETE) tarafinda guncellenir.
        except Exception as cari_ex:
            current_app.logger.error(f"[CARI] _build_cari_rows hatası: {cari_ex}", exc_info=True)
            cari_rows_all = []
            cari_donem_modu = False
            cari_devreden_bakiye = None
            firma_genel_bakiye = Decimal(str(firma.cari_bakiye_kdvli or 0))
        c_pag = ListPagination(total=len(cari_rows_all), page=cari_page, per_page=cari_per_page)
        cb = (c_pag.page - 1) * c_pag.per_page
        ce = cb + c_pag.per_page
        cari_rows_sayfa = cari_rows_all[cb: ce]
        cari_goster_devreden_satiri = (
            cari_donem_modu
            and cari_devreden_bakiye is not None
            and (c_pag.pages == 0 or c_pag.page == c_pag.pages)
        )

        return render_template(
            'firmalar/bilgi.html',
            start_date=start_date,
            end_date=end_date,
            cari_donem_modu=cari_donem_modu,
            cari_devreden_bakiye=cari_devreden_bakiye,
            cari_goster_devreden_satiri=cari_goster_devreden_satiri,
            firma_genel_bakiye=firma_genel_bakiye,
            **finans_verileri,
            hareketler=hareketler_sayfa,
            hareket_pagination=h_pag,
            hareket_per_page=hareket_per_page,
            kiralamalar_sayfa=kiralamalar_sayfa,
            kiralama_pagination=k_pag,
            kiralama_per_page=kiralama_per_page,
            cari_rows=cari_rows_sayfa,
            cari_pagination=c_pag,
            cari_per_page=cari_per_page,
            tab=tab,
            today=today_date,
            kiralama_tarih_filtre_aktif=kiralama_tab_ctx['kiralama_tarih_filtre_aktif'],
            kiralama_filt_baslangic=kiralama_tab_ctx['kiralama_filt_baslangic'],
            kiralama_filt_bitis=kiralama_tab_ctx['kiralama_filt_bitis'],
            bilgi_sd_input=bilgi_sd_input,
            bilgi_ed_input=bilgi_ed_input,
            kiralama_duzenle_next=request.url,
            hareket_ph_page=h_pag.page,
            hareket_ph_per_page=hareket_per_page,
            cari_ph_page=cari_page,
            cari_ph_per_page=cari_per_page,
        )
    except ValidationError as e:
        flash(str(e), "danger")
        return redirect(url_for('firmalar.index'))
    except Exception as e:
        current_app.logger.error(f"Cari özet yükleme hatası (Firma ID: {id}): {str(e)}", exc_info=True)
        flash("Finansal veriler şu an hesaplanamıyor.", "danger")
        return redirect(url_for('firmalar.index'))


@firmalar_bp.route('/bilgi/<int:id>/kiralama-listesi', methods=['GET'])
@login_required
def bilgi_kiralama_listesi(id):
    """Firma bilgi — Kiralamalar sekmesi HTML parçası (cari/dönem mantığından bağımsız süzüm)."""
    try:
        kiralama_page = request.args.get('kiralama_page', 1, type=int)
        kiralama_per_page = request.args.get('kiralama_per_page', 25, type=int)
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        hp = max(1, request.args.get('hareket_page', 1, type=int) or 1)
        hpp = request.args.get('hareket_per_page', 25, type=int)
        if hpp not in {10, 25, 50, 100}:
            hpp = 25
        cp = max(1, request.args.get('cari_page', 1, type=int) or 1)
        cpp = request.args.get('cari_per_page', 25, type=int)
        if cpp not in {10, 25, 50, 100}:
            cpp = 25

        fin_ctx = _get_firma_bilgi_context(id)
        firma = fin_ctx['firma']

        kt = _build_kiralama_bilgi_tab(
            firma, start_date_str, end_date_str, kiralama_page, kiralama_per_page
        )

        today_date = bugun()
        sd_r, ed_r = _resolve_bilgi_start_end_dates(request.args, firma, today_date)
        bilgi_sd_input = start_date_str or (sd_r.strftime('%Y-%m-%d') if sd_r else '')
        bilgi_ed_input = end_date_str or (ed_r.strftime('%Y-%m-%d') if ed_r else '')

        next_kwargs = {
            'id': id,
            'tab': 'kiralama',
            'kiralama_page': kt['kiralama_pagination'].page,
            'kiralama_per_page': kt['kiralama_per_page'],
            'hareket_page': hp,
            'hareket_per_page': hpp,
            'cari_page': cp,
            'cari_per_page': cpp,
        }
        if (start_date_str or '').strip():
            next_kwargs['start_date'] = (start_date_str or '').strip()
        if (end_date_str or '').strip():
            next_kwargs['end_date'] = (end_date_str or '').strip()
        kiralama_duzenle_next = url_for('firmalar.bilgi', **next_kwargs)

        return render_template(
            'firmalar/_bilgi_kiralama_fragment.html',
            firma=firma,
            kiralama_tarih_filtre_aktif=kt['kiralama_tarih_filtre_aktif'],
            kiralama_filt_baslangic=kt['kiralama_filt_baslangic'],
            kiralama_filt_bitis=kt['kiralama_filt_bitis'],
            kiralama_pagination=kt['kiralama_pagination'],
            kiralamalar_sayfa=kt['kiralamalar_sayfa'],
            kiralama_per_page=kt['kiralama_per_page'],
            bilgi_sd_input=bilgi_sd_input,
            bilgi_ed_input=bilgi_ed_input,
            hareket_ph_page=hp,
            hareket_ph_per_page=hpp,
            cari_ph_page=cp,
            cari_ph_per_page=cpp,
            kiralama_duzenle_next=kiralama_duzenle_next,
        )
    except ValidationError:
        return '', 404
    except Exception as ex:
        current_app.logger.error(f"Kiralama listesi parçası hatası (Firma ID: {id}): {ex}", exc_info=True)
        return '', 500


# -------------------------------------------------------------------------
# YAZDIR: Firma Cari Hareketleri (Tüm Kayıtlar, Sayfalama Yok)
# -------------------------------------------------------------------------
@firmalar_bp.route('/bilgi/<int:id>/yazdir')
@login_required
def bilgi_yazdir(id):
    try:
        tab = request.args.get('tab', 'cari')
        finans_verileri = FirmaService.get_financial_summary(id)
        hareketler = finans_verileri.pop('hareketler')
        firma = finans_verileri['firma']
        today_date = bugun()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        sd_r, ed_r = _resolve_bilgi_start_end_dates(request.args, firma, today_date)
        cari_export_period = tab == 'cari' and _cari_period_mode_active(start_date_str, end_date_str)
        if tab == 'kiralama':
            kiralamalar_all, filt_aktif, filt_bas, filt_bit = _filter_kiralamalar_bilgi(
                firma,
                start_date_str,
                end_date_str,
                key_fn=lambda k: k.id or 0,
            )
            if filt_aktif:
                rapor_tarihi = f"{filt_bas.strftime('%d.%m.%Y')} – {filt_bit.strftime('%d.%m.%Y')}"
            else:
                rapor_tarihi = today_date.strftime('%d.%m.%Y')
        else:
            kiralamalar_all = sorted(firma.kiralamalar, key=lambda k: k.id, reverse=True)

        cari_rows: list = []
        cari_devreden_print = None
        filt_start_print = None
        if tab == 'cari':
            build_end = ed_r if cari_export_period else today_date
            raw_rows = _build_cari_rows(firma, build_end)
            if cari_export_period:
                cari_rows, cari_devreden_print, tb, ta, gb = build_period_filtered_cari(raw_rows, sd_r, ed_r)
                filt_start_print = sd_r
                finans_verileri['toplam_borc'] = tb
                finans_verileri['toplam_alacak'] = ta
                finans_verileri['guncel_bakiye'] = gb
                rapor_tarihi = f"{sd_r.strftime('%d.%m.%Y')} – {ed_r.strftime('%d.%m.%Y')}"
            else:

                def _yz_sort(row):
                    d = row.get('sort_date') or row.get('baslangic') or row.get('form_tarihi')
                    dd = d or date.min
                    try:
                        if hasattr(dd, 'date') and callable(getattr(dd, 'date')):
                            dd = dd.date()
                    except Exception:
                        pass
                    return (dd, int(row.get('id') or 0))

                cari_rows = sorted(raw_rows, key=_yz_sort)
                rapor_tarihi = today_date.strftime('%d.%m.%Y')
        elif tab != 'kiralama':
            rapor_tarihi = today_date.strftime('%d.%m.%Y')

        cari_foot_borc = None
        cari_foot_alacak = None
        cari_foot_bakiye = None
        if tab == 'cari' and cari_rows:
            cari_foot_borc, cari_foot_alacak, cari_foot_bakiye = _cari_yazdir_footer_totals(
                cari_rows, cari_devreden_print
            )

        return render_template(
            'firmalar/yazdir.html',
            **finans_verileri,
            hareketler=hareketler,
            kiralamalar=kiralamalar_all,
            cari_rows=cari_rows,
            tab=tab,
            rapor_tarihi=rapor_tarihi,
            cari_export_period=bool(tab == 'cari' and cari_export_period),
            cari_devreden_bakiye=cari_devreden_print,
            filt_start_print=filt_start_print,
            cari_foot_borc=cari_foot_borc,
            cari_foot_alacak=cari_foot_alacak,
            cari_foot_bakiye=cari_foot_bakiye,
        )
    except Exception as e:
        current_app.logger.error(
            f"Yazdırma hatası (Firma ID: {id}): {e}",
            exc_info=True,
        )
        flash("Yazdırma verisi yüklenemedi.", "danger")
        return redirect(url_for('firmalar.bilgi', id=id))


@firmalar_bp.route('/bilgi/<int:id>/excel')
@login_required
def bilgi_excel(id):
    try:
        tab = request.args.get('tab', 'hareket')
        finans_verileri = FirmaService.get_financial_summary(id)
        hareketler = finans_verileri.pop('hareketler')
        firma = finans_verileri['firma']
        kiralamalar = sorted(firma.kiralamalar, key=lambda k: k.id, reverse=True)
        today_date = bugun()
        sd_s_x = request.args.get('start_date')
        ed_s_x = request.args.get('end_date')
        sd_x, ed_x = _resolve_bilgi_start_end_dates(request.args, firma, today_date)
        cari_ex_period = tab == 'cari' and _cari_period_mode_active(sd_s_x, ed_s_x)

        cari_rows: list = []
        if tab == 'cari':
            build_end_x = ed_x if cari_ex_period else today_date
            raw_x = _build_cari_rows(firma, build_end_x)
            if cari_ex_period:
                cari_rows, _, tb_x, ta_x, gb_x = build_period_filtered_cari(raw_x, sd_x, ed_x)
                finans_verileri['toplam_borc'] = tb_x
                finans_verileri['toplam_alacak'] = ta_x
                finans_verileri['guncel_bakiye'] = gb_x
                rapor_tarihi = f"{sd_x.strftime('%d.%m.%Y')} – {ed_x.strftime('%d.%m.%Y')}"
            else:
                def _x_sort(row):
                    d = row.get('sort_date') or row.get('baslangic') or row.get('form_tarihi')
                    dd = d or date.min
                    try:
                        if hasattr(dd, 'date') and callable(getattr(dd, 'date')):
                            dd = dd.date()
                    except Exception:
                        pass
                    return (dd, int(row.get('id') or 0))

                cari_rows = sorted(raw_x, key=_x_sort)
                rapor_tarihi = date.today().strftime('%d.%m.%Y')
        elif tab == 'kiralama':
            kiralamalar, filt_aktif_x, filt_bas_x, filt_bit_x = _filter_kiralamalar_bilgi(
                firma,
                sd_s_x,
                ed_s_x,
                key_fn=lambda k: k.id or 0,
            )
            if filt_aktif_x:
                rapor_tarihi = f"{filt_bas_x.strftime('%d.%m.%Y')} – {filt_bit_x.strftime('%d.%m.%Y')}"
            else:
                rapor_tarihi = date.today().strftime('%d.%m.%Y')
        else:
            rapor_tarihi = date.today().strftime('%d.%m.%Y')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Kiralamalar' if tab == 'kiralama' else 'Cari Hareketler'
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        title_font = Font(name='Calibri', size=14, bold=True, color='1F1F1F')
        meta_font = Font(name='Calibri', size=10, color='44546A')
        header_font = Font(name='Calibri', size=10, bold=True, color='18324A')
        body_font = Font(name='Calibri', size=10, color='1F1F1F')
        total_font = Font(name='Calibri', size=10, bold=True, color='1F1F1F')

        header_fill = PatternFill(fill_type='solid', fgColor='DBE7F3')
        total_fill = PatternFill(fill_type='solid', fgColor='F0F0F0')

        thin_blue = Side(style='thin', color='D9E2F0')
        dark_blue = Side(style='medium', color='18324A')

        header_border = Border(top=dark_blue, bottom=dark_blue)
        cell_border = Border(bottom=thin_blue)
        total_border = Border(top=dark_blue)

        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        def set_cell(row, column, value, font=None, fill=None, border=None, alignment=None, number_format=None):
            cell = sheet.cell(row=row, column=column, value=value)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
            if number_format:
                cell.number_format = number_format
            return cell

        def item_value(item, key, default=None):
            if isinstance(item, dict):
                return item.get(key, default)
            return getattr(item, key, default)

        if tab == 'kiralama':
            headers = ['#', 'Form No', 'Ekipmanlar', 'Durumları', 'Genel Durum']
            column_widths = {'A': 6, 'B': 18, 'C': 54, 'D': 18, 'E': 22}
            sheet.merge_cells('A1:E1')
            set_cell(1, 1, 'Kiralama Listesi', font=title_font, alignment=left_alignment)
            sheet.merge_cells('A2:C2')
            set_cell(2, 1, f'Firma: {firma.firma_adi}', font=meta_font, alignment=left_alignment)
            sheet.merge_cells('D2:E2')
            set_cell(2, 4, f'Rapor Tarihi: {rapor_tarihi}', font=meta_font, alignment=left_alignment)

            header_row = 4
            for col_idx, header in enumerate(headers, start=1):
                align = center_alignment if col_idx in {1, 4, 5} else left_alignment
                set_cell(header_row, col_idx, header, font=header_font, fill=header_fill, border=header_border, alignment=align)

            current_row = 5
            for index, kiralama in enumerate(kiralamalar, start=1):
                kalemler = list(kiralama.kalemler or [])
                toplam = len(kalemler)
                biten = len([kalem for kalem in kalemler if kalem.sonlandirildi])
                if biten == 0:
                    genel_durum = 'Tamamı Sahada'
                elif biten == toplam:
                    genel_durum = 'Sözleşme Kapandı'
                else:
                    genel_durum = f'Parçalı Teslimat\n({biten}/{toplam} makine döndü)'

                grup_satir_sayisi = max(1, toplam)
                grup_baslangic = current_row
                grup_bitis = current_row + grup_satir_sayisi - 1

                if grup_satir_sayisi > 1:
                    sheet.merge_cells(start_row=grup_baslangic, start_column=1, end_row=grup_bitis, end_column=1)
                    sheet.merge_cells(start_row=grup_baslangic, start_column=2, end_row=grup_bitis, end_column=2)
                    sheet.merge_cells(start_row=grup_baslangic, start_column=5, end_row=grup_bitis, end_column=5)

                set_cell(grup_baslangic, 1, index, font=body_font, border=cell_border, alignment=center_alignment)
                set_cell(grup_baslangic, 2, kiralama.kiralama_form_no or 'Belirtilmemiş', font=body_font, border=cell_border, alignment=left_alignment)
                set_cell(grup_baslangic, 5, genel_durum, font=body_font, border=cell_border, alignment=center_alignment)

                if not kalemler:
                    set_cell(grup_baslangic, 3, 'Ekipman bilgisi yok', font=body_font, border=cell_border, alignment=left_alignment)
                    set_cell(grup_baslangic, 4, '-', font=body_font, border=cell_border, alignment=left_alignment)
                else:
                    for satir_ofset, kalem in enumerate(kalemler):
                        satir_no = grup_baslangic + satir_ofset
                        ekipman = getattr(kalem, 'ekipman', None)
                        if ekipman:
                            marka = (getattr(ekipman, 'marka', '') or '').strip()
                            model = (getattr(ekipman, 'model', '') or '').strip()
                            ekipman_adi = f'{marka} - {model}'.strip(' -')
                            kod = (getattr(ekipman, 'kod', '') or '').strip()
                            ekipman_metni = f'({kod}) {ekipman_adi}'.strip() if kod else (ekipman_adi or getattr(ekipman, 'tipi', 'Ekipman bilgisi yok'))
                        else:
                            marka = (getattr(kalem, 'harici_ekipman_marka', '') or '').strip()
                            model = (getattr(kalem, 'harici_ekipman_model', '') or '').strip()
                            ekipman_adi = f'{marka} - {model}'.strip(' -')
                            ekipman_metni = f'(Dış Tedarik) {ekipman_adi}'.strip() if ekipman_adi else 'Ekipman bilgisi yok'

                        if getattr(kalem, 'is_dis_tedarik_ekipman', False) and getattr(kalem, 'harici_tedarikci', None):
                            ekipman_metni = f"{ekipman_metni}\nTedarikçi: {kalem.harici_tedarikci.firma_adi}"

                        if kalem.sonlandirildi:
                            durum_metni = f"Bitti\n{kalem.kiralama_baslangici.strftime('%d.%m.%Y')} - {kalem.kiralama_bitis.strftime('%d.%m.%Y')}"
                        else:
                            durum_metni = f"Aktif\n{kalem.kiralama_baslangici.strftime('%d.%m.%Y')} tarihinden beri sahada"

                        set_cell(satir_no, 3, ekipman_metni, font=body_font, border=cell_border, alignment=left_alignment)
                        set_cell(satir_no, 4, durum_metni, font=body_font, border=cell_border, alignment=left_alignment)

                current_row = grup_bitis + 1

            if kiralamalar:
                sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                set_cell(current_row, 1, f'TOPLAM: {len(kiralamalar)} kiralama kaydı', font=total_font, fill=total_fill, border=total_border, alignment=left_alignment)

        else:
            headers = ['#', 'Tarih', 'İşlem Türü', 'Açıklama', 'Matrah', 'KDV Oranı', 'KDV Tutarı', 'Borç', 'Alacak', 'Bakiye']
            column_widths = {'A': 6, 'B': 14, 'C': 16, 'D': 42, 'E': 14, 'F': 12, 'G': 14, 'H': 14, 'I': 14, 'J': 14}
            sheet.merge_cells('A1:J1')
            set_cell(1, 1, 'Cari Hesap Ekstresi', font=title_font, alignment=left_alignment)
            sheet.merge_cells('A2:E2')
            set_cell(2, 1, f'Firma: {firma.firma_adi}', font=meta_font, alignment=left_alignment)
            sheet.merge_cells('F2:J2')
            set_cell(2, 6, f'Rapor Tarihi: {rapor_tarihi}', font=meta_font, alignment=left_alignment)
            sheet.merge_cells('A3:D3')
            set_cell(3, 1, f"Toplam Borç: {finans_verileri.get('toplam_borc', 0):,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.'), font=meta_font, alignment=left_alignment)
            sheet.merge_cells('E3:G3')
            set_cell(3, 5, f"Toplam Alacak: {abs(finans_verileri.get('toplam_alacak', 0)) :,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.'), font=meta_font, alignment=left_alignment)
            sheet.merge_cells('H3:J3')
            set_cell(3, 8, f"Güncel Bakiye: {finans_verileri.get('guncel_bakiye', 0):,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.'), font=meta_font, alignment=left_alignment)

            header_row = 5
            for col_idx, header in enumerate(headers, start=1):
                align = center_alignment if col_idx in {1, 2, 3, 6} else right_alignment if col_idx >= 5 else left_alignment
                set_cell(header_row, col_idx, header, font=header_font, fill=header_fill, border=header_border, alignment=align)

            current_row = 6
            if tab == 'cari':
                for index, row in enumerate(cari_rows, start=1):
                    islem_tarih = row.get('sort_date') or row.get('baslangic') or row.get('form_tarihi')
                    islem_turu = row.get('islem_turu') or 'İşlem'
                    aciklama = row.get('aciklama') or ''
                    matrah = float(row.get('matrah') or 0) if row.get('matrah') is not None else None
                    kdv_orani = float(row.get('kdv_orani') or 0) if row.get('kdv_orani') is not None else None
                    kdv_tutari = float(row.get('kdv_tutar') or 0) if row.get('kdv_tutar') is not None else None
                    toplam = float(row.get('toplam') or 0)
                    borc = toplam if toplam > 0 else None
                    alacak = abs(toplam) if toplam < 0 else None
                    bakiye = float(row.get('bakiye') or 0)

                    row_data = [
                        index,
                        islem_tarih.strftime('%d.%m.%Y') if islem_tarih else '-',
                        islem_turu,
                        aciklama,
                        matrah,
                        kdv_orani,
                        kdv_tutari,
                        borc,
                        alacak,
                        bakiye,
                    ]

                    for col_idx, value in enumerate(row_data, start=1):
                        if col_idx in {5, 6, 7, 8, 9, 10} and value is not None:
                            number_format = '0.00' if col_idx == 6 else '#,##0.00'
                            set_cell(current_row, col_idx, value, font=body_font, border=cell_border, alignment=right_alignment, number_format=number_format)
                        else:
                            display_value = value if value not in (None, '') else '-'
                            align = center_alignment if col_idx in {1, 2, 3} else left_alignment
                            set_cell(current_row, col_idx, display_value, font=body_font, border=cell_border, alignment=align)
                    current_row += 1
            else:
                for index, islem in enumerate(hareketler, start=1):
                    islem_tur = item_value(islem, 'tur', '')
                    islem_ozel_id = item_value(islem, 'ozel_id')
                    islem_tur_tipi = item_value(islem, 'tur_tipi', '')
                    islem_aciklama = item_value(islem, 'aciklama', '')
                    islem_belge_no = item_value(islem, 'belge_no')
                    islem_tarih = item_value(islem, 'tarih')
                    islem_tutar = item_value(islem, 'tutar')
                    islem_kdv_tutari = item_value(islem, 'kdv_tutari')
                    islem_borc = item_value(islem, 'borc')
                    islem_alacak = item_value(islem, 'alacak')
                    islem_bakiye = item_value(islem, 'kumulatif_bakiye')
                    islem_kiralama_alis_kdv = item_value(islem, 'kiralama_alis_kdv')
                    islem_nakliye_alis_kdv = item_value(islem, 'nakliye_alis_kdv')
                    islem_kdv_orani = item_value(islem, 'kdv_orani')

                    is_kiralama = islem_tur == 'Kiralama' and islem_ozel_id
                    is_nakliye = islem_tur == 'Nakliye' and islem_ozel_id
                    if is_kiralama:
                        islem_turu = 'Kiralama'
                    elif is_nakliye:
                        islem_turu = 'Nakliye'
                    elif 'Fatura' in islem_tur or 'Hizmet' in islem_tur:
                        islem_turu = 'Fatura'
                    elif islem_tur == 'Tahsilat (Giriş)':
                        islem_turu = 'Tahsilat'
                    elif islem_tur == 'Ödeme (Çıkış)':
                        islem_turu = 'Ödeme'
                    else:
                        islem_turu = islem_tur

                    aciklama = islem_aciklama or ''
                    if islem_belge_no:
                        belge_metni = f"Kasa: {islem_belge_no}" if islem_tur_tipi == 'odeme' else f"No: {islem_belge_no}"
                        aciklama = f"{aciklama}\n{belge_metni}" if aciklama else belge_metni

                    if islem_kiralama_alis_kdv is not None:
                        kdv_orani = float(islem_kiralama_alis_kdv)
                    elif islem_nakliye_alis_kdv is not None:
                        kdv_orani = float(islem_nakliye_alis_kdv)
                    elif islem_kdv_orani is not None:
                        kdv_orani = float(islem_kdv_orani)
                    else:
                        kdv_orani = None

                    matrah = None
                    if islem_tur_tipi != 'odeme' and islem_tutar not in (None, 0):
                        matrah = abs(float(islem_tutar))

                    kdv_tutari = float(islem_kdv_tutari) if islem_kdv_tutari is not None else None
                    borc = float(islem_borc) if islem_borc not in (None, 0) else None
                    alacak = float(islem_alacak) if islem_alacak not in (None, 0) else None
                    bakiye = float(islem_bakiye) if islem_bakiye is not None else 0

                    row_data = [
                        index,
                        islem_tarih.strftime('%d.%m.%Y') if islem_tarih else '-',
                        islem_turu,
                        aciklama,
                        matrah,
                        kdv_orani,
                        kdv_tutari,
                        borc,
                        alacak,
                        bakiye,
                    ]

                    for col_idx, value in enumerate(row_data, start=1):
                        if col_idx in {5, 6, 7, 8, 9, 10} and value is not None:
                            number_format = '0.00' if col_idx == 6 else '#,##0.00'
                            set_cell(current_row, col_idx, value, font=body_font, border=cell_border, alignment=right_alignment, number_format=number_format)
                        else:
                            display_value = value if value not in (None, '') else '-'
                            align = center_alignment if col_idx in {1, 2, 3} else left_alignment
                            set_cell(current_row, col_idx, display_value, font=body_font, border=cell_border, alignment=align)
                    current_row += 1

            if (tab == 'cari' and cari_rows) or (tab != 'cari' and hareketler):
                sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                set_cell(current_row, 1, 'NET BAKİYE:', font=total_font, fill=total_fill, border=total_border, alignment=right_alignment)
                set_cell(current_row, 8, float(finans_verileri.get('toplam_borc', 0) or 0), font=total_font, fill=total_fill, border=total_border, alignment=right_alignment, number_format='#,##0.00')
                set_cell(current_row, 9, abs(float(finans_verileri.get('toplam_alacak', 0) or 0)), font=total_font, fill=total_fill, border=total_border, alignment=right_alignment, number_format='#,##0.00')
                set_cell(current_row, 10, float(finans_verileri.get('guncel_bakiye', 0) or 0), font=total_font, fill=total_fill, border=total_border, alignment=right_alignment, number_format='#,##0.00')

        for column_letter, width in column_widths.items():
            sheet.column_dimensions[column_letter].width = width

        max_row = sheet.max_row
        start_row = 5 if tab == 'kiralama' else 6
        for row_idx in range(start_row, max_row + 1):
            sheet.row_dimensions[row_idx].height = 42 if tab == 'kiralama' else 34

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        file_stub = 'firma_kiralama_listesi' if tab == 'kiralama' else 'firma_cari_ekstresi'
        file_name = f'{file_stub}_{id}_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        current_app.logger.error(f'Excel export hatası (Firma ID: {id}): {str(e)}')
        flash('Excel çıktısı hazırlanamadı.', 'danger')
        return redirect(url_for('firmalar.bilgi_yazdir', id=id, tab=request.args.get('tab', 'hareket')))


# -------------------------------------------------------------------------
# 7. Firmayı Arşive Kaldır (Silme - Kontrollü)
# -------------------------------------------------------------------------
@firmalar_bp.route('/sil/<int:id>', methods=['POST'])
@login_required
def sil(id):
    try:
        FirmaService.archive_with_check(id, actor_id=get_actor_id())
        flash("Firma başarıyla arşive kaldırıldı.", 'success')
    except ValidationError as e:
        flash(str(e), 'warning')
    except Exception as e:
        current_app.logger.error(f"Arşivleme hatası (ID: {id}): {str(e)}")
        flash("İşlem sırasında sistemsel bir hata oluştu.", 'danger')
    return redirect(url_for('firmalar.index'))

# -------------------------------------------------------------------------
# 8. Firmayı Aktifleştir (Arşivden Geri Al)
# -------------------------------------------------------------------------
@firmalar_bp.route('/aktiflestir/<int:id>', methods=['POST'])
@login_required
def aktiflestir(id):
    try:
        FirmaService.update(id, {'is_active': True}, actor_id=get_actor_id())
        flash("Firma başarıyla tekrar aktif hale getirildi.", "success")
    except Exception as e:
        current_app.logger.error(f"Aktifleştirme hatası (ID: {id}): {str(e)}")
        flash("İşlem başarısız oldu.", "danger")
    return redirect(url_for('firmalar.pasif_index'))

# -------------------------------------------------------------------------
# 9. İmza Yetkisi Kontrolü
# -------------------------------------------------------------------------
@firmalar_bp.route('/imza-kontrol/<int:id>', methods=['POST'])
@login_required
def imza_kontrol(id):
    try:
        FirmaService.update(
            id, 
            {'imza_yetkisi_kontrol_edildi': True, 'imza_yetkisi_kontrol_tarihi': date.today()}, 
            actor_id=get_actor_id()
        )
        flash("İmza yetkisi başarıyla onaylandı.", "success")
    except ValidationError as e:
        flash(str(e), "danger")
    except Exception as e:
        current_app.logger.error(f"İmza kontrol hatası (ID: {id}): {str(e)}")
        flash("Onay işlemi yapılamadı.", "danger")
    return redirect(url_for('firmalar.bilgi', id=id))


# -------------------------------------------------------------------------
# Sözleşme No Düzelt (Admin)
# -------------------------------------------------------------------------
from app.firmalar.forms import SozlesmeNoDuzeltForm

@firmalar_bp.route('/sozlesme_no_duzelt', methods=['POST'])
@login_required
@admin_required
def sozlesme_no_duzelt():
    form = SozlesmeNoDuzeltForm()
    if not form.validate_on_submit():
        flash('Form doğrulaması başarısız oldu. Lütfen tekrar deneyin.', 'danger')
        return redirect(url_for('firmalar.index'))
    firma_id = int(form.firma_id.data)
    yeni_sozlesme_no = form.sozlesme_no.data.strip()
    yeni_sozlesme_tarihi = form.sozlesme_tarihi.data
    firma = db.session.get(Firma, firma_id)
    if not firma:
        flash('Firma bulunamadı.', 'danger')
        return redirect(url_for('firmalar.index'))
    # Aynı sözleşme_no başka bir firmada var mı?
    if yeni_sozlesme_no:
        mevcut = Firma.query.filter(Firma.sozlesme_no == yeni_sozlesme_no, Firma.id != firma_id).first()
        if mevcut:
            flash('Bu sözleşme numarası başka bir firmada zaten kullanılıyor.', 'danger')
            return redirect(url_for('firmalar.index'))
    firma.sozlesme_no = yeni_sozlesme_no
    if yeni_sozlesme_tarihi:
        firma.sozlesme_tarihi = yeni_sozlesme_tarihi
    db.session.commit()
    flash('Sözleşme numarası ve tarihi başarıyla güncellendi.', 'success')
    return redirect(url_for('firmalar.index'))
