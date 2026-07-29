"""Nakliye soft-delete listeleme, silme ve geri alma testleri."""
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO

import pytest

from app.extensions import db
from app.firmalar.models import Firma
from app.cari.models import HizmetKaydi
from app.nakliyeler.models import Nakliye
from app.nakliyeler.routes import _nakliye_filtered_query
from app.services.nakliye_services import CariServis, NakliyeService
from app.services.kiralama_services import KiralamaService
from app.kiralama.models import Kiralama, KiralamaKalemi
from app.auth.models import User
from app.auth.session_security import (
    SESSION_LAST_PING_KEY,
    SESSION_TOKEN_KEY,
    new_session_token,
    utc_now,
)


def _vergi_no():
    return f"{uuid.uuid4().int % 10**10:010d}"


def _login_user(client, user_id: int) -> None:
    token = new_session_token()
    now = utc_now()
    user = db.session.get(User, user_id)
    user.active_session_token = token
    user.active_session_started_at = now
    user.active_session_seen_at = now
    db.session.commit()

    with client.session_transaction() as session:
        session["_user_id"] = str(user_id)
        session["_fresh"] = True
        session[SESSION_TOKEN_KEY] = token
        session[SESSION_LAST_PING_KEY] = now.isoformat()


def _firma(adi_prefix="Nak Soft", **kwargs):
    defaults = dict(
        firma_adi=f"{adi_prefix} {uuid.uuid4().hex[:4]}",
        yetkili_adi="Yetkili",
        iletisim_bilgileri="Adres",
        vergi_dairesi="Istanbul VD",
        vergi_no=_vergi_no(),
        is_musteri=True,
        is_tedarikci=True,
        bakiye=Decimal("0"),
        is_active=True,
    )
    defaults.update(kwargs)
    firma = Firma(**defaults)
    db.session.add(firma)
    db.session.flush()
    return firma


def _nakliye(firma, tutar="100.00", plaka="34ABC123", is_active=True, is_deleted=False, **kwargs):
    n = Nakliye(
        firma_id=firma.id,
        tarih=date.today(),
        islem_tarihi=date.today(),
        guzergah="Test guzergah",
        plaka=plaka,
        tutar=Decimal(tutar),
        toplam_tutar=Decimal(tutar),
        is_active=is_active,
        is_deleted=is_deleted,
        **kwargs,
    )
    db.session.add(n)
    db.session.flush()
    return n


def test_filtered_query_excludes_inactive_and_soft_deleted(app):
    with app.app_context():
        firma = _firma()
        aktif = _nakliye(firma, plaka="34AKTIF1")
        pasif = _nakliye(firma, plaka="34PASIF1", is_active=False)
        silinmis = _nakliye(firma, plaka="34SILIN1", is_active=False, is_deleted=True)
        db.session.commit()

        rows = _nakliye_filtered_query(None, None, None, None, str(firma.id)).all()
        ids = {r.id for r in rows}
        assert aktif.id in ids
        assert pasif.id not in ids
        assert silinmis.id not in ids


def test_plaka_dropdown_query_excludes_inactive(app):
    with app.app_context():
        firma = _firma()
        _nakliye(firma, plaka="34DROP01")
        _nakliye(firma, plaka="34DROP02", is_active=False, is_deleted=True)
        db.session.commit()

        plakalar = (
            db.session.query(Nakliye.plaka)
            .filter(*Nakliye.active_filters(), Nakliye.plaka.isnot(None))
            .distinct()
            .all()
        )
        values = {p[0] for p in plakalar}
        assert "34DROP01" in values
        assert "34DROP02" not in values


def test_nakliye_index_print_and_excel_exclude_soft_deleted(app, client):
    with app.app_context():
        user = User(username=f"nak_endpoint_{uuid.uuid4().hex[:6]}", rol="admin")
        user.set_password("Secret123!")
        db.session.add(user)
        firma = _firma("Nak Endpoint")
        aktif = _nakliye(firma, plaka="34HTMLAKT")
        silinmis = _nakliye(
            firma,
            plaka="34HTMLSIL",
            is_active=False,
            is_deleted=True,
        )
        db.session.commit()
        user_id = user.id
        firma_id = firma.id
        baslangic = date.today().isoformat()
        bitis = date.today().isoformat()

    _login_user(client, user_id)
    query = f"?baslangic={baslangic}&bitis={bitis}&firma_id={firma_id}"
    response = client.get(f"/nakliyeler/{query}")
    assert response.status_code == 200
    assert b"34HTMLAKT" in response.data
    assert b"34HTMLSIL" not in response.data

    response = client.get(f"/nakliyeler/yazdir{query}")
    assert response.status_code == 200
    assert b"34HTMLAKT" in response.data
    assert b"34HTMLSIL" not in response.data

    response = client.get(f"/nakliyeler/excel{query}")
    assert response.status_code == 200
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(response.data), read_only=True, data_only=True)
    values = [cell.value for row in workbook.active.iter_rows() for cell in row]
    assert "34HTMLAKT" in values
    assert "34HTMLSIL" not in values


def test_nakliye_cari_temizle_soft_deletes_hizmet(app):
    with app.app_context():
        firma = _firma()
        nakliye = _nakliye(firma)
        CariServis.musteri_nakliye_senkronize_et(nakliye)
        db.session.commit()

        hizmet = HizmetKaydi.query.filter_by(nakliye_id=nakliye.id, yon="giden").one()
        CariServis.nakliye_cari_temizle(nakliye.id)
        db.session.commit()

        db.session.refresh(hizmet)
        assert db.session.get(HizmetKaydi, hizmet.id) is not None
        assert hizmet.is_deleted is True
        assert hizmet.is_active is False


def test_taseron_maliyet_soft_deletes_instead_of_hard(app):
    with app.app_context():
        musteri = _firma("Musteri")
        taseron = _firma("Taseron", is_musteri=False, is_tedarikci=True)
        nakliye = _nakliye(musteri)
        nakliye.nakliye_tipi = "taseron"
        nakliye.taseron_firma_id = taseron.id
        nakliye.taseron_maliyet = Decimal("50.00")
        CariServis.taseron_maliyet_senkronize_et(nakliye)
        db.session.commit()

        maliyet = HizmetKaydi.query.filter_by(nakliye_id=nakliye.id, yon="gelen").one()
        nakliye.nakliye_tipi = "oz_mal"
        nakliye.taseron_maliyet = Decimal("0.00")
        CariServis.taseron_maliyet_senkronize_et(nakliye)
        db.session.commit()

        db.session.refresh(maliyet)
        assert db.session.get(HizmetKaydi, maliyet.id) is not None
        assert maliyet.is_deleted is True


def test_deleted_nakliye_cari_sync_never_reactivates_customer_or_cost(app):
    with app.app_context():
        musteri = _firma("Cari Koruma")
        taseron = _firma("Cari Koruma Taseron", is_musteri=False)
        nakliye = _nakliye(musteri)
        CariServis.musteri_nakliye_senkronize_et(nakliye)
        nakliye.nakliye_tipi = "taseron"
        nakliye.taseron_firma_id = taseron.id
        nakliye.taseron_maliyet = Decimal("50.00")
        CariServis.taseron_maliyet_senkronize_et(nakliye)
        db.session.commit()

        musteri_hizmet = HizmetKaydi.query.filter_by(
            nakliye_id=nakliye.id, yon="giden"
        ).one()
        maliyet = HizmetKaydi.query.filter_by(
            nakliye_id=nakliye.id, yon="gelen"
        ).one()
        nakliye.is_active = False
        nakliye.is_deleted = True
        CariServis.musteri_nakliye_senkronize_et(nakliye)
        CariServis.taseron_maliyet_senkronize_et(nakliye)
        db.session.commit()

        assert musteri_hizmet.is_deleted is True
        assert maliyet.is_deleted is True


def test_guncelle_cari_toplam_ignores_deleted_nakliye(app):
    with app.app_context():
        firma = _firma("Cari Toplam")
        kiralama = Kiralama(
            kiralama_form_no=f"PF-CARI-{uuid.uuid4().hex[:6]}",
            firma_musteri_id=firma.id,
            kdv_orani=20,
        )
        db.session.add(kiralama)
        db.session.flush()
        nakliye = _nakliye(firma, kiralama_id=kiralama.id)
        CariServis.musteri_nakliye_senkronize_et(nakliye)
        db.session.commit()
        hizmet = HizmetKaydi.query.filter_by(nakliye_id=nakliye.id, yon="giden").one()

        nakliye.is_active = False
        nakliye.is_deleted = True
        hizmet.is_deleted = True
        hizmet.is_active = False
        db.session.commit()

        KiralamaService.guncelle_cari_toplam(kiralama.id, auto_commit=False)
        db.session.commit()
        assert hizmet.is_deleted is True
        assert hizmet.is_active is False


def test_standalone_delete_requires_password_and_supports_undo(app, client):
    with app.app_context():
        user = User(username=f"nak_user_{uuid.uuid4().hex[:6]}", rol="admin")
        user.set_password("Secret123!")
        db.session.add(user)

        firma = _firma()
        nakliye = _nakliye(firma)
        CariServis.musteri_nakliye_senkronize_et(nakliye)
        db.session.commit()
        nakliye_id = nakliye.id
        user_id = user.id

    _login_user(client, user_id)
    json_headers = {"X-Requested-With": "XMLHttpRequest"}

    resp = client.post(
        f"/nakliyeler/sil/{nakliye_id}",
        data={"delete_confirm_password": "wrong", "ajax": "1"},
        headers=json_headers,
    )
    assert resp.status_code == 400

    with app.app_context():
        assert db.session.get(Nakliye, nakliye_id).is_deleted is False

    resp = client.post(
        f"/nakliyeler/sil/{nakliye_id}",
        data={"delete_confirm_password": "Secret123!", "ajax": "1"},
        headers=json_headers,
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["ok"] is True
    assert payload["undo_seconds"] == 60

    with app.app_context():
        n = db.session.get(Nakliye, nakliye_id)
        assert n.is_deleted is True
        assert n.is_active is False
        hizmet = HizmetKaydi.query.filter_by(nakliye_id=nakliye_id, yon="giden").one()
        assert hizmet.is_deleted is True

    resp = client.post(
        f"/nakliyeler/geri-al/{nakliye_id}",
        headers=json_headers,
    )
    assert resp.status_code == 200
    assert resp.get_json()["ok"] is True

    with app.app_context():
        n = db.session.get(Nakliye, nakliye_id)
        assert n.is_deleted is False
        assert n.is_active is True
        hizmet = HizmetKaydi.query.filter_by(nakliye_id=nakliye_id, yon="giden").one()
        assert hizmet.is_deleted is False


def test_kiralama_delete_marks_nakliye_soft_deleted_and_keeps_prior_passive(app):
    with app.app_context():
        musteri = _firma("Kir Soft")
        kiralama = Kiralama(
            kiralama_form_no=f"PF-SOFT-{uuid.uuid4().hex[:6]}",
            firma_musteri_id=musteri.id,
            kdv_orani=20,
        )
        db.session.add(kiralama)
        db.session.flush()

        aktif_n = _nakliye(musteri, kiralama_id=kiralama.id, plaka="34KIRAKT")
        pasif_n = _nakliye(
            musteri,
            kiralama_id=kiralama.id,
            plaka="34KIRPAS",
            is_active=False,
            tutar="0.00",
        )
        db.session.commit()

        snapshot = KiralamaService.delete_with_relations(kiralama.id)
        db.session.refresh(aktif_n)
        db.session.refresh(pasif_n)

        assert aktif_n.is_deleted is True
        assert aktif_n.is_active is False
        assert pasif_n.is_deleted is False
        assert pasif_n.is_active is False


def test_archive_restore_only_restores_children_deleted_with_parent(app):
    with app.app_context():
        firma = _firma("Archive Timestamp")
        kiralama = Kiralama(
            kiralama_form_no=f"PF-ARCH-{uuid.uuid4().hex[:6]}",
            firma_musteri_id=firma.id,
            kdv_orani=20,
        )
        db.session.add(kiralama)
        db.session.flush()
        active_child = _nakliye(firma, kiralama_id=kiralama.id, plaka="34ARCHNEW")
        old_child = _nakliye(
            firma,
            kiralama_id=kiralama.id,
            plaka="34ARCHOLD",
            is_active=False,
            is_deleted=True,
        )
        db.session.commit()

        KiralamaService.delete_with_relations(kiralama.id)
        parent_deleted = kiralama.deleted_at
        old_child.deleted_at = parent_deleted - timedelta(days=1)
        db.session.commit()
        KiralamaService.restore_archived_with_relations(kiralama.id)

        db.session.refresh(active_child)
        db.session.refresh(old_child)
        assert active_child.is_deleted is False
        assert active_child.is_active is True
        assert old_child.is_deleted is True
        assert old_child.is_active is False


def test_undo_snapshot_rejects_wrong_actor_and_expiry():
    snapshot = {
        "nakliye_id": 10,
        "actor_id": 7,
        "expires_at": (datetime.now(timezone.utc) - timedelta(seconds=1)).isoformat(),
    }
    with pytest.raises(Exception, match="yalnızca silme"):
        NakliyeService._validate_restore_snapshot(10, snapshot, actor_id=8)
    snapshot["actor_id"] = 7
    with pytest.raises(Exception, match="süresi doldu"):
        NakliyeService._validate_restore_snapshot(10, snapshot, actor_id=7)

        KiralamaService.restore_with_relations(kiralama.id, snapshot=snapshot)
        db.session.refresh(aktif_n)
        db.session.refresh(pasif_n)

        assert aktif_n.is_deleted is False
        assert aktif_n.is_active is True
        assert pasif_n.is_active is False


def test_donus_recreate_soft_deletes_old_and_avoids_duplicate_active(app):
    with app.app_context():
        musteri = _firma("Donus Soft")
        kiralama = Kiralama(
            kiralama_form_no=f"PF-DON-{uuid.uuid4().hex[:6]}",
            firma_musteri_id=musteri.id,
            kdv_orani=20,
        )
        db.session.add(kiralama)
        db.session.flush()
        kalem = KiralamaKalemi(
            kiralama_id=kiralama.id,
            kiralama_baslangici=date.today() - timedelta(days=5),
            kiralama_bitis=date.today(),
            kiralama_brm_fiyat=Decimal("100.00"),
            sonlandirildi=True,
            is_active=False,
            donus_nakliye_satis_fiyat=Decimal("80.00"),
        )
        db.session.add(kalem)
        db.session.flush()

        aciklama = f"Dönüş: {kiralama.kiralama_form_no} #{kalem.id}"
        eski = _nakliye(musteri, kiralama_id=kiralama.id, plaka="34DONUS1")
        eski.aciklama = aciklama
        db.session.commit()

        NakliyeService.soft_delete_matching(
            Nakliye.kiralama_id == kiralama.id,
            Nakliye.aciklama == aciklama,
            soft_delete_cari=False,
        )
        yeni = Nakliye(
            kiralama_id=kiralama.id,
            firma_id=musteri.id,
            tarih=date.today(),
            islem_tarihi=date.today(),
            guzergah="Yeni donus",
            tutar=Decimal("80.00"),
            toplam_tutar=Decimal("80.00"),
            aciklama=aciklama,
            is_active=True,
            is_deleted=False,
        )
        db.session.add(yeni)
        db.session.commit()

        db.session.refresh(eski)
        assert eski.is_deleted is True
        aktifler = Nakliye.active_query().filter_by(
            kiralama_id=kiralama.id, aciklama=aciklama
        ).all()
        assert len(aktifler) == 1
        assert aktifler[0].id == yeni.id
