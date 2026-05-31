"""audit_firma_bakiye_defter_mismatch.py read-only rapor testleri."""

from __future__ import annotations

import importlib.util
import sys
import uuid
from datetime import date
from decimal import Decimal
from pathlib import Path

from app.cari.models import HizmetKaydi
from app.extensions import db
from app.firmalar.models import Firma

ROOT = Path(__file__).resolve().parents[1]
_AUDIT_PATH = ROOT / "scripts" / "audit_firma_bakiye_defter_mismatch.py"
_spec = importlib.util.spec_from_file_location("audit_firma_bakiye_mod", _AUDIT_PATH)
_audit_mod = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
sys.modules["audit_firma_bakiye_mod"] = _audit_mod
_spec.loader.exec_module(_audit_mod)
run_audit = _audit_mod.run_audit
write_csv = _audit_mod.write_csv
write_xlsx = _audit_mod.write_xlsx


def _unique_vergi_no() -> str:
    return f"B{uuid.uuid4().hex[:9].upper()}"


def _firma(name: str, bakiye=Decimal("0")):
    return Firma(
        firma_adi=name,
        yetkili_adi="Yetkili",
        iletisim_bilgileri="Adres",
        vergi_dairesi="VD",
        vergi_no=_unique_vergi_no(),
        is_musteri=True,
        is_tedarikci=False,
        bakiye=bakiye,
    )


def test_firma_bakiye_mismatch_report_classifies_zero_cache(app):
    firma = _firma("AUDIT FIRMA ZERO CACHE", bakiye=Decimal("0"))
    db.session.add(firma)
    db.session.flush()

    db.session.add(
        HizmetKaydi(
            firma_id=firma.id,
            tarih=date(2026, 5, 1),
            islem_tarihi=date(2026, 5, 1),
            tutar=Decimal("1000.00"),
            yon="giden",
            aciklama="Test fatura",
            kdv_orani=20,
        )
    )
    db.session.commit()

    report = run_audit(today=date(2026, 5, 26), tolerance=Decimal("0.01"))

    row = next(r for r in report.firma_mismatches if r.firma_id == firma.id)
    assert row.firma_bakiye_cache == Decimal("0")
    assert row.defter_net_bakiye == Decimal("1000.00")
    assert "FIRMA_BAKIYE_ZERO_CACHE" in row.sebep_tahmini
    assert "CARI_CACHE_ZERO" in row.sebep_tahmini


def test_firma_bakiye_mismatch_csv_is_report_only(app, tmp_path):
    firma = _firma("AUDIT FIRMA CSV", bakiye=Decimal("9999.00"))
    db.session.add(firma)
    db.session.add(
        HizmetKaydi(
            firma=firma,
            tarih=date(2026, 5, 1),
            islem_tarihi=date(2026, 5, 1),
            tutar=Decimal("100.00"),
            yon="giden",
            aciklama="Test fatura",
            kdv_orani=0,
        )
    )
    db.session.commit()

    csv_path = tmp_path / "firma_bakiye.csv"
    report = run_audit(today=date(2026, 5, 26), tolerance=Decimal("0.01"))
    write_csv(report.firma_mismatches, str(csv_path))

    content = csv_path.read_text(encoding="utf-8-sig")
    assert "firma_id,firma_adi,firma_bakiye_cache,defter_net_bakiye" in content
    assert "AUDIT FIRMA CSV" in content

    db.session.refresh(firma)
    assert firma.bakiye == Decimal("9999.00")


def test_firma_bakiye_mismatch_xlsx_report_has_sheets(app, tmp_path):
    from openpyxl import load_workbook

    firma = _firma("AUDIT FIRMA XLSX", bakiye=Decimal("0"))
    db.session.add(firma)
    db.session.add(
        HizmetKaydi(
            firma=firma,
            tarih=date(2026, 5, 1),
            islem_tarihi=date(2026, 5, 1),
            tutar=Decimal("250.00"),
            yon="giden",
            aciklama="Test fatura",
            kdv_orani=20,
        )
    )
    db.session.commit()

    xlsx_path = tmp_path / "firma_bakiye.xlsx"
    report = run_audit(today=date(2026, 5, 26), tolerance=Decimal("0.01"))
    write_xlsx(report, str(xlsx_path))

    wb = load_workbook(xlsx_path, data_only=True)
    assert wb.sheetnames == ["Ozet", "Firma Uyusmazliklari", "Dis Kiralama"]
    ws = wb["Firma Uyusmazliklari"]
    assert ws["A1"].value == "firma_id"
    assert any(row[1] == "AUDIT FIRMA XLSX" for row in ws.iter_rows(min_row=2, values_only=True))

    db.session.refresh(firma)
    assert firma.bakiye == Decimal("0")
